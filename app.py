"""
Adchor SOW Builder -- Streamlit Web App
Team URL: deploy to Streamlit Community Cloud or sow.adchor.com

Flow:
  Step 1 -- Upload filled creative brief PDF + call transcript
  Step 2 -- Review & edit AI-generated SOW content
  Step 3 -- Build pricing (live auto-total, growing library)
  Step 4 -- Download PDF → send via Adobe Sign
"""
import streamlit as st
import json
import os
import sys
import base64
from pathlib import Path
from datetime import datetime

# ── Ensure app directory is always in sys.path ────────────────────────────────
_APP_DIR = os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else os.getcwd()
if _APP_DIR not in sys.path:
    sys.path.insert(0, _APP_DIR)

# ── Import all app modules upfront (no lazy imports -- prevents runpy issues) ──
from brief_extractor import extract_brief_fields, format_for_prompt
from sow_generator import generate_sow_content, get_empty_sow
from sow_pdf import build_sow_pdf

# ── Page Config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Adchor SOW Builder",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Adchor Brand CSS ──────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;500;600;700;800;900&display=swap');

/* ── Global ── */
html, body, [class*="css"],
.stMarkdown, .stTextInput, .stTextArea,
.stSelectbox, .stButton, .stNumberInput,
.stFileUploader, p, span, div, label {
    font-family: 'Montserrat', sans-serif !important;
}

/* Preserve Material Icons font */
[data-testid="stIconMaterial"] {
    font-family: 'Material Symbols Rounded', 'Material Icons', 'Material Icons Round' !important;
}

/* ── App background ── */
.stApp { background: #0a0c12 !important; }
[data-testid="stAppViewBlockContainer"] { background: #0a0c12; }

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: #000000 !important;
    border-right: 1px solid #0d0f1a !important;
}
[data-testid="stSidebar"] > div { padding-top: 56px !important; }

/* ── Sidebar collapse button — keep it above brand row ── */
[data-testid="stSidebarCollapseButton"] {
    position: absolute !important;
    top: 10px !important;
    right: 10px !important;
    z-index: 1000 !important;
    background: transparent !important;
    border: none !important;
}
[data-testid="stSidebarCollapseButton"] button {
    background: transparent !important;
    color: #1e2540 !important;
}
[data-testid="stSidebarCollapseButton"] button:hover {
    color: #014bf7 !important;
    background: rgba(1,75,247,0.08) !important;
}

/* ══════════════════════════════════════════
   SIDEBAR — Brand Row & Logo
══════════════════════════════════════════ */
.sb-brand-row {
    display: flex; align-items: center;
    margin-bottom: 22px; padding: 0 2px;
}
/* Logo visibility: dark mode shows dark logo, light mode shows light logo */
.sb-logo { width: 100%; height: auto; display: block; max-width: 220px; }
.sb-logo-light { display: none; }
@media (prefers-color-scheme: light) {
    .sb-logo-dark  { display: none; }
    .sb-logo-light { display: block; }
}
html.light .sb-logo-dark  { display: none !important; }
html.light .sb-logo-light { display: block !important; }
html.dark  .sb-logo-dark  { display: block !important; }
html.dark  .sb-logo-light { display: none !important; }
.sb-brand-mark {
    width: 40px; height: 40px; border-radius: 12px;
    background: linear-gradient(135deg, #014bf7, #021de0);
    display: flex; align-items: center; justify-content: center;
    font-size: 17px; font-weight: 900; color: white;
    box-shadow: 0 4px 20px rgba(1,75,247,0.6), 0 0 40px rgba(1,75,247,0.2);
    flex-shrink: 0;
}
.sb-bname { font-size: 13px; font-weight: 800; color: #fff; letter-spacing: 3px; text-transform: uppercase; }
.sb-bsub  { font-size: 8px; font-weight: 600; color: #14a4fe; letter-spacing: 2px; text-transform: uppercase; margin-top: 2px; }

/* ── Progress bar ── */
.sb-prog-wrap { margin-bottom: 18px; padding: 0 2px; }
.sb-prog-label {
    font-size: 10px; font-weight: 700; color: #1e2540; letter-spacing: 2px;
    text-transform: uppercase; margin-bottom: 7px;
    display: flex; justify-content: space-between;
}
.sb-prog-label span { color: #014bf7; }
.sb-prog-track { height: 3px; background: #0d0f1a; border-radius: 99px; }
.sb-prog-fill  { height: 3px; background: linear-gradient(90deg, #014bf7, #14a4fe); border-radius: 99px; box-shadow: 0 0 8px rgba(20,164,254,0.6); }

/* ── Step items ── */
.sb-step-item {
    display: flex; align-items: center; gap: 12px;
    padding: 10px 10px; border-radius: 11px; margin-bottom: 4px;
}
.sb-step-icon {
    width: 36px; height: 36px; border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 15px; font-weight: 900; flex-shrink: 0;
}
.sb-step-name { font-size: 13px; font-weight: 700; }
.sb-step-desc { font-size: 11px; font-weight: 500; margin-top: 2px; }

.step-done .sb-step-icon { background: rgba(0,255,121,0.1); color: #00ff79; }
.step-done .sb-step-name { color: #00ff79; }
.step-done .sb-step-desc { color: rgba(0,255,121,0.5); }

.step-active { background: rgba(1,75,247,0.1); }
.step-active .sb-step-icon { background: #014bf7; color: white; box-shadow: 0 4px 14px rgba(1,75,247,0.55); }
.step-active .sb-step-name { color: #ffffff; }
.step-active .sb-step-desc { color: rgba(255,255,255,0.45); }

.step-pend .sb-step-icon { background: rgba(255,255,255,0.03); color: #1e2540; border: 1px solid #0d0f1a; }
.step-pend .sb-step-name { color: #1e2540; }
.step-pend .sb-step-desc { color: #141828; }

/* ── Active Client chip ── */
.sb-client-chip {
    margin-top: 20px;
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.06);
    border-radius: 12px; padding: 12px 14px;
}
.sb-chip-label {
    font-size: 8px; font-weight: 700; color: #1e2540;
    letter-spacing: 2px; text-transform: uppercase; margin-bottom: 5px;
}
.sb-chip-row { display: flex; align-items: center; gap: 9px; }
.sb-chip-dot {
    width: 8px; height: 8px; border-radius: 50%;
    background: #00ff79; box-shadow: 0 0 6px rgba(0,255,121,0.7); flex-shrink: 0;
}
.sb-chip-name { font-size: 13px; font-weight: 800; color: #9aa0b4; }

/* ══════════════════════════════════════════
   HERO BAR — Step header
══════════════════════════════════════════ */
.hero-bar {
    background: linear-gradient(135deg, #010c30 0%, #014bf7 60%, #021de0 100%);
    padding: 20px 26px;
    border-radius: 12px;
    position: relative; overflow: hidden;
    margin: 0 0 20px;
    box-shadow: 0 6px 28px rgba(1,75,247,0.4);
}
.hero-bar::before {
    content: '';
    position: absolute; top: -40px; right: -40px;
    width: 140px; height: 140px; border-radius: 50%;
    background: rgba(20,164,254,0.12);
}
.hero-bar::after {
    content: '';
    position: absolute; bottom: -25px; right: 110px;
    width: 80px; height: 80px; border-radius: 50%;
    background: rgba(0,255,121,0.07);
}
.hero-eyebrow {
    font-size: 9px; font-weight: 700; letter-spacing: 3px;
    text-transform: uppercase; color: rgba(255,255,255,0.45); margin-bottom: 5px;
}
.hero-title {
    font-size: 20px; font-weight: 900; color: white;
    letter-spacing: -0.5px; line-height: 1.1;
}
.hero-sub {
    font-size: 11px; font-weight: 500; color: rgba(255,255,255,0.5); margin-top: 5px;
}

/* ── Section header bar (sub-section within a step) ── */
.sec-bar {
    background: linear-gradient(135deg, #030a20 0%, #014bf7 100%);
    color: white;
    padding: 13px 22px;
    border-radius: 10px;
    font-weight: 800;
    font-size: 11px;
    letter-spacing: 2.5px;
    text-transform: uppercase;
    margin: 0 0 18px;
    box-shadow: 0 4px 20px rgba(1,75,247,0.3);
    position: relative; overflow: hidden;
}
.sec-bar::before {
    content: '';
    position: absolute; top: -20px; right: -20px;
    width: 80px; height: 80px; border-radius: 50%;
    background: rgba(20,164,254,0.1);
}

/* ── Scope block header ── */
.scope-header {
    background: #014bf7;
    color: white;
    padding: 11px 18px;
    border-radius: 8px 8px 0 0;
    font-weight: 700;
    font-size: 11px;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    margin-top: 14px;
    box-shadow: 0 4px 12px rgba(1,75,247,0.25);
}

/* ── Pricing total / Investment bar ── */
.pricing-total {
    background: linear-gradient(135deg, #021de0 0%, #014bf7 100%);
    color: white;
    padding: 18px 24px;
    border-radius: 14px;
    text-align: right;
    font-size: 22px;
    font-weight: 800;
    margin-top: 14px;
    box-shadow: 0 8px 32px rgba(1,75,247,0.4);
    border: 1px solid rgba(20,164,254,0.25);
    letter-spacing: -0.5px;
    position: relative; overflow: hidden;
}
.pricing-total::before {
    content: '';
    position: absolute; top: -25px; right: -25px;
    width: 100px; height: 100px; border-radius: 50%;
    background: rgba(20,164,254,0.12);
}
.pricing-total .sub { font-size: 12px; opacity: 0.65; font-weight: 500; letter-spacing: 0; }

/* ── Info box ── */
.info-box {
    background: #0f111a;
    border: 1px solid #1e2235;
    border-left: 4px solid #014bf7;
    border-radius: 10px;
    padding: 14px 18px;
    margin: 8px 0 14px;
    font-size: 13px;
    color: #c8ccd8;
}

/* ── Download CTA ── */
.dl-note {
    background: rgba(0,255,121,0.05);
    border: 1px solid rgba(0,255,121,0.2);
    border-radius: 10px;
    padding: 16px 20px;
    margin-top: 14px;
    font-size: 13px;
    color: #c8ccd8;
}

/* ── AI instruction panel (Step 2 global editor) ── */
.ai-instr-panel {
    background: linear-gradient(135deg, #090d1c 0%, #070b17 100%);
    border: 1px solid #1a2545;
    border-left: 4px solid #14a4fe;
    border-radius: 12px;
    padding: 14px 20px 12px;
    margin: 8px 0 4px;
}
.ai-instr-label {
    font-size: 10px;
    font-weight: 800;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: #14a4fe;
    margin-bottom: 3px;
}
.ai-instr-sub {
    font-size: 11px;
    color: #5a6278;
    line-height: 1.5;
}

/* ── AI reword box ── */
.ai-box {
    background: #0d0f1a;
    border: 1px solid #1e2540;
    border-left: 4px solid #14a4fe;
    border-radius: 10px;
    padding: 14px 18px;
    margin: 6px 0 10px;
    font-size: 12px;
    color: #9aa0b4;
    white-space: pre-wrap;
    word-break: break-word;
    line-height: 1.6;
}

/* ── Input fields ── */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea,
.stNumberInput > div > div > input {
    background: #0f111a !important;
    border: 1px solid #1e2235 !important;
    border-radius: 8px !important;
    color: #e8eaf0 !important;
    font-family: 'Montserrat', sans-serif !important;
    font-size: 13px !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: #014bf7 !important;
    box-shadow: 0 0 0 2px rgba(1,75,247,0.15) !important;
}

/* ── Buttons ── */
.stButton > button {
    font-family: 'Montserrat', sans-serif !important;
    font-weight: 700 !important;
    letter-spacing: 0.5px !important;
    border-radius: 8px !important;
    transition: all 0.2s !important;
}
.stButton > button[kind="primary"] {
    background: #014bf7 !important;
    border: none !important;
    color: #ffffff !important;
    box-shadow: 0 4px 16px rgba(1,75,247,0.4) !important;
}
.stButton > button[kind="primary"] p,
.stButton > button[kind="primary"] span,
.stButton > button[kind="primary"] div {
    color: #ffffff !important;
}
.stButton > button[kind="primary"]:hover {
    background: #0240d4 !important;
    box-shadow: 0 6px 22px rgba(1,75,247,0.6) !important;
    transform: translateY(-1px) !important;
}
.stButton > button[kind="secondary"] {
    background: transparent !important;
    border: 1px solid #1e2235 !important;
    color: #9aa0b4 !important;
}
.stButton > button[kind="secondary"]:hover {
    border-color: #014bf7 !important;
    color: white !important;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    background: #0f111a !important;
    border: 1.5px dashed #1e2235 !important;
    border-radius: 10px !important;
}
[data-testid="stFileUploader"]:hover {
    border-color: #014bf7 !important;
}

/* ── Dividers ── */
hr { border-color: #1a1d2e !important; margin: 16px 0 !important; }

/* ── Success/warning/error ── */
.stSuccess { background: rgba(0,255,121,0.08) !important; border-color: rgba(0,255,121,0.3) !important; color: #00ff79 !important; }
.stAlert { border-radius: 8px !important; }

/* ── Expander ── */
[data-testid="stExpander"] {
    background: #0f111a !important;
    border: 1px solid #1e2235 !important;
    border-radius: 10px !important;
}

/* ── Selectbox ── */
[data-testid="stSelectbox"] > div > div {
    background: #0f111a !important;
    border: 1px solid #1e2235 !important;
    border-radius: 8px !important;
    color: #e8eaf0 !important;
}

/* ── Caption/small text ── */
.stCaption { color: #5a6278 !important; font-size: 12px !important; }

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #0a0c12; }
::-webkit-scrollbar-thumb { background: #1e2235; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #014bf7; }

/* ══════════════════════════════════════════
   LIGHT MODE
══════════════════════════════════════════ */
@media (prefers-color-scheme: light) {
    .stApp { background: #eef1fb !important; }
    [data-testid="stAppViewBlockContainer"] { background: #eef1fb !important; }
    [data-testid="stSidebar"] {
        background: #ffffff !important;
        border-right: 1px solid #e4e8f4 !important;
        box-shadow: 2px 0 20px rgba(1,75,247,0.06) !important;
    }
    /* Sidebar typography */
    .sb-bname { color: #0d1025; }
    .sb-bsub  { color: #014bf7; }
    .sb-prog-label { color: #c8d0e4; }
    .sb-prog-track { background: #eef1f8; }
    .sb-prog-fill  { box-shadow: none; }
    /* Step states */
    .step-done .sb-step-icon { background: rgba(0,168,84,0.1); color: #00a854; }
    .step-done .sb-step-name { color: #00a854; }
    .step-done .sb-step-desc { color: rgba(0,168,84,0.6); }
    .step-active { background: #f0f4ff; }
    .step-active .sb-step-icon { background: #014bf7; box-shadow: 0 4px 14px rgba(1,75,247,0.35); }
    .step-active .sb-step-name { color: #014bf7; }
    .step-active .sb-step-desc { color: #8090b8; }
    .step-pend .sb-step-icon { background: #f4f6fc; color: #c8d0e4; border-color: #eef1f8; }
    .step-pend .sb-step-name { color: #c8d0e4; }
    .step-pend .sb-step-desc { color: #d8dff0; }
    /* Client chip */
    .sb-client-chip { background: #f4f7ff; border-color: #e4e8f4; }
    .sb-chip-label  { color: #c0c8e0; }
    .sb-chip-dot    { background: #00a854; box-shadow: 0 0 5px rgba(0,168,84,0.5); }
    .sb-chip-name   { color: #3a4060; }
    /* Primary button text always white */
    .stButton > button[kind="primary"],
    .stButton > button[kind="primary"] p,
    .stButton > button[kind="primary"] span { color: #ffffff !important; }
    /* Inputs */
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea,
    .stNumberInput > div > div > input {
        background: #ffffff !important;
        border-color: #e4e8f4 !important;
        color: #1a1d2e !important;
    }
    /* Info / AI boxes */
    .info-box { background: #f0f4ff; border-color: #c8d4f8; color: #4a5280; }
    .ai-instr-panel { background: #f0f4ff; border-color: #c8d4f8; }
    .ai-instr-sub   { color: #6070a0; }
    .ai-box { background: #f4f6fc; border-color: #c8d4f8; color: #5a6080; }
    /* Expander */
    [data-testid="stExpander"] { background: #ffffff !important; border-color: #e4e8f4 !important; }
    /* Selectbox */
    [data-testid="stSelectbox"] > div > div {
        background: #ffffff !important;
        border-color: #e4e8f4 !important;
        color: #1a1d2e !important;
    }
    /* File uploader */
    [data-testid="stFileUploader"] { background: #ffffff !important; border-color: #c8d4f8 !important; }
    /* Scrollbar */
    ::-webkit-scrollbar-track { background: #eef1fb; }
    ::-webkit-scrollbar-thumb { background: #c8d0e4; }
    /* Dividers */
    hr { border-color: #e4e8f4 !important; }
    /* Caption */
    .stCaption { color: #8090b4 !important; }
}

/* ══════════════════════════════════════════
   LIGHT MODE  (Streamlit theme toggle)
   Streamlit adds html.light / html.dark when
   the user changes theme in the settings menu.
══════════════════════════════════════════ */
html.light .stApp { background: #eef1fb !important; }
html.light [data-testid="stAppViewBlockContainer"] { background: #eef1fb !important; }
html.light [data-testid="stSidebar"] {
    background: #ffffff !important;
    border-right: 1px solid #e4e8f4 !important;
    box-shadow: 2px 0 20px rgba(1,75,247,0.06) !important;
}
html.light .sb-bname { color: #0d1025; }
html.light .sb-bsub  { color: #014bf7; }
html.light .sb-prog-label { color: #c8d0e4; }
html.light .sb-prog-track { background: #eef1f8; }
html.light .sb-prog-fill  { box-shadow: none; }
html.light .step-done .sb-step-icon { background: rgba(0,168,84,0.1); color: #00a854; }
html.light .step-done .sb-step-name { color: #00a854; }
html.light .step-done .sb-step-desc { color: rgba(0,168,84,0.6); }
html.light .step-active { background: #f0f4ff; }
html.light .step-active .sb-step-icon { background: #014bf7; box-shadow: 0 4px 14px rgba(1,75,247,0.35); }
html.light .step-active .sb-step-name { color: #014bf7; }
html.light .step-active .sb-step-desc { color: #8090b8; }
html.light .step-pend .sb-step-icon { background: #f4f6fc; color: #c8d0e4; border-color: #eef1f8; }
html.light .step-pend .sb-step-name { color: #c8d0e4; }
html.light .step-pend .sb-step-desc { color: #d8dff0; }
html.light .sb-client-chip { background: #f4f7ff; border-color: #e4e8f4; }
html.light .sb-chip-label  { color: #c0c8e0; }
html.light .sb-chip-dot    { background: #00a854; box-shadow: 0 0 5px rgba(0,168,84,0.5); }
html.light .sb-chip-name   { color: #3a4060; }
html.light .stButton > button[kind="primary"],
html.light .stButton > button[kind="primary"] p,
html.light .stButton > button[kind="primary"] span { color: #ffffff !important; }
html.light .stButton > button[kind="secondary"] {
    background: #ffffff !important;
    border-color: #c8d0e4 !important;
    color: #4a5280 !important;
}
html.light .stTextInput > div > div > input,
html.light .stTextArea > div > div > textarea,
html.light .stNumberInput > div > div > input {
    background: #ffffff !important;
    border-color: #e4e8f4 !important;
    color: #1a1d2e !important;
}
html.light .info-box { background: #f0f4ff; border-color: #c8d4f8; color: #4a5280; }
html.light .ai-instr-panel { background: #f0f4ff; border-color: #c8d4f8; }
html.light .ai-instr-sub   { color: #6070a0; }
html.light .ai-box { background: #f4f6fc; border-color: #c8d4f8; color: #5a6080; }
html.light .dl-note { background: #f0fff4; border-color: rgba(0,180,80,0.3); color: #3a4060; }
html.light [data-testid="stExpander"] { background: #ffffff !important; border-color: #e4e8f4 !important; }
html.light [data-testid="stSelectbox"] > div > div {
    background: #ffffff !important;
    border-color: #e4e8f4 !important;
    color: #1a1d2e !important;
}
html.light [data-testid="stFileUploader"] { background: #ffffff !important; border-color: #c8d4f8 !important; }
html.light hr { border-color: #e4e8f4 !important; }
html.light .stCaption { color: #8090b4 !important; }
</style>
""", unsafe_allow_html=True)

# ── Base directory (works with both direct run and runpy) ─────────────────────
_BASE_DIR = Path(__file__).parent if "__file__" in dir() else Path(os.getcwd())

# ── Pricing Library (load once per session) ───────────────────────────────────
LIBRARY_PATH     = _BASE_DIR / "pricing_library.json"
SOW_LIBRARY_PATH = _BASE_DIR / "sow_library.json"

def load_library():
    if LIBRARY_PATH.exists():
        with open(LIBRARY_PATH) as f:
            return json.load(f)
    return {"items": []}

def save_library(library: dict):
    """Persist pricing library to disk immediately."""
    with open(LIBRARY_PATH, "w") as f:
        json.dump(library, f, indent=2)

def load_sow_library() -> dict:
    """Load saved SOW history from disk."""
    if SOW_LIBRARY_PATH.exists():
        with open(SOW_LIBRARY_PATH) as f:
            return json.load(f)
    return {"sows": []}

def save_sow_library(library: dict):
    """Persist SOW library to disk immediately."""
    with open(SOW_LIBRARY_PATH, "w") as f:
        json.dump(library, f, indent=2)

def autosave_pricing_items_to_library(pricing_items: list):
    """
    Automatically add any quoted line items into the pricing library
    so they can be reused in future SOWs. Skips items with no name
    or that are already in the library (matched by name).
    """
    lib = st.session_state.get("pricing_library", load_library())
    existing_names = {it["name"].strip().lower() for it in lib.get("items", []) if it.get("name")}
    changed = False
    for item in pricing_items:
        name = (item.get("name") or "").strip()
        if not name or name.lower() in existing_names:
            continue
        lib["items"].append({
            "name":        name,
            "description": item.get("description", ""),
            "category":    item.get("category", "Quoted Service"),
            "unit_price":  item.get("unit_price", 0),
        })
        existing_names.add(name.lower())
        changed = True
    if changed:
        st.session_state.pricing_library = lib
        save_library(lib)

def upsert_sow_to_library(sow_data: dict, pricing_items: list, sow_discount: float):
    """Save or update the current SOW in the library (matched by client+project)."""
    import uuid as _uuid
    lib = st.session_state.get("sow_library", load_sow_library())
    sows = lib.get("sows", [])
    client  = (sow_data or {}).get("client_name", "").strip() or "Unknown Client"
    project = (sow_data or {}).get("project_name", "").strip() or "Untitled Project"
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
    # Match by client + project name to allow updates
    existing = next(
        (s for s in sows if s.get("client_name") == client and s.get("project_name") == project),
        None,
    )
    if existing:
        existing["sow_data"]      = sow_data
        existing["pricing_items"] = pricing_items
        existing["sow_discount"]  = sow_discount
        existing["updated_at"]    = now
        existing["status"]        = "draft"
    else:
        sows.insert(0, {
            "id":            str(_uuid.uuid4()),
            "client_name":  client,
            "project_name": project,
            "created_at":   now,
            "updated_at":   now,
            "status":       "draft",
            "sow_data":      sow_data,
            "pricing_items": pricing_items,
            "sow_discount":  sow_discount,
        })
    lib["sows"] = sows
    st.session_state.sow_library = lib
    save_sow_library(lib)


# ════════════════════════════════════════════════════════════════════════════════
# NEW: AI Reword Helper
# ════════════════════════════════════════════════════════════════════════════════
def _ai_reword(text: str, instruction: str, api_key: str) -> str:
    """Call Claude to reword/improve a piece of SOW text."""
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[{
            "role": "user",
            "content": (
                "You are a professional copywriter for a creative agency. "
                "You write polished, client-facing statements of work.\n\n"
                f"Task: {instruction}\n\n"
                f"Text to improve:\n{text}\n\n"
                "Return only the improved text — no preamble, no explanation, "
                "no markdown formatting, no code fences."
            ),
        }],
    )
    return msg.content[0].text.strip()


# ════════════════════════════════════════════════════════════════════════════════
# NEW: AI SOW Update — whole-document revision from a free-form instruction
# ════════════════════════════════════════════════════════════════════════════════
def _ai_update_sow(sow_data: dict, instruction: str, api_key: str) -> dict:
    """
    Call Claude to apply a free-form instruction to the entire SOW.
    Returns the complete updated sow_data dict (all fields, even unchanged ones).
    """
    import anthropic as _ant
    import json as _json

    _client = _ant.Anthropic(api_key=api_key)
    _sow_json = _json.dumps(sow_data, indent=2, ensure_ascii=False)

    _msg = _client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        messages=[{
            "role": "user",
            "content": (
                "You are an expert editor of Statement of Work (SOW) documents "
                "for a creative advertising agency called Adchor.\n\n"
                "Here is the current SOW data in JSON:\n"
                f"{_sow_json}\n\n"
                f"User instruction: {instruction}\n\n"
                "Apply the instruction and return the COMPLETE updated SOW as a single "
                "valid JSON object. Rules:\n"
                "1. Include EVERY field — even ones you did not change.\n"
                "2. Preserve existing values unless the instruction explicitly changes them.\n"
                "3. For list fields (assumptions, out_of_scope, scope_sections[].services, "
                "scope_sections[].deliverables), always return JSON arrays of strings.\n"
                "4. Keep the same JSON structure and key names as the input.\n"
                "5. Return raw JSON only — no markdown, no code fences, no commentary."
            ),
        }],
    )

    _text = _msg.content[0].text.strip()
    # Strip markdown code fences if the model wrapped in them anyway
    if _text.startswith("```"):
        _lines = _text.splitlines()
        _inner = _lines[1:-1] if _lines and _lines[-1].strip() == "```" else _lines[1:]
        _text = "\n".join(_inner).strip()

    return _json.loads(_text)


# ════════════════════════════════════════════════════════════════════════════════
# NEW: Widget keys that must be cleared when resetting a SOW
# ════════════════════════════════════════════════════════════════════════════════
_SOW_FIELD_WIDGET_KEYS = [
    "ta_why_now", "ta_project_overview", "ta_core_message",
    "ta_assumptions", "ta_out_of_scope", "ta_timeline_notes",
]
_AI_WIDGET_KEYS = [
    "ai_field_select", "ai_preset", "ai_custom_instr",
    "ai_input_text", "ai_result_area",
    "ai_sow_instruction",
]


# ════════════════════════════════════════════════════════════════════════════════
# NEW: Centralised reset helper (used by sidebar + Done button)
# ════════════════════════════════════════════════════════════════════════════════
def _reset_sow_state():
    """Full SOW workflow reset — preserves pricing_library and api_key."""
    reset_vals = {
        "step": 1,
        "brief_fields": {},
        "transcript": "",
        "sow_data": None,
        "pricing_items": [],
        "sow_discount": 0,
        "sow_total": 0,
        # new keys
        "pdf_downloaded": False,
        "ai_reword_result": "",
        "ai_reword_target": "",
        "ai_reword_widget_key": None,
        "ai_reword_field_type": "",
        "ai_input_text": "",
        "ai_sow_instruction": "",
        "ai_sow_update_status": "",
    }
    for k, v in reset_vals.items():
        st.session_state[k] = v
    # Clear SOW field widget state so they re-initialise with fresh sow_data
    for k in _SOW_FIELD_WIDGET_KEYS + _AI_WIDGET_KEYS:
        st.session_state.pop(k, None)
    # Clear scope-section widget keys (support up to 30 sections)
    for i in range(30):
        for k in [f"st_{i}", f"sd_{i}", f"ss_{i}", f"del_{i}"]:
            st.session_state.pop(k, None)


# ── Session State Init ────────────────────────────────────────────────────────
defaults = {
    "step": 1,
    "brief_fields": {},
    "transcript": "",
    "sow_data": None,
    "pricing_items": [],
    "sow_discount": 0,
    "sow_total": 0,
    "pricing_library": load_library(),
    "sow_library":     load_sow_library(),
    # ── NEW keys ──────────────────────────────────────────────────────────────
    "pdf_downloaded": False,      # True only after the download button is clicked
    "ai_reword_result": "",       # Stores Claude's latest reword suggestion
    "ai_reword_target": "",       # sow field key that was targeted
    "ai_reword_widget_key": None, # widget key for the targeted field
    "ai_reword_field_type": "",   # "text" | "list"
    "ai_input_text": "",          # content in the AI reword input box
    "ai_sow_instruction": "",     # global SOW AI editor instruction box
    "ai_sow_update_status": "",   # success/error banner after AI SOW update
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Pending widget updates (applied before any widget renders) ────────────────
# When "Apply to field" fires, the target widget has already been rendered in
# that pass, so we can't set its session-state key directly.  Instead we store
# a deferred update here and apply it at the top of the NEXT rerun, before the
# widget is instantiated.
if "_pending_widget_update" in st.session_state:
    _pwu = st.session_state.pop("_pending_widget_update")
    if isinstance(_pwu, dict):
        for _k, _v in _pwu.items():
            st.session_state[_k] = _v

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    # ── Brand row — real logos (dark/light) embedded as base64 ────────────────
    _LOGO_DARK  = (
    "iVBORw0KGgoAAAANSUhEUgAACMoAAAjKCAYAAABtSXCKAAAACXBIWXMAABcRAAAXEQHKJvM/AAAgAElEQVR4nOzaMQEAIAzAMMC/580HTSz07Z2ZAwAAAAAAAAAAv3sKAwAAAAAAAABQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAACw7Nw/apRBHIBhlv02uyGBFIoa8QiCteAN0til1CNYegRLj6GdR7CzFmzsJX/EwsKYzX6BlRQiCIqiEeF9nmaYaob5tS9DglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAGGyNCQAACAASURBVAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEgZjBgD4tw4Oj/cnk8nubDa7f3HwMExvDcNw5eslNhfznel0Ovn+UqvVuFqN4+ev++Xy7PXFen5+/mq9Xh/v3rj21CgBAAAAAAB+bLJerz0PAMAleHdwdG+YTveGYbi7WMzvbMxm2xsbs0sNlVer8Xx5dvZhtRrfjuP4QjwDAAAAAADwjVAGAOAvuQhjZsPwcHNzsbeYz69edhTzqz6dfP54erp8OY7js5u715+bNwAAAAAAUCWUAQD4A4dH7x/N5xsPtre2bv8vYczPXPw48+nk5M1yefZENAMAAAAAANQIZQAAftPB4fH+YjF/vPOFvbuJjetM7wV/1CXRFEmRRVmiSIkt0uPOzSS4d8wGF0OA7haDWcxqIPWqMyuxb7I3Gxhg7gVmYHnVyMo07iYLBy2tZoIsWkI2dxZBU4kJcHEJS+mgk3Tc12Kb4pcokZRFyqZU4ODIp9plWiKLZNX5/P0Agv6QWKfe9+WpqvP+z/N0nnqrVCody+r4PX36xdMnm1v/35dffjnRf6FvLgWHBAAAAAAA0FSCMgAAdZi/vzhw4sSJ/7u73HU1C5VjDmp94/Fvn2xujfef7/0oW0cOAAAAAABQP0EZAIA9zC8svd12svW/ZL16TL0ef/5kaXNz6y/6ensms3HEAAAAAAAA9ROUAQB4iTAg09Hedr3c1flmEcfnyebW+vrG4/9NhRkAAAAAACBPBGUAAGoUPSCz24uWTE82/5f+C31z6ToyAAAAAACAgxOUAQCIrD5cm+oud/6wCC2WDqJSqew8Wtu4dfbM6R9l56gBAAAAAAC+7TvGBAAousWllYnt7WfPzrzefUlI5tvCMTl75vSVp0+/2FpYXP5x2o4PAAAAAACgXirKAACFNX9/caCjo/3vtFk6mNWHa7fPvN49lqVjBgAAAAAACFSUAQCKKqwi03P2zCdCMgcXVt55srm1Nr+w9HbWjh0AAAAAACg2FWUAgMJZfbg2FYY9zPzRVCqVnQerj/6i99zZ/5zl5wEAAAAAABSHoAwAUBhhq6VyuetOR3tb2aw3jlZMAAAAAABAVmi9BAAUwsLi8o/DVktCMo0XVud5/PmTxTCIlLfnBgAAAAAA5IuKMgBA7i0tP/jZ2TOn/89SqXTMbDfP06dfPH34aO2P+i/0zeX1OQIAAAAAANkmKAMA5NryyuqH53rO/JlZjkelUtlZXln938/3nfvrIjxfAAAAAAAgW7ReAgByS0gmfmHVnnM9Z/6fsNVV0Z47AAAAAACQfoIyAEAuCckkR1gGAAAAAABIK0EZACB3hGSSJywDAAAAAACkkaAMAJArQjLpISwDAAAAAACkjaAMAJAbQjLpIywDAAAAAACkiaAMAJALYRDjzOvd/9Fspo+wDAAAAAAAkBbHdnZ2TAYAkGnz9xcH+np7Pg0DGWYyvZ4+/eLpw0drf9R/oW+u6GMBAAAAAAAkQ1AGAMi8p0+/2Dp5svWkmUy/J5tb6x3tbd1FHwcAAAAAACAZWi8BAJm2vvH4EyGZ7Ohobys/Wlv/uOjjAAAAAAAAJENQBgDIrKXlBz8rd3W+aQaz5XR3eSicu6KPAwAAAAAAED+tlwCATJpfWHq779zZvy+VSsfMYPZUKpWdxaWVN/ov9M0VfSwAAAAAAID4CMoAAJn0ZHNrLWzjY/ay68nm1npHe1t30ccBAAAAAACIj9ZLAEDmPFh99AshmewL5zCcy6KPAwAAAAAAEB9BGQAgU8KWS6e7uy6btXwI5zKc06KPAwAAAAAAEA+tlwCATNFyKX8ef/5kqfNUR1/RxwEAAAAAAGg+FWUAgMxYWn7wMyGZ/Ok81dEbzm3RxwEAAAAAAGg+FWUAgEyYv7840HP2zCctLSeOm7H82d5+9nzlwer3+i/0zRV9LAAAAAAAgOZRUQYAyITW1tYbQjL5Fc5tOMdFHwcAAAAAAKC5BGUAgNSbX1h6u7vc+UMzlW9nXu++FM510ccBAAAAAABoHkEZACD1OtrbrpdKpWNmKv/KXZ1/W/QxAAAAAAAAmkdQBgBItYXF5R+XuzrfNEvF0NHeVl5afvCzoo8DAAAAAADQHMd2dnYMLQCQWk82t9bC8IQZKo7t7WfPW1pOnCj6OAAAAAAAAI2nogwAkFqLSysTQjLF09Jy4vjqw7Wpoo8DAAAAAADQeCrKAACptb397FkYmjBDxVOpVHYWlx/8sP9870dFHwsAAAAAAKBxVJQBAFLpweqjXwjJFFepVDrW0d52vejjAAAAAAAANJagDACQOvP3FwdOd3ddNjPFVu7qfDNsv1X0cQAAAAAAABpH6yUAIHXWNx5/EoYkzAxPn37x9OTJ1rbCDwQAAAAAANAQKsoAAKmysLj8YyEZqk6ebD0ZtuEyIAAAAAAAQCOoKAMApMqTza21jva2slmhqlKp7CwurbzRf6FvzqAAAAAAAABHoaIMAJAaS8sPfiYkw26lUulYW9vJmwYGAAAAAAA4KkEZACAV5u8vDpzuLv8fZoOXOd1dHgrbchkcAAAAAADgKARlAIBUaG1tvdHScuK42eBVOjtP/aXBAQAAAAAAjkJQBgBI3PzC0tvd5c4fmgn2ErblWl5Z/dAgAQAAAAAAh3VsZ2fH4AEAiXr8+ZPFzlMdvWaB/WxvP3u+8mD1e/0X+uYMFgAAAAAAcFAqygAAiVpcWpkQkqFeYXuusE2XAQMAAAAAAA5DRRkAIFFPn36xdfJk60mzwEHMLyz9oP9870cGDQAAAAAAOAgVZQCAxDxYffQLIRkOo/NUx98YOAAAAAAA4KAEZQCARMzfXxw43d112ehzGGG7rrBtl8EDAAAAAAAOQuslACARj9bWPz7dXR4y+hzW9vaz5y0tJ04YQAAAAAAAoF4qygAAsZtfWHpbSIajamk5cTxs32UgAQAAAACAeqkoAwDE7snm1lpHe1vZyHNUlUplZ3Fp5Y3+C31zBhMAAAAAANiPijIAQKyWlh/8TEiGRimVSsc6Otr/zoACAAAAAAD1UFEGAIjV9vazZ2HLHKNOIy0sLv/p+b5zf21QAQAAAACAvagoAwDEZvXh2pSQDM3Q2XnqLw0sAAAAAACwH0EZACAW8wtLb3eXO39otGmGsJ1X2NbL4AIAAAAAAHvRegkAiMX6xuNPyl2dbxptmmV7+9nzlQer3+u/0DdnkAEAAAAAgJdRUQYAaLrFpZUJIRmaLWzr1draesNAAwAAAAAAr6KiDADQdE+ffrF18mTrSSNNs1UqlZ3F5Qc/7D/f+5HBBgAAAAAAdlNRBgBoquWV1Q+FZIhLqVQ61tHedt2AAwAAAAAALyMoAwA0zfz9xYEzr3f/RyNMnMI2X2G7L4MOAAAAAADsJigDADRNa2vrjbDChxEmbqdOdbxr0AEAAAAAgN0EZQCApphfWHr7zOvdl4wuSehobyuHbb8MPgAAAAAAUOvYzs6OAQEAGu7J5tZaGFYwsiSlUqnsLC6tvNF/oW/OJAAAAAAAAIGKMgBAMywurUwIyZC0sO1X2P7LRAAAAAAAAFUqygAADbe9/exZS8uJ40aWNJhfWPpB//nej0wGAAAAAACgogwA0FCrD9emhGRIk85THX9jQgAAAAAAgEBQBgBopPn7iwPd5c4fGlTSpPNUR2/YDsykAAAAAAAAWi8BAA2zvvH4k3JX55tGlLR5+vSLpydPtraZGAAAAAAAKDYVZQCAhggrdgjJkFYnT7aeXF5Z/dAEAQAAAABAsakoAwA0xJPNrbWO9ray0SStKpXKzuLSyhv9F/rmTBIAAAAAABSTijIAwJGFlTqEZEi7Uql0rK3t5E0TBQAAAAAAxSUoAwAcyfz9xYFyufMnRpEsON1dHppfWHrbZAEAAAAAQDEJygAAR9La2nrjtZYW7ynIjHJX59+aLQAAAAAAKCabWgDAoYWVOc683n3JCJIlYZuwxaWVCZMGAAAAAADFc2xnZ8e0AwCH8vjzJ4udpzp6jR5Zs7397HlLy4kTJg4AAAAAAIpFRRkA4FDCihxCMmRVS8uJ4w9WH/3CBAIAAAAAQLGoKAMAHMr29rNnYdjA6JFVlUplZ3Fp5Y3+C31zJhEAAAAAAIpBRRkA4MDCShxCMmRdqVQ61tZ28qaJBAAAAACA4hCUAQAOZP7+4sDp7q7LRo08ON1dHlpYXP6xyQQAAAAAgGIQlAEADqSjo/3vwkocRo286Ow89ZcmEwAAAAAAikFQBgCoW1h5o9zV+aYRI0862tvKS8sPfmZSAQAAAAAg/47t7OyYZgCgLk82t9bCUIHRIm+2t589b2k5ccLEAgAAAABAvqkoAwDUJay4ISRDXrW0nDi++nBtygQDAAAAAEC+qSgDAOxr/v7iQM/ZM5+EYQKjRV5VKpWdxaWVN/ov9M2ZZAAAAAAAyCcVZQCAfbW2tt4QkiHvSqXSsY6O9r8z0QAAAAAAkF+CMgDAnuYXlt7uLnf+0ChRBOWuzjcXFpd/bLIBAAAAACCftF4CAPb0+PMni52nOnqNEkXxZHNrvaO9rduEAwAAAABA/mihAAC80uLSykRfb4+QTINUKpUwhPGNH/bll9svvupVOl4K2ttOfuNPHz9+/Fv/jcPraG8rL6+sfniu58yfG0YAAAAAAMgXFWUAgFd6+vSLrZMnWyUw9rG59TR4/vz570MvzyvPg83Npy/+0ubWVvD8eSXW4+nqPPXi+/EXoZq2F//c2dnxjf/H3ra3nz1febD6vf4LfXOGCgAAAAAA8kNFGQDgpaKKGkIykWo1mDAU8+WXX74Iwnzx5ZcHqgYTl43Hn//+kR4+Wv/Wo1YDNK+91hK0vvZa0N5+Mngt/K4qze+1tJw43traeiMIgrGUHBIAAAAAANAAKsoAAN8yf39xoK+359NSqXSsiKMThmE2X4Ritl4EYpKoCpOU9va2F4GZ6vciV6CpVCo7i8sPfth/vvejFBwOAAAAAADQAIIyAMC3PFpb//h0d3moCCMTVooJK7CEgZjwe201Fr5SG5qp/nNRrG88/m25q/N7lgIAAAAAAOSDoAwA8A3zC0tv95/v/Ye8jko1GPPV15MXlWM4mLB1Uxia6ew89VV4JufBmcWllZ/29fZMpuBQAAAAAACAIxKUAQC+4cnm1lpHe1s5T6MShmIera0LxjRJGJw53V1+EZp5/XQ5KJVKuXp+Tza31jva27pTcCgAAAAAAMARCcoAAL+3uLQy0dfb837WRySsGvPw0XoUjvk8eP68koKjKo6wPdPr3eXg9OlybqrNLK+s/tW5njN/noJDAQAAAAAAjkBQBgD4ve3tZ89aWk4cz+KIfPnldvBwbT1YefBQ1ZgUee21lhdVZnrOnsl0aKZSqewsLq280X+hby4FhwMAAAAAABySoAwA8MLyyup/Pddz5n/N0mgIx2RL1kMzqw/Xbp95vXssBYcCAAAAAAAckqAMABDM318c6Ovt+bRUKh1L+2jUtlUKv5NNYWjm3NkzQc/Z11/8c1bMLyz9oP9870eWHQAAAAAAZJOgDAAQrG88/qTc1flmmkdi4/HnLyrHhAGZ588rKTgiGqW9vS0439vzotpMqVRK9bg+/vzJUuepjr4UHAoAAAAAAHAIgjIAUHALi8s/Pt937v9N4yiE1WOWHzwMj/FFmyXy7fjxUnC6uxyc7zuX6tZMi0srP+3r7ZlMwaEAAAAAAAAHJCgDAAX3ZHNrraO9rZymUdjcevoiHBNWkKGYqlVmwtZMafP06RdPT55sbbM0AQAAAAAgewRlAKDAlpYf/Kz33Nn/lJYRCIMx4VfYZgmCqMrM+d5zwfm+nlS1ZVpeWf2rcz1n/jwFhwIAAAAAAByAoAwAFNj29rNnLS0njic5AtorUa+wuszF/vPBa6+1JD5mlUplZ3Fp5Y3+C31ziR8MAAAAAABQN0EZACio1YdrU2de776U1LMPAzILiyvBwtJy8Px5xTKkbmFgJvzq6jyV6KCtPly7feb17jEzBwAAAAAA2SEoAwAF9Ght/X/q6jx1p1QqHYv72QvI0ChhUOa7/X2JBmbmF5Z+0H++9yOTCgAAAAAA2SAoAwAFtL7x+JNyV+ebcT5zARmaJcnAzJPNrfWO9rZukwsAAAAAANkgKAMABbO4tDLR19vzflzPWkCGuIRBmTcGvxu0t52MdcwXl1Z+2tfbM2miAQAAAAAg/QRlAKBgnmxurXW0t5XjeNafzS8KyBC7nrOvBxf7zwevvdYSy0Nvbz973tJy4oSZBgAAAACA9BOUAYACWV5Z/fBcz5k/a/YzXnnwMPjd/ELw5ZfblheJCcMy5/t6glKp1PRDeLD66ObZM6d/ZLYBAAAAACDdBGUAoCDm7y8O9Jw980lLy4njzXrGG48/f1FFJvwOaXD8eCn4bhiY6e1p6tFUKpWdxaWVN/ov9M2ZeAAAAAAASC9BGQAoiNWHa1NnXu++1IxnW6lUgv9+77MXlWQgjdrb24I3BvqDrs5TTTu6R2vrd053l79vAQAAAAAAQHoJygBAAcwvLL3dd+7s35dKpWONfrYLSyvBZ/MLwfPnFUuJ1Hv9dDn4gzcHm9aOaX5h6Qf953s/shIAAAAAACCdBGUAoADWNx5/Uu7qfLORz3Rz62nwb7+9F2xubllCZEoz2zE92dxa72hv67YiAAAAAAAgnb5jXgAg3xaXViYaGZIJ2yx9OvdZcOcffy0kQyaF1Y8+vfdZcOdX//wi8NVIHe1t5fn7i/+XlQEAAAAAAOmkogwA5NyTza21cPO+Ec9y4/HnL6rIfPnltmVDblzsPx98t7+vYU9ne/vZ85aWEyesEAAAAAAASB8VZQAgx5ZXVj9sREimWkXmn379GyEZcud38wsNrS7T0nLi+PLK6n+1UgAAAAAAIH1UlAGAnJq/vzjQc/bMJ+Gm/VGeoSoyFEmjqstUKpWdxaWVN/ov9M1ZQAAAAAAAkB5H2jgDANKrtbX1xlFCMmEVmbDSxsLiilmmMMI1/3BtPfiDNweD9raTh37apVLpWEdH+98FQfA9qwcAAAAAANJDRRkAyKH5haW3+86d/ftws/4wzy5sQRNWkdnc3LI8KKTjx0vBd/vPB+d7e4709BcWl//0fN+5v7aKAAAAAAAgHQRlACCH1jcef1Lu6nzzMM9sYWkl+Gx+IXj+vGJpUHivny6/qC5TKpUONRRPNrfWO9rbuos+jgAAAAAAkBbfMRMAkC+LSysThwnJhK2W/uU3vw0+vfeZkAxEHj5aDz7+x18HG48/P9SQdLS3lZeWH/zMeAIAAAAAQDqoKAMAOfNkc2st3Jw/yLMKWy39879+Enz55bblAK/wxuB3D9WKaXv72fOWlhMnjCsAAAAAACRPRRkAyJHlldUPDxqSWXnwMLjzj78WkoF9hNWWwqpLYfWlg2hpOXF89eHalPEFAAAAAIDkqSgDADkxf39xoOfsmU/CTfl6n9G//fbei6AMUL/29rbgD94cDNrbTtb9dyqVys7i8oMf9p/v/chQAwAAAABAclSUAYCcaG1tvVFvSCasiHHnV/8sJAOHsLm5FfzTr/81eLS2XvdfLpVKxzra264bbwAAAAAASJagDADkwPzC0tvd5c4f1vNMNreeBv/t41+92OwHDuf580rwz//62+Cz+cW6/365q/PNxaWVCUMOAAAAAADJ0XoJAHJgfePxJ+Em/H7PJKwg8+ncZy82+YHG6Dn7evA/DH43rBqz7897srm13tHe1m3oAQAAAAAgGSrKAEDGhRUq6gnJhJUv/u2394RkoMHCANqvfv2bFy3N9tPR3lZeXln90BwAAAAAAEAyVJQBgIx7srm1Fm6+7/UswoBMuJkPNM9rr7UEf/SH3wva207u+Rjb28+erzxY/V7/hb450wEAAAAAAPFSUQYAMiysTLFXSCascPFPv/6NkAzE4Msvt4N/+vW/BptbT/d8sJaWE8dbW1tvmBMAAAAAAIifijIAkFHz9xcHes6e+STcdH/ZMwhDMmE7mM3NLVNMnt0NF5cFEQAAIABJREFUgmC95vndi75Cg9FXVRgqeyuOsfiDNweDnrOv7/ln5heWftB/vvejOI4HAAAAAAD4yks31gCA9AsrUrwqJBNWtAjbLQnJkBNhi6I7NV/royPDU0d5atMzs0NRcGasJlBzqVHDFf7+hfYKy3Se6vibIAj6GvWYAAAAAADA/lSUAYAMml9Yerv/fO8/vOzIw5BM2P7l+fOKqSWrwmDMzSAIwjDM1OjI8HpczyMK0AxFAZrwa+AoP2+/yjKLSys/7evtmTzKYwAAAAAAAPUTlAGADHr8+ZPFzlMdvbuPXEiGDLsdhWNujo4M30vL05iemR2sCc1cCYKg66A/IwzKhIGZl3n69IunJ0+2tjXh0AEAAAAAgJcQlAGAjFlcWpno6+15f/dRC8mQQXeDILietnDMXqZnZquBmfGDhGb2Css8WH108+yZ0z+K4fABAAAAAKDwBGUAIGOePv1i6+TJ1pO1Ry0kQ8bcCIJgcnRk+E6WJ256ZrYamLlcz59/VVimUqnsLC6tvNF/oW+uGccJAAAAAAB8TVAGADJkeWX1w3M9Z/6s9oiFZMiIjTAcEwVk1vM0aVF7pjAwM7FflZlXhWUera3fOd1d/n4zjxMAAAAAABCUAYDMmL+/ONDX2/NpqVQ6Vj1mIRkyILcBmd2mZ2bLUVhmz8DMq8IyC4vLf3q+79xfx3jIAAAAAABQOIIyAJARqw/Xps683n2perRCMqRcYQIyu9UTmHlZWObJ5tZ6R3tbdwKHDAAAAAAAhSEoAwAZML+w9Hb/+d5/qB6pkAwp90EQBNeKFpDZLQrMXAuC4J2X/f/zfT3BGwPf/cZ/W15Z/atzPWf+PIHDBQAAAACAQhCUAYAMePz5k8XOUx29gZAM6XY7rKIyOjJ8xzx9bXpmdiyqrvPW7v8XVpUJq8tUVSqVncWllTf6L/TNJXW8AAAAAACQZ4IyAJByi0srE329Pe8HX22iB//t418JyZA2G1FA5rqZebXpmdmwusy7u//A7rDMo7X1O6e7y99PxUEDAAAAAEDOCMoAQMo9ffrF1smTrSfDkMyvfv2bYHNzy5SRJreCIBgvepulek3PzA4FQXAzCIKB2r+yOyyzsLj8p+f7zv11Sp8GAAAAAABklqAMAKTY8srqh+d6zvyZkAwptBEFZG6anIOZnpktB0EQVt+5XP2Lx4+Xgn//x38YtLedfPHvTza31jva27oz8YQAAAAAACBDBGUAIKXm7y8O9PX2fFoqlY7922/vBSsPHpoq0uJuEARXRkeG75mRw5uemR0PgmAyCIKu4CVhmeWV1b8613PmzzP55AAAAAAAIKUEZQAgpdbXH8+Uy53/s5AMKfPB6MjwhElpjN2tmNrb24L/8Mf/LiiVSkGlUtlZXFp5o/9C31wOnioAAAAAAKTCd0wDAKTP/MLS22FIJgzICMmQEmGrpR8JyTTW6MjwnSAIwrDMrfAHh+3VwjZrobCaVGtr643cPFkAAAAAAEgBFWUAIIWebG6tVSqV8j9FG+aQsLmo1dIdE9E80zOzYQjp/fABes6+HvzBm4MvHmt+YekH/ed7P8rlkwYAAAAAgJipKAMAKbO4tDJx7Nix8r/85remhjS4G1Y8EZJpvtGR4ckgCP4krN4TVpJaWFp58Zjlrs6/zfczBwAAAACA+AjKAEDKdHV1/sW//fZe8Px5xdSQtNtBEIyNjgyvm4l4jI4MT0WtmO5+eu+z4NHaetDR3lYOA3RFeP4AAAAAANBsWi8BQIo8WH30i4eP1q48fCSXQOJujI4Mj5uGZEzPzJaDIJg8frx09d//8R8GlUplqfNUR18RxwIAAAAAABpJUAYAUmL+/uLA8+eVf7m/uNxqTkiYkExKTM/Mjre3t/38P/zxvwsWlx/8oP9870dFHxMAAAAAADgKrZcAICXmPlv4qZAMKSAkkyKjI8PXNze3vv/p3PzTtpOt/6Xo4wEAAAAAAEelogwApMD0zOxgEASfmgsSdnd0ZHjIJKRP2IrpzTcuzvSeO/s/Fn0sAAAAAADgKI4bPQBIheumgYTdDYJgzCSk0+jI8HoQBEIyAAAAAABwRCrKAEDCpmdmw3DCL80DCdoIgmAwCmMAAAAAAADk1ndMLQAkTjUZkhSGZMaEZAAAAAAAgCIQlAGABE3PzF4LgmDAHJCgidGR4TsmAAAAAAAAKAKtlwAgIdMzs+UgCO4FQdBlDkjIjdGR4XGDDwAAAAAAFIWKMgCQnEkhGRI0F1aTMQEAAAAAAECRqCgDAAmYnpkdCoLgY2NPgr6v5RIAAAAAAFA0KsoAQDImjTsJek9IBgAAAAAAKCIVZQAgZtMzs1eCIPiFcSchc6Mjw4MGHwAAAAAAKCIVZQAgfqrJkKRxow8AAAAAABSVoAwAxGh6ZvZaEAQDxpyE3BodGZ4y+AAAAAAAQFEJygBATKZnZstBEEwYbxKyYf0BAAAAAABFJygDAPEJWy51GW8SMjk6MnzP4AMAAAAAAEUmKAMAMZiemR0LguCqsSYhc1FQCwAAAAAAoNAEZQAgHteMMwm6NjoyvG4CAAAAAACAohOUAYAmm56ZHQ+C4JJxJiG3R0eGrxt8AAAAAAAAQRkAaKrpmdmyajIkbMIEAAAAAAAAfEVQBgCaKwwpDBhjEnJjdGT4jsEHAAAAAAD4iqAMADTJ9MzsoGoeJGjD+gMAAAAAAPgmQRkAaJ6w5VKX8SUh10ZHhtcNPgAAAAAAwNeO7ezsGA4AaLDpmdmxIAh+aVxJyNzoyPCgwQcAAAAAAPgmFWUAoDkmjSsJGjf4AAAAAAAA33bcmABAY03PzIYhhbcMKwm5PToyPGXwAQAAAAiCoBwEwVBKB+Je9LWXoeg57OdOEARpbEM+GH3tJ63HD5BLgjIA0EDTM7Nl1WRImGoyAAAAAFQNpbhF/HtBEFzb58+E11ov1fGzPgiCYKJBx9VI4Q1tA3X8vD+J/iwAMdB6CQAaK/ww1mVMSch7oyPD+92FAwAAAAB5M15n5Zk4jdUZkgEgZoIyANAg0zOzYQnNd40nCdlQzQgAAACAgupKYaXl/arlAJAQQRkAaBwhBZI0MToyrI8xAAAAAEWVptZLQ3W2jAIgAYIyANAA0zOzYRnNy8aShNwdHRm+bvABAAAAKLCBFFWVSVNoB4BdBGUAoDFUkyFJPngDAAAAQDquk4Ut+q+m4DgAeAVBGQA4oumZ2fDD11vGkYTcGh0ZnjL4AAAAAPDiOu1YwsOQlqo2ALzCcQMDAIc3PTNbDoLgmiEkIRuqyQAAAACwhztBEPzJIQdovI7KKHePcH3q3iH/3n7C40nqxrKy63UA6ScoAwBHE4ZkuowhCZkcHRlu1gUFAAAAALJv/QihkXoqsxzl5zfL5aj9URLXza64XgyQfoIyAHBI0zOz4Yetd4wfCZkLgzIGHwAAaHCLiXsJbSwCQCNdS6gFkurjABkgKAMAh3fd2JGga6Mjw+smAAAoiKGojH0jrEctCCBPftnA5/KeTT4AcuBK9P4xzutn4WMOWDwA6ScoAwCHMD0zG96td8nYkZDboyPDgloAQJFMNvD99+0GV98AACB9wvZHEzGHPyesA4Bs+I55AoBDEVIgST50AwAAAMDe4ryGNuTGSoDsEJQBgAOanpmdUEKTBN0YHRnWKgAAAAAA9hZWlRmPaYzc2AaQIYIyAHAA0zOzZb3aSdCGD90AAAAAULc4ruUOBkFw1ZQAZIegDAAczGR0JwIkYXJ0ZHjdyAMAAABAXcLK4GNNHqq4qtYA0CCCMgBQp+mZ2SF3BpCgudGRYdWMAAAAAOBgmnlNrawCNED2CMoAQP0mjRUJ8oEbAAAAAA7uUtQeqRmuqEAOkD2CMgBQh+mZ2SvRBypIwu3RkeGbRh4AAAAAvuVWHUPSrKoy+/3cuSY9LgBHICgDAPVRTYYkqSYDAAAAAC93J7zRbJ+xuRq1SWqk8ObKgX1+nlbqACkkKAMA+5iemb1WxwceaJYPRkeG7xhdAAAAAHilem50bPTNaPv9vI0gCK43+DEBaABBGQDYw/TM7KBqHiRow10nAAAAALCvm3W0OZpoYFWZsTpa9atSDpBSgjIAsLcwpNBljEjItdGR4XWDDwAAAAD72u+Gs66oXVIjjNfxMwRlAFJKUAYAXmF6ZnYs6l0LSZgbHRn2YRoAAAAA6nMzqtC8l0ZUbx6s47rxjSAI3AAHkFKCMgDwalrekKR67koBAAAAAL6yXkcVl4EGVJWpp1W/a8sAKSYoAwAvMT0zO15Hj1loltujI8NTRhcAAAAADuR6HX+4nqDLq5TruMHtdhAE90wbQHoJygDALtMzs2WJfxKmmgwAAAAAHNy9qO3RXsIbJIcOObbhdbuuff6Ma8sAKScoAwDfNhGV4IQkvDc6MuyOEwAAAAA4nHqCKoetKrPf35sLgkClaICUE5QBgBrTM7ODQRC8a0xIyEYdfZQBAAAAgFe7F7U/2svVIAgGDziG43XcYKmaDEAGCMoAwDcJKZCkidGR4XUzAAAAAABH0oyqMvu1Sw9vgrtu2gDST1AGACLTM7NjQRBcNh4k5O7oyLAP0gAAAABwdFNRG6S9hMGXcp2PFF47vrTPn3ETJkBGCMoAwNd8kCFJh+2LDAAAAAB8235VZbrqqBJTVc+fc30ZICMEZQDgq2oyYUjhLWNBQm6NjgxPGXwAAAAAaJjrUTukvdRz89pgEARX9/kzN4Ig0FIdICMEZQAovOmZ2XKdPWuhWVSTAQAAAIDG26/Ky0Ad1WLquXbn+jJAhgjKAMBXH2K6jAMJeW90ZPiewQcAAACAhqunHdJeQZlyHUGa20EQuL4HkCGCMgAU2vTMbFg2852ijwOJmdO7GAAAAACaZj1qi7SXS0EQjL3i/4/XcZOlajIAGSMoA0DRXS/6AJCoa6Mjw3oXAwAAAEDz1BNkeVXVmP3aLoU3wk2ZO4BsEZQBoLCmZ2bHorsFIAm3R0eGBbUAAAAAoLnCtki39nmEq0EQDO76b2F4ZmCfv6eaDEAGCcoAUGRCCiTJh2gAAAAAiEc97c93X697VZWZqg3XmAGySVAGgEKanpm9VsfdANAsN0ZHhpVkBQAAAIB4hNfi7u7zSFeCIChH/1xPNfJ6wjcApJCgDACFMz0zW66jtyw0y4ZqMgAAAAAQu/2CLV011433qyYTCMoAZJegDABFNBl96IEkTI6ODN8z8gAAAAAQq7BN0tw+DxgGZAaDILi6z5+7EQTBuukDyCZBGQAKZXpmdqiODznQLHOjI8OqyQAAAABAMq7v86gDUZum/bjGB5BhgjIAFI1ymCRJyy8AAAAASM5k1Bp9LwP7/P9bQRCoGA2QYYIyABTG9MxsWDbzkhknIbdHR4ZvGnwAAAAASMx6HVVl9uNmTICME5QBoBCmZ2bLymGSMNVkAAAAACB5Rwm63K2zNRMAKSYoA0BRTNRRMhOa5cboyPAdowsAAAAAiQvbJt045EGoJgOQA4IyAOTe9MzsoGoeJGjD+gMAAACAVDlM+6W5BrRtAiAFBGUAKIIw5d9lpknItdGR4XWDDwAAAACpEbZPun3AgxGSAcgJQRkAcm16ZnYsCILLZpmEzI2ODCvHCgAAAADpc5Dgy4a2SwD5ISgDQN758EKSxo0+AAAAAKTS9aidUj1uBkGgajRATgjKAJBb0zOzYUjhLTNMQm6PjgxPGXwAAAAASK16b7S8ZgoB8kNQBoBcmp6ZLasmQ8JUkwEAAACAdLsetVXay60gCO6ZR4D8EJQBIK/ChH+X2SUh742ODPvwDAAAAADptl7HDZduyATIGUEZAHJnemZ2MAiCd8wsCdnw4RkAAAAAMuP6Hgd6NwgC7dUBcua4CQUgh/b6YAPNNjE6MrxulAEAAADIuGvRV5LGYnjssDL0sSb+/Gb+bAAOQUUZAHJlemb2ShAEl8wqCbk7OjIsqAUAAAAAAJBSKsoAkDda3pCkCaMPkHmD0Vew65/3ElYSu1Pz/+9E/w2giIaCICjX3P29+1y6340Nd2vOobXn16ld3wGa6aDvCb0fBADIEEEZAHJjemY2LAM6YEZJyI3RkWEX7QGyYyzazB2s+d7o9xFzUQnvOzXfbZpwUNXQQe0mXTn673u5s0/YAI6qev4cis6pjTqPvrXr3y9H39+t+W/V8+tUzbn1nhktnHpacTjnsZfyrnNY9bzW1cBR836Q3apr7VWB0iBqK69iMQA0kaAMALkwPTNbVs2DBG2koF8zAK9WvRBdDcfs3oRtloHoa3f1hLlog2SqZpMXGh062Ktqx8aucMEdm8nUIVyTV2rOpUnepPCy8+tczXl1SnAmF4Z2nRvLRwgxbNSEE3YHFijWmhqL+Ty21/vBqZrXYGsxv8ZqvuptF+99GQA0maAMAHkx2eA7fuAgJkdHhl2IB0iXcPNjPLooHVcwpl7VDZNqlYRw8+5mdEH8pjuMC6Ma4LqSQOigK9qo2b1Zc3dX0MBa5ErNOk179c7w+K5GX0G0nm9GXzags+Ewm8n16qr5mZd3/Z3bNWGFmwUa76K4UnMuS9N5bPc5a65mDXoNzr4rNV+uVwJACh3b2dkxLwBk2vTMbHix45dmkYSEF7OGRkeGXcQCSN5gVGEuCxu6e7lVs7nr9SVfqgGurKzR2zVrMelQ8FQDN85v19mypajGatZpXjb35qJ1PJnjSjONvMj7XowVM9O4mXyrJrDghohsykNIofp+UPudr13b1YbvKI414fiqn0XGG7Du4jwPA0Ahfce0A5ADPjiSpAkhGYDEjUd3gn8aBME7GQ/JBNGd7j8PgmAt2iC5koJj4vAGazbnP87YGg2DKe9Hv1t3os2fwRQcF403GH2uuhfdhHA1Z3fAD0S/e59GAYjxFBxTkQ1Fm//h56hfpHC9Xa4591XXSzkFx8Xeas9jaVxXB1V9P7ge/b4MZevwC2UsmqPqZxEVZAAgAwRlAMi06ZnZ8SaUZIZ63R4dGVaaGyAZ5WgzZD3aREhbe6VGuRxt9tyLQgo26rJjPNpgzUuA662ajWMBrvyo3dx7NwfrtB6XajafJ4W/YlUNtn6coRBDdb3cs15SK+/nsa7o9+Xj6PdH0C89BqP3er+saaEFAGSEoAwAmTU9M1tWTYaETZgAgNgNRpsha9FmSFHu2ByIQgr3ovc/NurSqRy9P7gXbazmNdBdG+C6JsCVSWM29168flSrzFx3Xm2q8ZrzYlaDrdZL+hTxPPZWTXBLgDo55Zpwlpv3ACCjBGUAyLKJgtzxSDrdGB0ZvmNuAGJTDch+WvA7NruigFB1o84GSTqUa9o9vF+g96gD0XoUmMmO2o1lm3tfuyoA0RS1AZk8nRetl2Q5j30zQO0GnnhVA9EqyABAxgnKAJBJ0zOzg9FFeUjChotRALGqXpD22v9NVwUUUqF2fRalwtFuXQIzqWdjuT4CEI0xFK23vAVkdrsquBqralVB57GvddUEZrRkaq5qm6X3C/x+DwByRVAGgKyaNHMk6NroyPC6CQBourGaCh0uSL9cNaBwxwZJ7KzPb6sNzAgVp4ON5cOpBiAEvw4uHLOPC7berjrvNd216L2OKh4vNxAF06aioBqNdSVaf15HASBHBGUAyJzpmdlwU+KymSMhc6Mjw4JaAM1VjkKxv9RmsW61GySqIDRXOL43rc891d7hPpbi48w7G8tH965KDXUbitZbUau/Vc97Xocbq3ZdCaXu71IUVJsU8muY8LX0F9YfAOSPoAwAWXTdrJEgF8kBmmss2hB5xzgfyqWaKgg03kS0PoW26zMQBYq0JYmXjeXG6hJE3NdEtDn/VsqPMw6Xot8/1WWO7pp1dWjvROtQWPVormv9CgD5JSgDQKZMz8xOuHOXBN0aHRmeMgEATaOKTONU2zHZ1G2MwWiTXJulw6m2JbmSxYPPGBvLzSOI+G3laCP5/bQdWMKq1WVuCgkeStGrEzVKNayqIu7BlVVkA4D8O26OAciK6ZnZsouSJMxdgaRW6eLcYM2G+NABLspXw1/rld8N3DHDJKTaysbGbmO9VXNXu4p8h3clGj8BmaPpiloXfBC9p1/P8pNJoaFonTqPNt+70XnhShQAK6py9D7Smnu1y9Hr8JXoO/sbj4IdXnMb552ossy4dVgX5zYAKAhBGQCy5JqLJSTovdGR4SJfCCcFShfnytFG2FgULBiM/v0o58bf36lZujgXftuILqDeiTZ/7lR+N6CSEs0khNBc1ZYhY1FgRjihfuVos87dxI1lw67xbCzHr+hBxKFoI9ma299ANFZCq3vzmttcb1mHdRGSAYACEZQBIBOmZ2aHoovqkIQN5YpJQuniXDUUU/0eRzuarqi1wKXqf4gCNHeji4ZTld8N3LQgaJBryurH5mp0Lil6BYR6qXLUXDbsGue6jeXEFDWIKCRzcNW1MqRK6UsJJ8TDOtzfpHUIAMUhKANAVggpkKSJ0ZFhd+DTdFHFmLFoIzuuYEy93oq+3omCM7eiTeSbld8N+P3gMGzuxq9aAWFMJY892QSOhw27o7GxnB5FCiI6Px7NO9Hv7niWn0SDWVPxe6fmnOVz3NdUNAKAgvmOCQcg7aZnZq/UVjaAmN0eHRl2pzNNVbo4d6V0cS4MnawFQfCL6AJdmkIyL3M52uBcC4+9dHHOBX/qVY5CGi5EJyPciPrYJt0rjUfjY8MuPu9EwctyUZ5wAwxFgQwhmfSoDSLmlUBDY1x1zvu9K9ZUYi5FY28dfuWKKtYAUDwqygCQatMzs2XVZEjYNRNAM5Quzg1Gd9CP5+DicBiauVy6ODcZVQmZrPxuQGsXXiarFRDm6qwSMJSh3+efR/PhfdbXtAJLzuXo3DDm7vZ9jUe/t2k/12xEwZE70ZzeO0C1lWrYZLDmK+0B4iCak18GQfCTHLYUE5JpLOe8r85lP0/BcRzE7Tr/bFbOWW9F5+WiVxoc1AYSAIpJUAaAtJvIyAUG8unG6MjwlLmlkUoX58aizdg8Vsrqiu7EC9szhReSr1V+N+B3iKqh6CJ0WkMyczWbutWN3cOu3+rm7lDN9zT+zr8fHZvqMtlqBVbdqKs3eFBdj+WUh9SqFTmuaA32SmneWL4dnTOnas6hh/Wqc+9YzTl1LMXrOW8txdIUkqkGV9frPE9UQ1dpfA1+q8BhmTSfy+5Ga+teND/1rrWXKde8FxyM5jptoequmnVY1Nfe60KAAFBMgjIApNb0zOxgji4ukj0bqsnQSFFromsFCv+FGxK/FJghksY70eei1gdTNRshjXKvZoOl1lj0dSVFG7zVcEiRwzJpDcnc3RXeOmr4oGr3ht1Qil6bBmzYvVLaKh7VnkNvxvSYu8+p5V3n1TS9x3onOr6sn1vL0fwm8fp9OzoPTEWvqUc9J5RrzntjKQnPFDEsk7aQzN2a81ijP6+8KnQ9VHPuutzgxzyMIodlJhI8F2zsOscdJAQIADTAsZ2dHeMIQCpNz8xm6c5e8ue90ZFhQRmOLKogc111rBebHROV3w248Fc8aQrJ3I1+H28eoAVIswxGG7vjKQnN3ChoWCZN7zc3dgUP4tw0ra7HtGzabbxkw26qgZtZt2sqTWRBWtZpdY1eb8KGciNU13GaqpLGeW5t5EXe96IWX3G2S6yur5tNCLC+TLkmZJX079et6DjybiKqZpe0NL0fLEdzfyUFr78ve+1Nm0aGRruj+Y/rM8rGrlCWVsUAkDBBGQBSaXpmdizq7w5JmBsdGR408hxF6eLcULTBkMcWS0dxIwrMFK3EfFGVY74A/TIb0WbI9RRf+K9W0RtPeKyKFpZJQ/igdmM4rqoc+ynXhLiSfA3bvWFX1KBMGtbpXPSe5nqGql5UWx9dSUFQM65za6ODMoMxrb0bKTgHVs97SVZ/zPtrcNKVZKrvBydTHFAo17wfTGodpj0s08igzI2YznG3aoJZAECKfMdkAJBSkyaGBGn5xaGVLs6VSxfnwnPYx0IyLxVejLxXujjn9yz/yglXkgk3dn9SE0JJ892x96JjHIw2JzcSOo6r0YX8Ikg6fFC7PsdTtnmyHo1PuFH2RhAEHyS0JqutIIYSeOy0SNM6ncxYa5g70e9W0ufVIKPn1vEmr72NaF7eSMk5sHreC9fLn0RhurhdzXHr3yRDMtXzWDWEkuYqHuvRGhiMjnkugWMo0mtvs19fb0TnuCtCMgCQTirKAJA60zOzaetZTbHcHh0ZzlIpflJEm6UDC8uej2vHlEvlmNs11JqLNhmyHPiobuZMJBQ0+iDnodEkwwe3o/WZxrY1e0lyTW5Em4Y3C1ZRJsl1mofz6MuMR4GfpAKcza4YkoWLvBvRHGQheJXU+/qf5Ox3L6nrO3k5j40nVOloIwrLpC1Y1MiKMs1yIzpOrZUAIOVUlAEgVaZnZsuqyZCwIrWcoEFqqsj8UkjmQMIQxceli3N5vXu2yK4nEJIJL+j/NNpQz/qmSO0dxR8k8Pjv5Pj18FpC4YPbUZWCsQyGZIJdazLu6hzVu9vLMT5m0iYSWqcbNS138lhd6nrCFWauFvyzxo1o/K9lpDrRVHS8P415vUzmqJrHUAIhmTy9HwwSPG91RQHVIr32HtVc9F5vXEgGALJBUAaAtLmWgh7yFNcHoyPDLmhwIKWLc0PRhfR3jNyhvVu6ODdVujg3mNHj55vCC/qXYx6TGzWtQfJkPdow/35UgSlOP49KxefJeAJ3IVdbPmQ1ILNbbWDmRoyP+1ZCFaqSEK7T9xN43FvRpnbew6vVNTwUPee45fHcup/azeMste+qqgZX4mrH1BW9l8p6QGEogde9vL4fDBI6b72Vk/cucbiR0JoHAI5AUAaA1JiemR200UyCNgqwMUCDlS7OjSfYXiZvwnYad0oX54q2eZQ34zFXQcj65lu97kQX39+L+XGv5+iu9rEE7mr/IBq/PFbmWI9+7/4k+j2kMZJYp+F74B9F4Y38DqDSAAAgAElEQVQiBcbvRc/5RwlUl8nTuXU/H+Rk8/he9Pv505ge762Mhz3K0TqP6yaoorwfTOK89VZO38c00k8KsPYAIJcEZQBIEx++SdK10ZFhFzaoW+ni3PVoM0sVrMYJx/IXURsrsifu8vp52Xw7iGtRdZm4ggl5Kbs/GD2PuMxF8zRRgE2Tqej3MIkWYXkzFPM6DaIKGXH/fqTNzWgM4qoWEuSoYsheqgGsvJ0HJ2N8Hb6a4epDcbbgvFXA94M3Y65yVPS2ca+yEQW0XMsEgIwSlAEgFaZnZq9E1QQgCXdHR4ZtzFOX0sW5ctgmKOaqGUXzTtSKSU/87CjHuEGR1823et2JufT+QMY30cvR8ccVaqyW3r8T0+OlQbVFWBKVOfIi7uoLQVQZY8wd8C+sx1wtJMhBxZC93E0o+BWX6utwHC0RsxiouhZjC86fRGGiIp7HqlWO4gqq/rxAlbDqMZejtpoAUFiCMgCkhZACSZow+tQjCm5MCfbFIhzjMCzjgmw2xBVEuBtdlC5y9YMg2hC6EmMrpksZbk84GdNd7RtK7//+Dvc4No/zJs7qC9U74H3++rbJaGziCnzlsUrD7eh1Ou9tvNaj892NJj9OV8aqVYRz/24Mj7MRVfZRyeOraxk/iemx8lBlsBHuFjAUDQC5JCgDQOKmZ2avRXcrQxJujY4MuwuIfUWBjXsxbmTx1VgLy6TftZjCY7eiDRgXpb92LcbNkXej8c+S8Ziqf1XvKrZh9/Ud7s3ePM6TOKsv3C1gi5KDmorWcFyBr8mo9VMe3ChglaLxGM53lzPSgqkcU5D5bvQ74/3g165HwaFmh/wGvNf5fWhfNTYAyAFBGQASNT0zW1bNgwRtWH/UIwpqTMXcEoGvdEVhGX3x02kopjuHbxS4tP5+4tocCTLWAmIopooZ7ir+tvWYNo/zIK7qC0FN2DDvlT4a4U6MYZmsVQx5lRs5rI5TrzjOd5MZeP2No33cbSGFV6qet5r9fvByga+hCMkAQM4IygCQtEkbzyRocnRk2GYBexKSSYVw7H8uLJNKcWzufVDgzbd6xbU5MpChdi1xbNjdsmGyp/EYKx5lUTnGgISw4cGtxxiWuZTx17kih2Sqmh2WGUh5OGEihspYRaxYdFBxhfyu5agSVr2EZAAghwRlAEjM9MzsWEzl8OFl5jK02UdChGRSR1gmXa7F0IrsJyp/1S2usMzVDLRgimNtCh7U57qwzCtdj6n9rBDD4cUZlslCxZCXuW19/V6zwzITKV0jg9HrbjM5j9UvjrBMXiph1UtIBgBySlAGgCQ1+2IK7GVidGTYhQ5eqXRxblBIJpWEZdIhjpZLPynYRfhGiCssk+YWTHGsTRt2ByMs821XYqi+EPqptXpkcYVlujL4+fhutJb52ngUHmqGrpSGh5tdJdhr7sFVWyA28/3gpYKE2YVkACDHBGUASMT0zOx49MEaknB7dGT4ppHnVUoX58IN4JtCMqklLJO8Zlfkek9I5tDiCMukuQVEs9emDbvDEZb5WjmmqoY/UT2xYeIKy7yToXYmG9G50Obxt11p4lp5N2VB1WaH/rzmHl4c7wevZbQSVr2c5wAg5wRlAIjd9MxsXBeH4VW08WA/UzG07eBoJqPWWMSv2WHXG6rOHdmdGF7r3k3hhu5EDGvTht3hXW9yW5KsuBZDyyVhw8aLo0JDkKHXv/HotYZva/ZaSctn2WZf17nrc/uR3Wlyu8yunF/bc54DgJwTlAEgCROqNJCgG6Mjwy528Eqli3PXhWQyIXwdmRKWiV25yZt4twQRGuZ61Hal2Y+RFs1emzbsGmO84GGZoahqSDMJGzbPnRhaDV3NQFWZG1HlRV6tmYHViZRU8ZhoYuhPu5vGudPkim5Xo9e2vPnAeQ4A8k9QBoBYTc/MDkZ3IEMSNmxysZfSxbmJ6GIf2RCGZa5HrbKIR7M3RYRkGmuyyaGES02+U/kgrjUxiG3DrrEmYmhhk1Zag2XfVAwhxDQHnXyeqt/1KADcaF0p+D0vN3EdaHfTeNej4Eez5K2qzIbAKQAUg6AMAHHTcokkXRsdGXbBjZcqXZwLN0HfNzqZ85a7/WLT7E2RKzZFmqLZoYQ0bCQMNrFKhw27xluPft+b3cImbcaa3BpM1aP4TDYpAFGV5qoyzocH06wWTEn/rk82MZyq3U1zhGvmdpN+dpqC040w6TwHAMUgKANAbKZnZsMPzpeNOAmZGx0Z/v/Zu7/jNpJzb8A4B74nvgQAOgLxFG5QhQvREYiOQNwIREewUATmRrBkBJYisHiBKtygLEZgDhM4ZAAsfzU6jV0sSUj80z09M/08VSr/uQCBmUbPYN5fvy2oxaNCRxJhi+56OxxXVv2ll3LrxLooctWmD9sjNwkLdYOWFEdSfv8V7NK4amALm7ZJuVXZra5HjavnhirhH23jfc2F++Vnu0nU/WWS8dq7n7ADp+1u0koZUu3Lb7FbC/wAoByCMgA0KeXDYfgRbej5nk8JAwA04+fQFYg0UnaTURRJ72viAkbO4oiCXXfVW9h8LOSzHifctm4gJJNFqgDExlG49raJUPLLfErUySPX79tU40BXrPRSzlt96SqjmwwAFERQBoBGLFfrk8QPh+F7Ps9n0y+OEI8JnUhSboVAcz6F7kDEl6qbjKJIc0572nJfwa7bFom3BmuLlAGDv+l6lM2XEKhLYa9lc9BF+Ly8TIpzmaMrV6pw6q2FLY35lHDruD6E6SzwA4CCCMoAkNxytR5ZfUZmCl08ajiuDupOJI5Ob+x5uJlMqnlUUaRZKbdgynGtTVmwK21boJz6Pg8cJVwwcGGLiOwWCbdgatN3w+/516nDbOeRX3Mvw7Uq1ZhcCPw1KtX9YNe7ylzYChYAyiIoA0ATTm1pQkYf57Ophx3sIlTRP++G40qBO67jRNfxj4oijbtKWOx8F4IrTUr1WRYKJY362vMtmFKFyAS62iHlViaTloRlLnWTiSLFNavJOSDVNpyXAn+NSzlvdTn8artNACiMoAwASS1X64NEK33hKW49dGOXsOXSGweol05twRRViqJIZX7O5jThVjdNdpUZJSoQKtjlkbIrR04HCbd3PA7FTvL7knArkzYUnc2JcVwlGCdNBmWOEgWndX/N41OiLTnfZwhOxyIoAwCF+ZMTDkBiHqqR08l8NlVA4AFbLvXeJBRdPXh/vYNEgbKFAm9W9XfjnwnewHGD37tUnY7MG/nUx/4fPfxMKVwo6LXOcbhm9pGxFs9p6MAWy14Yd0106Esxn53rVpRVfU7/leANHHdwu7ZKN0EAKM9//ec//3HaAUhiuVof9fBhN91xMZ9Nu7w/NgkNx9WXhCu8aY8/311PPPB8nbqg8yHya9YFXvNzfqnmwb82VFS9CqG4mM47vmVAHzR1fW5iHhqFcZoi0PVnBT0ekeIh72dbfEUX+/r1sYFQwkGiQIW5LL+zBF2gq4RdZRaJFry4BwSAAtl6CYAklqv1SDcZMuvaCiYaMhxXJ0IyxTgr/QBEkOKBsfm5HVKdhyYKqgcJQjIDY7MV+nQOUm1T8lFhmQbpJhNf7PvTJsLHKe4HfzGXtUKK6+6kg6F4nY0AoECCMgCkcpKogAFPcT6fTT3o4IHhuNpXCC3K2+G40rnk5VIUeS88iG6NL+F8xPY+dNJIKUXB7lzBrhVSjcscUozTW4sRaJhrdnyxgzJNbPmVYj4zl7XDVbgHiq1r3Vma2L4MAGgZQRkAoluu1vsJ9+OHH7k1/viO00Sru2kvwaiXS9EZxPlol1TnI3VATaejfuvDudhP1L2uvo+5SfC68JhKgDCJq3BsY9lLHJZJEZwWTm2XFNfdrm3ZJigDAAUSlAEgBYVocjqdz6YKCDwQOou8c2SKo6vMy8V+wF1Zmd469fm4TPCmUhZHFOz6rw9dZVJ8B3SToWmu2enE3tIqZVAmxX20cGq7XCW47u51KCyT4l4YAOgAQRkAolqu1grR5FTNZ1MP3dgldptzukOXqec7TBBGMD+3U4rCe8rCiIJdGbp+zU4xTnWToWk6LKQTO4S0n/C9xr6mXwintlKK+8GuLFZwbQWAQgnKABCbVY7kpBjOo4bjqi6CThydYr0bjquUBYQ+ShF0iL16mjg+hU4VMaXcBkLBrgxnCcZlU0aJFg4I/NI0QZl0YgdlUgUSDhL8hjKXtdOnyFuCDTrUUcZcBwCFEpQBIJrlan08GAzeOKJkcjGfTRVheSAEJISoMAaeJ3bB5dxqzda6SRRiSlG0U7ArS1fPTYqxb3swclA8TucmcihhlOidxp7Pbl13Wy32/eAkcbejWPxGAYBCCcoAEMVytR7pJkNmx04AOywSbCFD95gjnm6UIPgqyNhuXQnKKNiVRVDmd+ZQclA8Titm+C3VgqXY85m5rN1K3n4JACiQoAwAsShEk9Mv89nUKlseGI6r+sHce0eG+ho1HFfCMk+TIoygMNJuKbZfSrH1koJdWepuFpcd/MSxx2llrJJBF797XRN7+6UUXWVcd8tyleC734WgjO5ZAFAoQRkAXm25WtetVD84kmRyG4Ja8Bhjg21d2Sc/t9gBB0WRbuhCu30Fu/LELiSnpiMXfaGbTHqxj3Hs+7eDyIuhBKe7IfZ1twtBGfMdABRKUAaAGLSsJ6eT+WzqwQYPhO4hbx0ZtrwbjqsUq237JvYD7a4VukuV4jzFDMrELtgNFOw6oWu/M1J0UjJOoZ/a3sUi9nzmfrAbYl93J4m6HQEAvJqgDACvslytDxWiyehyPpsKavFACEPoJsNjdJX5MYWRMqUoxscMXcUel58jvx5pfE2wLVhKKbauM4cCT9H2Lm7msm5Icd1NESIFAHg1QRkAXktIgZxOHH12OAmr1+A+QZnv24/ctaMaDAZXKd4o0dXd2S4jv2jsjjIxKdh1R5fOVexCtXFKLsZe98Sef8xn5Spx+yUAoECCMgC82HK1XihEk9Hn+WzqYRsPhG4yQlTsYvul71MUKVvsbSAEZYhBUAboo7Z/v2N2Dr7twFZT/C722Ix9bQQAiEJQBoAXWa7WCtHkdGv88R2nkTti0D9WNe4W+9goinRL7PMVM9wSOyhjbHZHl8IisbekFZQBcogdbHDN7ZY2B6cBAKIRlAHgpRSiyel0PpvayoMHhuOqfgj33pHhBwRldovdbUdhpFtin69Y94qjyPedFxFfi/S6Mo+kKASaQ4EcdMcqW+zzFTtECgAQhaAMAM+2XK0PFKLJqApBLXjMmaPCEwjK7KZrR9lSFLJiFNuMSy47cARiF5YFuoBcYl93LXLpntjXXVvfAgCtIygDwEsIKZDTYj6b3jgD3DccV4dWq/FEb4bjysPax8U8LvU2eebr7qkiv+MY4YHY31cFu+7pQrhJRy6gL1x3iX3OYoevAABeTVAGgGdZrtbHCtFkdDGfTXUMYRchPp7Dw9rHvYn4Woq83dTGYpaOMnShyBp7nAoaArnYegn3SgBA7wnKAPBky9W6XlW0cMTI6MTB5zHDcXUcucBP/9l+KT1F3m4qYdW3le3dU2LBTmEZyCV2UIbuiX0fb5ECANA6gjIAPEcdUpg4YmRyPp9NrWpiFyE+nsvD2odiF0XM2d0UO0QSI5QWO9gmKNM9XQje2dIP4KFLx6STYt/Hu0YCAK0jKAPAkyxX633dPMjo1vhjl+G4Wgjx8QJWyj7kmABtVeLWSzrKALnEvCfUYRAAgFYSlAHgqepC9J6jRSaL+WzqARsPDMfVSIiKF7JVFzyu78X5ixa8B55PFyCA5liEgM6QAEDvCcoA8EPL1bpud//ekSKTaj6bnjr47CDEx4sNx5Xtl9LSDYFYdDsCAGhO7IVKsbfRBAB4NUEZAJ5CSIGcjh19HjMcV3Xh9IODwyvYKx+6wcp2AOgmwWkAAFpJUAaA71qu1se2pyCji/ls6sEauywcGV5JRxmA7rgs6FzZIgwAAAASEpQBYKflaj3STYbMdJPhUcNxZUs4YtBRBqA7Ym8DAQAAABRKUAaA7zkZDAZ7jhCZfJzPplcOPjvoJkMMOsoAAAAAABRGUAaARy1X6/3BYPCzo0Mmt7oZsctwXNWdht46QESgowwAbSTICQAAAAkJygCwi5ACOZ3MZ1Pt9dlFNxkAoM909QT6QjAdAIBWEpQB4IHlan04GAzeOTJkcjmfTc8cfB4zHFd1SGbi4BCJFftp7ff5wwEA8EPutwEAaCVBGQAeo5sMOZ04+jxmOK5GxgeRWbGflqBMNylo0Va2XQSAZsTuBPTFeQMA2kZQBoA/WK7WdRH6jaNCJufz2dQDFHY5FWwASK6NWyRUEV9LEIhUvkZ+XWMVgFxcgwCA3hOUAeA3y9W6LowsHBEyuTX+2GU4ruoHde8dIEgqdpFXRxliuYr4WgKXpHIT+XXbGFoDynAR8VMKXAAA0EqCMgBsWygekNHpfDaNWQijX2wJB+nFLvIKynTTYQGfUQChe0ostJpDgT7wjKmbYl93PesBAFpHUAaAb5ardf0g9oOjQSaVIAS7DMfV0WAweOsAQeco8nZTCSESq9u7J/a4TFGw05UL4HGuu93ThesuAMCrCMoAsHHmSJDRYj6bxu5kQH8IUUFzYrbanzhvnfSmhW/6S+TXE0DontjnLEXBLva9rMIykEvs665Obt1TQodBAKBwgjIA1N1kDnVrIKOL+WwqqMWjhuNqodgOneYhe7eUUpgXlOmeLpyz2OEbQRmgL9wPdk/s627s8BUAwKsJygAw0E2GzBZOAI8Zjqt65eGJgwONir11iEJvt7Q1jBB7XCrYdU/scxZ7TA0SBGUmujAAmejkVrZR5MUqt6UfUACgnQRlAAq3XK1PdGsgo/P5bGplEbvUWy7tOTrQKB0RytbWAIktbYhdZE215edl5NczVoE+EFDtltjXnhThVACAVxOUASjYcrUe6eZBRrfGH7sMx1X9cO69A0RiVjc+pKNM2dpayIodqt2zur1T9hME+1MV7WKHDRWXgRxiX3d1yOqW2Nee2NdGAIAoBGUAyrbQrYGMTuezqQcm7HLqyNAAqxsfin1M3iiMdMYonK+2ih1sE0DojtiBu9uEHWVsEwb0RRX5c5jPuqML2x0CALyaoAxAoZardf3A+YPzTyaVIAS7DMfV0WAweOsAQRY3CQIJR05lJ7S9gCWAUK4uFexid2F4K2wIZOK6W67Yv8UFZQCAVhKUASiXkAI5LeazaaqVvHSf+Ymm6Gr1OIWRMrU90BQ7gGBcdkeXgjIpXlvYEMgh9nxmLuuGFOcp9j0cAEAUgjIABVqu1ro1kNPFfDY9cwZ4zHBcLcIe9tAEQZnHxX6YrTDSDaV1lJkk2NKH+PYTbAmWMihzk2C7EnMokEPs+0HX3W6Ifc257OuBAgC6T1AGoEy6NZDTwtHnMcNxVW8tcOLg0CBBmcfFLozsKfS23kEHQoopwg3HCV6TuFLMHam3gIg9h76z/RKQgQ5ZZYp9jnSTAQBaS1AGoDDL1Vq3BnL6PJ9NPShhl9NQUIemCMo8LsU8rTDSbl0IKdbf19vIr2lctl/sMNNtB4MyA2MVyOAmQTcQAdV2O0rwe9zzHwCgtQRlAAqyXK11ayA3449HDcdV3c3gvaNDw1IXS7vsIvJ7f68jQqt1pQhvG4iydG3bpY1PCV5TcRnIwXW3LCnuBwVlAIDWEpQBKItuDeT0cT6b6t7ALraEo2nV3fXkxlHfKcVDbYXedjru0P1higCCEG97pZgzUoyh+1J0YXiruAxkkOJ+0HW3nUYJFq5chmsiAEArCcoAFGK5WuvWQE63ghDsMhxXR6EABE0S3Ps+gYRydCnAlGpLG92O2inF2GxqZbs5FOiDFHOZ6247pbjGnPXhwAAA/SUoA1AOIQVyOpnPplYSsYv5iRy0Af++enuSKvJrTjq0xU8pDjsWVLxKMC73jMtWOgpzRkxVg1vupSguvw/bUQE06XPkv7Un+NdKXe3iBgDwYoIyAAVYrtbHujWQ0eV8NrWSiEcNx9UiQSEMnkJQ5sd0ROi/Lm6HlWJcLhK8Jq+TYq5oct5PETYcGKtABimuu7bjbJfjROFUHTwBgFYTlAHoueVqPfJAlcwURXnUcFyNjA9yubueCMr8WIqQ49vQxYT89ju6LWeKcTlRtGuVVJ2Oml7ZnmKs6ioDNC3F3Om62y4pnhnqGgsAtJ6gDED/nejWQEaf57OpYjS7LELrbWjahSP+JDoi9FtXCxjGZf+lOBe3PQnKDIxVoGE3CbZfGoS5bORkZpeim8zAtksAQBcIygD02HK13tetgcyMPx41HFf1/PTB0SETD26fLlVXGauI86o7drzr8Pu3ur2/UnWTybEN6FWiYGbdVeYgwesC7JKqm5vnBXml6kB9YdslAKALBGUA+k23BnL6OJ9NPRxhlxwFK9jQ6erpUnZEsIo4n653pEjVDUenjvxSnYNc9x2p/q4tLYAmfQqduWI7cT+YVaoO1H7rAwCd8CenCaCflqv1YVhtCDnceoDPLsNxlWq1ODxFdXc9+epIPdlVaLcfu/vIZhWxYELzjnswB286dcT+HJMwJo3LPI4Sjc3LsGVXDmdhPMUuRL4Nc6j77XY5TPRuvobtbyCnswQdQffCPKajW/NSdaC+FZQBALpCUAagvzw0JaeT+WzqYS67eHBGTrZder7TRNv0/BzOh+BSc0Y9ukc8SxSqOAmvrStes1KOzdxj/izMd7EtwhxqrLbDItF5vg0FbcjtNNHWue/DPKnjY7POEnWg9iwSAOgMWy8B9NByta5X47xxbsnkcj6bCkLwqOG4StXeGZ7Kw9vnqwsXVaLXdr1oVqqiSK7PkmIbiD3jMotU9we3LQhInhqrvXeQKCQzCOPHAgTaYNNlMIUzWzA1KlUHN52FAYBOEZQB6Jnlat2nlcJ0U4r2vfTAcFyNbGdBZpd31xMr718m1Xf3jXmhMUeJOgPllOqe9637mUb1PWRwk3ismkPzSxlY8tueNkk1HifmssbsJ5yzPgn2AQBdIigD0D8nPVopTPd8ns+mWiazy8L8RGaKTS93lrCrTF0gP2ziQxQsZVEkp1SdOgbhmmW7k/RGCcdmm1a2pxyr5tC8ThN2cz1XdKZl6t/6F4ne0gdzWSNSdhcUdgIAOkVQBqBHlqv1fsLVmPAjt1Zfs8twXO0n2tMenqoN2290XcqH35+03E/qU0+Diik7ddjWphmLhCGDNm1Zk3KsDsyh2Rwlvr9VdKaN3A921yLRlkuDEOzTuRMA6BRBGYB+sVKenE7ns6kHI+yi2EhuZ3fXE6uyXydlV5k9QaZkzhIGEdogZacO29qklTJk0KZuMhspx+pe6PSgwNycg8T3tx8VnWmplF1lNnMZ8R0mXFhn0RQA0EmCMgA9sVyt6x+975xPMqkEtdhlOK4OE65cg6cyR8WRMjTwVqguuuPBYPC+Z5/pvtSdOn4OgQ7iSh0yaFM3mY2bxIXEN651jRkl3r6kjUEv2JbyfvCN+8Ho9hMH0tt4zQUA+CFBGYD+8CCBnBbz2dSDEXYxP5Hb+d31xKrsOM4SriIehFDHcZMfqMfqIMKvhXzWlJ06BmHcHyR8/dKUHDKoP/dlwtd/776rEZ8Sd+o6UXSm5b6ErXZScT8YzyjxFpwWTQEAnSUoA9ADy9W6fpA2cS7J5GI+m3ogz6OG48r8RBvYOiWu1MfzV8WRVzsobOuC1F1lNluD2dYmjtJDBqnnN2GZtM4Sd0q8dP7oiEXikKr7wThKv+YCAOwkKAPQccvVeqQASGb2ouZRw3FlfqINdJOJL/Uq4kEojhzm+HA9sB/OUaqVw221CKuaU5mE4yos8zqpQwYXHQgZfB0MBr8k/hu6MaRx1sB2ds4bXXHVQCeRUx3dXqWJa27KLZ0AAJISlAHovkWBhRDa45f5bPrV+WAH8xO53QrzJXOSeBXxIDx4Vxx5ntTt9dsudYH5jYLQqzQRMujKnJ862DUIgUOB5XiaGL+/hCAVdEXquWwvhFTdDz5f6jnrVrAPAOg6QRmADluu1vXDgg/OIZncevjOLsNxtW9+ogVO764nWoGncdPAw3HFkecZheOVsr1+29Wf/3Pi9/jWtigvctxAyOBjh0IGTcyhtZ+N1yiaCMlUflvRUe4H26eJOWsRugoBAHSWoAxAt6Vucwvfs5jPpgrQ7KIoQ27V3fVEwSmtTw2EEjbFkaPcH7bl9oVkfnPcQLej965zz3IcupukdNnB30ZfQrgnNeP1dZooOA/C98RvK7roSwPbyQnLPF0Tc9aF55EAQB8IygB01HK1Pkq81zB8z+V8NvVghEcNx9Wh+YkW0Aq8GU2EEuriyD+c050OQhcNIZn/01SnjvchLDZq4G912UkDIZlBh0MGixDySe19KDIbr0+32cquiZDML+H8QFc1sZ2c8PSPNRGSseUSANAbgjIA3SWkQE4ejPA9Vi2T2+e764mCUzNuGixY/GpbigcOQ9For2XvK7cmuh3V3gkffFd9P/D3Bv5Ol7ZcesxRA4HDQQgx68jwNJsuXe8a+FuXrm30QFP3g8LTj9tsv9lU9ytbLgEAvSAoA9BBy9W6fpA2ce7I5ON8Nu1yMYKEhuPqxPxEZlY5Nq+p7UNqP+vi8Zt6nP9TSGan4wZWtw9CJx/hgz8aNbhdzUUPQgZXDQYON+PVdXK3o4a7dNlyib6ovzd/a+iz/Grx2G82wb4mOrqeh/twAIBeEJQB6Jjlar0fWphDDvWWS1Y88qjhuBpZEUsLHN1dTxScmrcIBesmvCs8mLAJITSxnU2XNdntaBM+OOznoXyWJle13/ZoC44vDRaY98L8cSZ0+AejUHj/R4MBxJ863g0J7jttqKNb7YOubt/uO5oK9l0KWQIAfSMoA9A9CyuHyciDEb7H/ERuv9hyKaumtg8ZbAUTSgsPHzQYQuiDJle374UOPyUHRg9Dd5SmOnEc9j1o4N0AACAASURBVKwTx2lYrd+U9+E7IuD1e7H5Q4N/89x2pfRUUx3dBqGLSpNdudpk0WBnwVvXCgCgjwRlADpkuVofKoyQkS2X2Gk4rvYbLi7AfZd31xMd1/K6CQ/RmwrL1IWBv4fgyH5bD0pEdUHkXw2GEPqi6fDBzwWNyW1NFuwGPe7EcRxW7TdlEs7baaFdGTZdZP7Z8Nahl7rE0mM3DYen90InqFLmsYNw/fu5wb/Zt2AqAMA3gjIA3WJLE3K5sOUSP2BFLDlZ5dgeXzMU/96Gv9vX61SOgkjfnDQcPnib6buQQ47x+UvP7zsOGx6vgxB2viqse+RJ+MxNB703W4YpOtNnXzPMJx/C3+1zd5kcoWlbxAEAvSUoA9ARy9X6ODx0h6bdFtrKmCcajqsj8xOZHd5dTxSc2uMsPFRv0l4o1Pep/f4oHMumCyJNF+ib0HS3o8G9jkcH+T56MqNMBbvzAgJIN5nCMvWY/TWM2T6HT4/DteLvmbYM3WxRBn33KcP94CR0l+nbtXczbzQdmv6bBTEAQJ8JygB0wHK13rSEhhyO5rOpAjTfY34ip5/uridWObbPWcPb3WxsF0i6WujdBBCuMmy52eftQHKEZQYhSPqv8J3oy3ZMx5m6HF0W1PGk6a1Ltr0NWxH1KTAz2grI/NrwNkvbdGagNLnuB/ty7T0Mc3HT28MNwnnzOx8A6DVBGYBuOMm02g1+ms+mX4o/Cuw0HFeLjMUG+Hh3PbHKsb2OMxVHBvcKvV0prO9vBWR+znDvd1nAFmY5t2R4v7VF2CjTe3itTcEuR9CghPF531WmcNfGZh7NsYVKLAehUJ47IDMIIRn3LJQo5/1gfe39d/judekash2QydG59bywrfgAgEIJygC03HK13s+wWhNq5/PZ1MNcdhqOq1EB2x/QXud315OF89N6OYsjg1Bc+DUUSRctXVV8FLYn+HemgMxgK4RQQge5Lxm2gtjYbBH2v6Fo15VtIY4zF+xKGp/3fQ2fvcr4Ht6EefQmdBdo+7jdD/enX0NHifctWHTyUUiGwuW+H3x/L/jXxsDqaKtjW67r7UBIBgAoiaAMQPt5oEYOn+ezqYcj/MipbldkUodkzFHdkbs4MghdBH4OYZSvoYiaMzRzGObQq7BV1LuM76XEEMJZxrDMxvtQxG9r0W67w9GvGQt2JYdkNr6GcMpl5vdR3/N9COP2KsxhbenQsJlTv4Z5/u8h4NMG5+G7BKVrw/3gJvj3vyGknPv6OwqB6bPwnn7NPHcJyQAARfmT0w3QXsvV+jDjQ2nKdenhCD8yHFcHocgHTROS6abNOWvDvPEmFFH/Hro0fNn6d5Xobx6EQu7mX1tChiWHEDZh9F8zv49N0a7+9zkU7lKOxe/ZDwW745aEDIRkfneztRVHG87NJIRmPoStoTbz6NfwnykdhLF6GP57m38vKzrDH7XpfvBd+PdruN5s3w+mvO5s3w+2af4yXwEAxRGUAWg33WRo2reCxHw2VZDgR04dITIQkum2NhVHNibh/Wze020o9H4NRZJNwfcm/H/fcxBWBo+2Crn7LS7iCiG0Jyyz8W6ru1B1L3zwo/H3Egf3QlyT5J/w6YzPhzZhmbPMXaju27s3dgdh/F5tzaFPnUv3t7p9jba2eToM/7stXWKe4qNOMvCoNt4Pvgn/PoT/vWsOu3pikHUzl23fEx60eA4TkgEAiiQoA9BSy9X6pGUPq+k/IRmeZDiujnS7IgMhmX44DoXSDy39NHthftvMcT9nfj+pCCH8rm1hmY37Ia5BOG83W0GupwS4NjZb5GwCXW2+jl+EzjbG50M3W9t0tLmz3yT86/tcustPFr3Ad7UxLLOtpDlMqA8AKJagDEALLVfrkR+qNExIhufQTYamCcn0y0ko7rctmFAKq4YfOgsBhLMWbYv1mM1K9D6HVY3PpzkOHQ7Mo+1yG65xQjLwY8fhfvDvjlU2Qn0AQNH+u/QDANBSpy1/SE+/CMnwZMNxtdDtioZ9FJLppfqh/F9DUZHmfBRC2OlT6LpiTOZjfD5PPY/+xZhtjdutrbGApzkNYQ3zWLNuw/XDfAUAFE1QBqBllqv1QcvbaNMvn4VkeKrhuBqFVbLQlJ/uric6rPXXJphwWfqBaMBtKET5Pn3f17A1kTHZLOPz5equMvthuyryuQzn4alboQG/Owv3g5Vj0ojN9ptfCvisAADfJSgD0D62NKEp5/PZ9EhIhmfQ7Yqm1EXT/7m7nljl2H9fw8P6z6UfiIQudTl4lqtwvM479J67rDI+X+0mHMOPHf8cXfVLCNj5TQUvtwmqCv2ldR6uF0J9AEDxBoIyAO2yXK2PBoPBW6eFBvw0n021tufJhuNKtyuaUhf1D+6uJx7glqMuLtb3QH8r/UAkoCDyMjdhCyBjMq3PoTBqfMaxCFtp6MrQjNuwhaBuixCH0F86t+Ge5lioDwDgd4IyAC2xXK1HusnQgPrB+f/MZ1Ordnku8xNN+FbUv7ueXDnaRarnmf+x7U0Um61sFEReZzMmBQ/i2hTsjozP6L6E8NEvPftcbXMRtlr6VPqBgAQWrr1RbToL+j0PAHCPoAxAe9Qr0SbOBwl9W7U7n02t2uVZhuNKtytS+1bUv7ueHN9dTxRNy7ZpvW818ctdhGMoFBvHV8GDqBTs0rsJvy3/IngY3SaEeCjkBUm59sbxUec2AIDdBGUAWmC5Wu9r2UxCt2GrpaP5bOqBLi+hmEVKl6GLjKI+2zariS8clSfbdOmoC7i6MsW1HTywwv3lFOyateku87cwP/A656GLjPsVaIbQ38tdhvvoRVc/AABAEwRlANqh/vG651yQQN1FZt9WS7zUcFwtdLsioY9315ODu+uJoimP+RpCHz8p8v7QpoAr2JiWbW1epg68/VnBLpvTMD+cF/r5X+siFOptZQd5fNnqNuh+8Ps2oWmhVACAJxCUAchsuVrXBaD3zgOR1aud/6KLDK8xHFcj3a5I5Nsqx7vriaIpT3EWirwKJA8p4DZvs8Jdx6Mfq+9H/6rLUSvchHnizwIzT7Y9fr905D1Dny2E/r7rF6FpAIDnEZQByM+PWGL6toJoPpvWXWQ80OW1TnW7IrJvc5QuMrzATSiQHCiQfLMJyCjg5rPpePRX2zE9cBuCbXXB7lPL3lvprgRmfqgKncyMX2gfob+HzsPxOBGaBgB4HkEZgIyWq3X9A/+Nc0AEvxUk5rOp8BWvNhxXB7pdEdm3VY531xNzFK+xXeQtscOMgEz7fAoF9Z8EZv4QkNExrN3uB2Z06/q/+fWvYfzathbaTejv94DMsa5tAAAvIygDkMlytR7pJkMEVdiDug7ILGyzRETmJ2L59hD37npycnc9MUcRy9VWC/6/FRBQOBeQab2zrcBMaVsyVfcCMub67tgUmzdj97Kwz3+7VWw+1EEGOmczh/2/cB3q+/3gJpAqIAMAEMGfHESAbBa2NOEVPtcFmfls6mEu0Q3H1dFgMHjryPIKt6FofHp3PfEAl5RuQrDvNBQ566LBUU/usarwuT4phHTKWfi3GY997s52sfV56babrXN5EMbucY9/r34Oc6uxC/2w2aJzEe4D6/nrXY/OrTkLACCB//rPf/7juAI0bLla1yv2/u2480yX4cHIp/lsqmBGMsNxVY+viSPMC1xuCvu6x5DRKBRJjjpYJKm2CiFfM72Hg8hdxUoPUmzG40lPtlxtwxilOUch9HXU8XvD29CN61P419d7lJgdx0qfu5vinKXT5fvBwfbzH4HpP9gEOWM6cU8DAGUSlAHIYLlaf9GtgSe62DzQFY6hCcNxVa/C+9nB5hl+K5reXU88YKSNNoXew5YGFS62Cri+Q/22v7XSvUuhmepeyIAyHWzNpYcd6DazmVu/2LIOijfaCv0dtjT4tx3o+yIcAwCQnqAMQMOWq3X9o/yfjjs7bB7o1oWyL/PZVEcGGjMcV6PwQM62cPzIJsT3RTiGjtkUSg62/rPJOa8K1/ivirfFa3vR7mKrWGee5zEH9/7lXAhyEe5ht+dXgF32790LNn0/OAgdY7bnLNdaAICGCcoANGy5WtvShEF4KLL9MPdqPpt6MEJWw3FVt3Z+7yxwzx8K+3fXE8Un+ma0VSQZbf3n6IVdPy7Dth43m2v81jVfAJZdRvcKdgcN/mYQMiCW/a0C9PZ8OnhFkKba6qywmU/vz68Ar7Vr/nrp/eAgXF8Hj9wLus4CALSAoAxAg5ardf1D+/SRv5hj9QpxXe4ofj14mGsLJdpoOK7qh4L/dnKKc3HvA28e2n57iCsUAw8cPvL/+Z6Q0v2C3aaQt21XAKG6FyL4uiPEBTltxrYwIdAVj12LB+4JAQC6RVAGAAAAAAAAAIAi/LfTDAAAAAAAAABACQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAAABRBUAYAAAAAAAAAgCIIygAAAAAAAAAAUARBGQAAAAAAAAAAiiAoAwAAAAAAAABAEQRlAAAAAAAAAAAogqAMAAAAAAAAAABFEJQBAAAAAAAAAKAIgjIAAAAAAAAA/H927UAAAAAAQJC/9SAXRwALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAClPCiwAAFVZJREFUAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAC1awcEAAAgDIPeP7U9HOQgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAA4L9tB1sbikqmuPFHAAAAAElFTkSuQmCC"
)
    _LOGO_LIGHT = (
    "iVBORw0KGgoAAAANSUhEUgAACMoAAAjKCAYAAABtSXCKAAAACXBIWXMAABcRAAAXEQHKJvM/AAAgAElEQVR4nOzaMQEAIAzAMMC/580HTSz07Z2ZAwAAAAAAAAAAv3sKAwAAAAAAAABQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAABAglEGAAAAAAAAAIAEowwAAAAAAAAAAAlGGQAAAAAAAAAAEowyAAAAAAAAAAAkGGUAAAAAAAAAAEgwygAAAAAAAAAAkGCUAQAAAAAAAAAgwSgDAAAAAAAAAECCUQYAAAAAAAAAgASjDAAAAAAAAAAACUYZAAAAAAAAAAASjDIAAAAAAAAAACQYZQAAAAAAAAAASDDKAAAAAAAAAACQYJQBAAAAAAAAACDBKAMAAAAAAAAAQIJRBgAAAAAAAACABKMMAAAAAAAAAAAJRhkAAAAAAAAAABKMMgAAAAAAAAAAJBhlAAAAAAAAAABIMMoAAAAAAAAAAJBglAEAAAAAAAAAIMEoAwAAAAAAAACw7Nw9alVBGIBhLrkn1wTBQs0PuATBWnAHaewsdQmWLsHSZWjnEuysBRt3YKKFjdFzcw5XUoggKIpGhPd5mmGqGeZrX4YEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAALSOeOEAACAASURBVAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkCGUAAAAAAAAAAEgQygAAAAAAAAAAkCCUAQAAAAAAAAAgQSgDAAAAAAAAAECCUAYAAAAAAAAAgAShDAAAAAAAAAAACUIZAAAAAAAAAAAShDIAAAAAAAAAACQIZQAAAAAAAAAASBDKAAAAAAAAAACQIJQBAAAAAAAAACBBKAMAAAAAAAAAQIJQBgAAAAAAAACABKEMAAAAAAAAAAAJQhkAAAAAAAAAABKEMgAAAAAAAAAAJAhlAAAAAAAAAABIEMoAAAAAAAAAAJAglAEAAAAAAAAAIEEoAwAAAAAAAABAglAGAAAAAAAAAIAEoQwAAAAAAAAAAAlCGQAAAAAAAAAAEoQyAAAAAAAAAAAkLI0ZAODf2ts/uLfZbA6n6ezu+cHzPN+Ypunq10usx/HKPM+L7y81DMN6OQynX/er1erV+bq1tXy5WCyO350cPzFKAAAAAACAH1tsNhvPAwBwAfYPDu/M83w0z9PtcRxvTWfT5bOz9YWGysOwPW2vtt8Pw/BmuRyei2cAAAAAAAC+EcoAAPwl52HMNE0PxvHz0XpcX7voKOZX7ezufrh0aefFcrl8enL89pl5AwAAAAAAVUIZAIA/cH1v/+F6vb7/6fT05v8SxvzM+Y8zO7u7r1er1WPRDAAAAAAAUCOUAQD4TXv7B/fGcXz08Qt7dxMb15nmi/0cl0WLlGRS/KZmLMpxZyZ9cTGuBkFk47bUyCKrwJpVT1ZS35m92UCAJEACy6vGrEzjbmbhQUurZDALS5jNzWLQ0owXFyAIl3a5Fx5YbHdE6oPil0iKlIkTHPpUm5YlskhWnc/fDyDkD4p16n1fnqo67/88z9O1d3d2dsKijt8bJ09u9vSc+n+6urqmFubvz+XgkAAAAAAAADpKUAYAoAWjY+fGnz9//n+ura5eKULlmMM68+ab/9bd3XP1wcL8F8U6cgAAAAAAgNYJygAA7GNkdOy9Z8+e/ceiV49p1anTpxe6u3v+9tHDB9PFOGIAAAAAAIDWCcoAALxEHJDZ3Ny4vra6+k4Vx6e7p2f5zJk3/ycVZgAAAAAAgDIRlAEA2KPqAZkXxS2ZenpO/Q8L8/fn8nVkAAAAAAAAhycoAwCQONvff3ttdfX9KrRYOoxarRa92dt368ni478szlEDAAAAAAD82GvGBACouqHhkamurjeeLy8tXRSS+bF4TJaeLF4+2d29MTwy+su8HR8AAAAAAECrVJQBACprdOzc+MbG+j9rs3Q4fWfP3ll68uRSkY4ZAAAAAAAgUFEGAKiquIrMk8XFr4RkDi+uvNNz6tTSyOjYe0U7dgAAAAAAoNpUlAEAKudsf//tOOxh5o+nVqtFfWf7//bxo4f/e5GfBwAAAAAAUB2CMgBAZcStllZXVxqbGxt9Zr19tGICAAAAAACKQuslAKAShkdGfxm3WhKSab+4Os/pM2fm4yBS2Z4bAAAAAABQLirKAAClNzg0/JvlpSf/687OTmi2O+eNkyc3+/rO/nRh/v5cWZ8jAAAAAABQbIIyAECpDQwOffZk8fFfm+V01Gq1qH9g8H9++GDhH6rwfAEAAAAAgGLRegkAKC0hmfTFVXueLD7+v+JWV1V77gAAAAAAQP4JygAApSQkkx1hGQAAAAAAIK8EZQCA0hGSyZ6wDAAAAAAAkEeCMgBAqQjJ5IewDAAAAAAAkDeCMgBAaQjJ5I+wDAAAAAAAkCeCMgBAKcRBjJXlpf9gNvNHWAYAAAAAAMiLMIoikwEAFNro2Lnxx48efh0HMsxkfr1x8uRmX9/Zny7M35+r+lgAAAAAAADZEJQBAArvZHf3xtazZ91mMv+6e3qWN9bXz1Z9HAAAAAAAgGxovQQAFNqbvb1fCckUx+bGRl9v39kvqz4OAAAAAABANgRlAIDCGhwa/s3a6uo7ZrBYVleW6/HcVX0cAAAAAACA9Gm9BAAU0sjo2HuLjx/9y87OTmgGi6dWq0WDQ8NvL8zfn6v6WAAAAAAAAOkRlAEACqnn1KmluI2P2Suu7p6e5Y319bNVHwcAAAAAACA9Wi8BAIXTPzD4uZBM8cVzGM9l1ccBAAAAAABIj6AMAFAoccul1ZXlD8xaOcRzGc9p1ccBAAAAAABIh9ZLAEChaLlUPqdOn154urY2VvVxAAAAAAAAOk9FGQCgMAaHhn8jJFM+60+fjsZzW/VxAAAAAAAAOk9FGQCgEEbHzo0/WVz86vnz7dfNWPmcONH1bf/AwE8W5u/PVX0sAAAAAACAzlFRBgAohK2tZzeEZMorntt4jqs+DgAAAAAAQGcJygAAuTcyOvbe2urq+2aq3JaXli7Gc131cQAAAAAAADpHUAYAyL3NzY3rOzs7oZkqv7W11X+q+hgAAAAAAACdIygDAOTa8MjoL9dWV98xS9WwubHRNzg0/JuqjwMAAAAAANAZYRRFhhYAyK2eU6eW4vCEGaqOEye6vt3e3jpR9XEAAAAAAADaT0UZACC3hoZHpoRkquf58+3Xz/b33676OAAAAAAAAO2nogwAkFtdXW88j0MTZqh6arVaNDA49P6Dhfkvqj4WAAAAAABA+6goAwDkUv/A4OdCMtW1s7MTbm5uXK/6OAAAAAAAAO0lKAMA5M7o2Lnx1ZXlD8xMta2trr4Tt9+q+jgAAAAAAADto/USAJA7b/b2fhWHJMwMb5w8uflsc7On8gMBAAAAAAC0hYoyAECuDI+M/lJIhqatZ8+64zZcBgQAAAAAAGgHFWUAgFzpOXVqaXNjo8+s0FSr1aLBoeG3F+bvzxkUAAAAAADgOFSUAQByY3Bo+DdCMrxoZ2cn3NzcvGlgAAAAAACA4xKUAQByYXTs3Pjqysr/YjZ4mdWV5XrclsvgAAAAAAAAxyEoAwDkwtbWsxvPn2+/bjZ4ladP1/7O4AAAAAAAAMchKAMAZG5kdOy9tdXV980E+4nbcg0MDn1mkAAAAAAAgKMKoygyeABApk6fOTO//vTpqFngICdOdH3bPzDwk4X5+3MGCwAAAAAAOCwVZQCATA0Nj0wJydCquD1X3KbLgAEAAAAAAEehogwAkKmT3d0bW8+edZsFDmN4ZPTnDxbmvzBoAAAAAADAYagoAwBkpn9g8HMhGY5iff3pPxo4AAAAAADgsARlAIBMjI6dG19dWf7A6HMUcbuuuG2XwQMAAAAAAA5D6yUAIBO9fWe/XF1Zrht9jurEia5vt7e3ThhAAAAAAACgVSrKAACpGxkde09IhuN6/nz79bh9l4EEAAAAAABapaIMAJC6nlOnljY3NvqMPMdVq9WiwaHhtxfm788ZTAAAAAAA4CAqygAAqRocGv6NkAztsrOzE25srP+zAQUAAAAAAFqhogwAkKqurjeexy1zjDrtNDQ88lcPHyz8g0EFAAAAAAD2o6IMAJCas/39t4Vk6ISnT9f+zsACAAAAAAAHEZQBAFIxMjr23trq6vtGm06I23nFbb0MLgAAAAAAsB+tlwCAVLzZ2/vV2urqO0abTjlxouvb/oGBnyzM358zyAAAAAAAwMuoKAMAdNzQ8MiUkAydFrf12tp6dsNAAwAAAAAAr6KiDADQcSe7uze2nj3rNtJ0Wq1WiwYGh95/sDD/hcEGAAAAAABepKIMANBRA4NDnwnJkJadnZ1wc3PjugEHAAAAAABeRlAGAOiY0bFz4yvLS//BCJOmuM1X3O7LoAMAAAAAAC8SlAEAOmZr69mNuMKHESZt6+tPPzLoAAAAAADAiwRlAICOGBkde295aemi0SULmxsbfXHbL4MPAAAAAADsFUZRZEAAgLbrOXVqKQ4rGFmyUqvVosGh4bcX5u/PmQQAAAAAACBQUQYA6ISh4ZEpIRmyFrf9itt/mQgAAAAAAKBJRRkAoO26ut54/vz59utGljwYHhn9+YOF+S9MBgAAAAAAoKIMANBWZ/v7bwvJkCfr60//0YQAAAAAAACBoAwA0E6jY+fG11ZX3zeo5Mn606ejcTswkwIAAAAAAGi9BAC0zZu9vV+tra6+Y0TJmzdOntx8trnZY2IAAAAAAKDaVJQBANoirtghJENebT171j0wOPSZCQIAAAAAgGpTUQYAaIueU6eWNjc2+owmeVWr1aLBoeG3F+bvz5kkAAAAAACoJhVlAIBjiyt1CMmQdzs7O+Hm5uZNEwUAAAAAANUlKAMAHMvo2LnxtbXVXxlFimB1Zbk+Mjr2nskCAAAAAIBqEpQBAI5la+vZjefb295TUBhra6v/ZLYAAAAAAKCabGoBAEcWV+ZYXlq6aAQpkrhN2NDwyJRJAwAAAACA6gmjKDLtAMCRnD5zZn796dNRo0fRnDjR9e329tYJEwcAAAAAANWiogwAcCRxRQ4hGYrq+fPt1/sHBj83gQAAAAAAUC0qygAAR9LV9cbzOGxg9CiqWq0WDQ4Nv70wf3/OJAIAAAAAQDWoKAMAHFpciUNIhqLb2dkJNzc3b5pIAAAAAACoDkEZAOBQRsfOja+uLH9g1CiD1ZXl+vDI6C9NJgAAAAAAVIOgDABwKBsb6/8cV+IwapTF06drf2cyAQAAAACgGgRlAICWxZU31lZX3zFilMnmxkbf4NDwb0wqAAAAAACUXxhFkWkGAFrSc+rUUhwqMFqUzYkTXd9ub2+dMLEAAAAAAFBuKsoAAC2JK24IyVBWz59vv362v/+2CQYAAAAAgHJTUQYAONDo2LnxJ4uLX8VhAqNFWdVqtWhwaPjthfn7cyYZAAAAAADKSUUZAOBAW1vPbgjJUHY7Ozvhxsb6P5toAAAAAAAoL0EZAGBfI6Nj762trr5vlKiCtdXVd4ZHRn9psgEAAAAAoJy0XgIA9nX6zJn59adPR40SVdHd07O8sb5+1oQDAAAAAED5aKEAALzS0PDIlJBM+9RqtaC7u+cHP6/rja7gRNcbLT/Gzs5O8Gxj40f/bXNz45V/h8PZ3NjoGxgc+mzx8aO/MXQAAAAAAFAuKsoAAK90srt7Y+vZs24jtL84/BKHYJqhl72BmO6e7/5fmp6ure0+2t4AzXry354+Xcvd+OXRiRNd3/YPDPxkYf7+XNXHAgAAAAAAykRFGQDgpeKKGkIy32uGX0729ARdXV27/xwHY7oOUQ0mLafPnPnjI/X29X33D2PfP/hugGZjI9je3gq2t7d3K9TEf6pK873nz7df39p6diMIgkt5OSYAAAAAAOD4VJQBAH5kdOzc+ONHD7/e2dkJqzg6cQimu6c7OBn/ufvP6VeFyUoclolDNJubm7sBmipXoKnVatHA4ND7Dxbmv8jB4QAAAAAAAG2gogwA8CObm5s3qxKSiQMwp0+f2a0UE/+5txpLFXUn4aC9muGZ9adrwebGZmUqz8S/A5ubG9eDIPhJDg4HAAAAAABoAxVlAIAfGBkde+/hg4V/LeuoNIMxp858F4p5MRTCweLWTU/X1narzayvrZU+ODM4NPzrRw8fTOfgUAAAAAAAgGMSlAEAfqDn1KmlzY2NvjKNShyMebOvTzCmQ+LgzMry0m7FmZXl5d1/L5Punp7ljfX1s+WbOQAAAAAAqB5BGQDgj4aGR6YeP3r4SdFHJK4a09vXF7zZe3Y3HBP/O+mJK8zEgZnV5eXSVJvpHxj8+8XHj/4mB4cCAAAAAAAcg6AMAPBHXV1vPH/+fPv1Io5IV1dX8Gbf2aB/YEDVmBzZ3t7aDc0sLS4WOjRTq9WiwaHhtxfm78/l4HAAAAAAAIAjEpQBAHYNDA39pyePH/+PRRoN4ZhiKXpopu/s2TtLT55cysGhAAAAAAAARyQoAwAEo2Pnxh8/evj1zs5OmPfR2NtWKf6TYopDM08WF4OlxcfB9vZ2YZ7D8Mjozx8szH+Rg0MBAAAAAACOQFAGAAje7O39am119Z08j8Tp02eCswMDQW/f2d2wDOURV5d5/PDBbrWZnZ2dXD+vU6dPLzxdWxvLwaEAAAAAAABHICgDABU3PDL6y0cPH/zfeRyFOBBzdmAwGBoeDrq63sjBEdFJcUhmZXkpePzwYa5bMw0ODf/60cMH0zk4FAAAAAAA4JAEZQCg4npOnVra3NjIVQ+j7u6eYHB4OOgfGMzB0ZCFZpWZuD1T3rxx8uTms83NHgsDAAAAAACKR1AGACpscGj4N4uPH/1veRmB/oGB4Gz/YHD6zJkcHA15EFeZefTwwW5oJk9tmfoHBv9+8fGjv8nBoQAAAAAAAIcgKAMAFdbV9cbz58+3X89yBLRXolVPFh8HD+bvB9vb25mPWa1WiwaHht9emL8/l/nBAAAAAAAALct0YwwAyM7Z/v7bWYZk4oDM4PBIMDQ8svvPcJC4FVf8FQdmlhYXg6dP1zIbs52dnXBr69mNIAgumTgAAAAAACgOFWUAoIJ6+87+xfrTtUa82Z/2sxeQoV2erq3tVpjJMjAzPDL68wcL81+YVAAAAAAAKAZBGQCooDd7e79aW119J81nLiBDp2QZmOnu6VneWF8/a3IBAAAAAKAYBGUAoGKGhkemHj96+Elaz1pAhrTEgZn7f/gm2NzcSHXMB4eGf/3o4YNpEw0AAAAAAPknKAMAFdNz6tTS5sZGXxrPemTsnIAMqXuy+Hi3wsz29nYqD33iRNe329tbJ8w0AAAAAADkn6AMAFTIwODQZ08WH/91p59x/8DAbkimq+sNy4vMLMzfDx4/fBDs7Ox0/BDO9g/cfLL4+C/NNgAAAAAA5JugDABUxOjYufEni4tfPX++/XqnnvHp02d2AzKnz5yxrMiFOCTTDMx0Uq1WiwaHht9emL8/Z+YBAAAAACC/BGUAoCLO9vffXl5autiJZxu3Vjr3p28F/QODlhO5tLm5Edz/5pvg6dO1jh3em719jZXlpZ9ZAQAAAAAAkF+CMgBQASOjY+8tPn70Lzs7O2G7n+3g8EhcrWY3LAN5t7K8HHwz93XH2jENj4z+/MHC/BcWAgAAAAAA5JOgDABUwJu9vV+tra6+085n2t3dE7x14cLun1AknWzH1N3Ts7yxvn7WggAAAAAAgHx6zbwAQLkNDY9MtTMk02yz9Gc//XdCMhRSvIb/pENreHNjo2907Nz/YWUAAAAAAEA+qSgDACXXc+rUUrx5345nefr0md0qMl1db1g2lEZcXebB/P22PZ0TJ7q+3d7eOmGFAAAAAABA/qgoAwAlNjA49Fk7QjLNKjLv/NmfC8lQOqNj59paXeb58+3XB4aG/pOVAgAAAAAA+aOiDACU1OjYufEni4tfxZv2x3mGqshQJe2qLlOr1aLBoeG3F+bvz1lAAAAAAACQH8faOAMA8mtr69mN44Rk4ioyI2PngqHhEbNMZcTVZXr7+oJv7t0LNjc3jvy0d3Z2wo2N9X8OguAnVg8AAAAAAOSHijIAUEIjo2PvLT5+9C/xZv1Rnl3cgiauItOuVjRQNDs7O7vVZR4/fHCsIx8aHvmrhw8W/sECAAAAAACAfBCUAYASerO396u11dV3jvLMBodHdqtqxBVloOpWlpeDb+a+3g3OHEV3T8/yxvr62aqPIwAAAAAA5MVrZgIAymVoeGTqKCGZOBhz4b/5SfAnf/qWkAwk4jZMf/bTfxecPn3mSEOyubHRNzg0/BvjCQAAAAAA+aCiDACUTM+pU0vx5vxhnlXcYunCO+8EXV1vWA7wCv/fH745UiumEye6vt3e3jphXAEAAAAAIHsqygBAiQwMDn122JBM/8DAbsUMIRnYX1xtKa66dNiKS8+fb79+tr//tuEFAAAAAIDsqSgDACUxOnZu/Mni4lfxpnyrz+it8QtB/8CgJQCHsLm5EXxz797un62q1WrRwODQ+w8W5r8w1gAAAAAAkB0VZQCgJLa2nt1oNSQTV8SIq8gIycDhxa3K3vmzPw96e1sv3rSzsxNubm5cN9wAAAAAAJAtQRkAKIGR0bH31lZX32/lmcSb/D/993+x+ydwNHHY7MI7PwlGxs61/PfXVlffGRoemTLkAAAAAACQHa2XAKAE3uzt/SrehD/omfQPDATn/vT87iY/0B5PFh8H9//wTVw15sCf193Ts7yxvn7W0AMAAAAAQDZUlAGAgosrVLQSkokrX7w1/raQDLRZ3MIsbsXUyu/W5sZG38Dg0GfmAAAAAAAAsqGiDAAUXM+pU0vx5vt+z+Kt8Qu7m/lA52xvbwX3/u3fgs3NjX0f48SJrm/7BwZ+sjB/f850AAAAAABAulSUAYACiytT7BeSiStcvPPf/rmQDKSgq+uN3coy3d09+z7Y8+fbr29tPbthTgAAAAAAIH0qygBAQY2OnRt/srj4Vbzp/rJnsBuSaWHTHgrubhAEy3uewr3kK3Yh+WqKQ2XvpvF0v5n7OniyuLjv9wyPjP78wcL8F2kcK61zfgAAIABJREFUDwAAAAAA8J2XbqwBAPkXV6R4VUgmDse8deGCkAxlEbcoauz5Wm7Mztw+znOrT0zWk+DMpT2BmovtGq+3xt/e/XO/sMz6+tN/DIJgrF2PCQAAAAAAHExFGQAooJHRsfcePlj415cdeRyOiSvJxBVloKDiYMzNIAjiMMztxuzMclpPIwnQ1JMATfw1fpyfd1BlmcGh4V8/evhg+jiPAQAAAAAAtE5QBgAK6PSZM/PrT5+OvnjkQjIU2J0kHHOzMTtzLy9Poz4xeWFPaOZyEAS9h/0ZTxYfB9/MvfwpvXHy5OazzU2lnwAAAAAAICWCMgBQMEPDI1OPHz385MWjFpKhgO4GQXA9b+GY/dQnJpuBmauHCc3sF5Y52z9w88ni479M4fABAAAAAKDyBGUAoGBOdndvbD171r33qIVkKJgbQRBMN2ZnGkWeuPrEZDMw80Er3/+qsEytVosGh4bfXpi/P9eJ4wQAAAAAAL4nKAMABTIwOPTZk8XHf733iIVkKIiVOByTBGSWyzRpSXumODAzdVCVmVeFZd7s7WusLC/9rJPHCQAAAAAACMoAQGGMjp0bf/zo4dc7Ozth85iFZCiA0gZkXlSfmOxLwjL7BmZeFZYZGh75q4cPFv4hxUMGAAAAAIDKEZQBgII4299/e3lp6WLzaIVkyLnKBGRe1Epg5mVhme6enuWN9fWzGRwyAAAAAABUhqAMABTAyOjYew8fLPxr80iFZMi5T4MguFa1gMyLksDMtSAIPnzZ/3/08EFw/w/f/OC/9Q8M/v3i40d/k8HhAgAAAABAJQjKAEABnD5zZn796dPRQEiGfLsTV1FpzM40zNP36hOTl5LqOu+++P++mfs6eLK4+Md/r9Vq0eDQ8NsL8/fnMjlYAAAAAAAoOUEZAMi5oeGRqcePHn4SfLeJHvz03/+FkAx5s5IEZK6bmVerT0zG1WU+evEbXgzLvNnb11hZXvpZLg4aAAAAAABKRlAGAHLuZHf3xtazZ91xOCauJBNXlIEcuRUEwdWqt1lqVX1ish4Ewc0gCMb3/pUXwzJDwyN/9fDBwj/k9GkAAAAAAEBhCcoAQI4NDA599mTx8V8LyZBDK0lA5qbJOZz6xGRfEARx9Z0Pmn9xZ2cn+Lf/+l+Czc2N3X/v7ulZ3lhfP1uIJwQAAAAAAAUiKAMAOTU6dm788aOHX+/s7IRvjV8I+gcGTRV5cTcIgsuN2Zl7ZuTo6hOTV4MgmA6CoDd4SVimf2Dw7xcfP/qbQj45AAAAAADIKUEZAMipN3t7//Pa6up/LyRDznzamJ2ZMint8WIrpjgkE4dl4tBMrVaLBoeG316Yvz9XgqcKAAAAAAC58JppAID8GRkdey8OyfQPDAjJkBdxq6W/FJJpr8bsTCPOywRBcCv+wXF7tbjNWvBdhZlwa+vZjdI8WQAAAAAAyAEVZQAgh3pOnVqqvVbra26YQ8bmklZLDRPROfWJyTiE9En8AE8WHwffzH3X2Wp4ZPTnDxbmvyjlkwYAAAAAgJSpKAMAOTM0PDIVREHfhXd+YmrIg7txhkNIpvMaszPTQRD8Iq7eE1eSGhwe2X3MtbXVfyr3MwcAAAAAgPQIygBAzjxdW/vbty5cCGq1mqkha3eCILjUmJ1ZNhPpaMzO3E5aMd39kz99K+jt7Qs2Nzb6dgN0AAAAAADAsWm9BAA50j8w+Hlv39nLvX19poWs3WjMzlw1C9moT0zGJ4HpnZ2dK//2X/9L8FrttYWna2tjVRwLAAAAAABoJ0EZAMiJ0bFz47Va7f8dGhk9aU7ImJBMTtQnJq9ubm78Ng7LDAwO/fzBwvwXVR8TAAAAAAA4Dq2XACAnRs/9ya+FZMgBIZkcaczOXO/u7vnZuT95a/PZs2f/serjAQAAAAAAx6WiDADkQH1i8kIQBF+bCzJ2tzE7UzcJ+RO3YvrD7+f+8+NHD/+7qo8FAAAAAAAcx+tGDwBy4bppIGN3gyC4ZBLyqTE7sxwEgZAMAAAAAAAck4oyAJCx+sRkHE74nXkgQytBEFxIwhgAAAAAAACl9ZqpBYDMqSZDluKQzCUhGQAAAAAAoAoEZQAgQ/WJyWtBEIybAzI01ZidaZgAAAAAAACgCrReAoCM1Ccm+4IguBcEQa85ICM3GrMzVw0+AAAAAABQFSrKAEB2poVkyNBcXE3GBAAAAAAAAFWiogwAZKA+MVkPguBLY0+GfqblEgAAAAAAUDUqygBANqaNOxn6WEgGAAAAAACoIhVlACBl9YnJy0EQfG7cychcY3bmgsEHAAAAAACqSEUZAEifajJk6arRBwAAAAAAqkpQBgBSVJ+YvBYEwbgxJyO3GrMztw0+AAAAAABQVYIyAJCS+sRkXxAEU8abjKxYfwAAAAAAQNUJygBAeuKWS73Gm4xMN2Zn7hl8AAAAAACgygRlACAF9YnJS0EQXDHWZGQuCWoBAAAAAABUmqAMAKTjmnEmQ9caszPLJgAAAAAAAKg6QRkA6LD6xOTVIAguGmcycqcxO3Pd4AMAAAAAAAjKAEBH1Scm+1STIWNTJgAAAAAAAOA7gjIA0FlxSGHcGJORG43ZmYbBBwAAAAAA+I6gDAB0SH1i8oJqHmRoxfoDAAAAAAD4IUEZAOicuOVSr/ElI9caszPLBh8AAAAAAOB7YRRFhgMA2qw+MXkpCILfGVcyMteYnblg8AEAAAAAAH5IRRkA6Ixp40qGrhp8AAAAAACAH3vdmABAe9UnJuOQwruGlYzcaczO3Db4AAAAAIRh2BdfsszpQNyLoujeft8QhmF87H0t/KxGFEW5a0MehmFc9bmVys+5PH6AshKUAYA2qk9M9qkmQ8ZUkwEAAACgqZ7jFvEfB0Fw7YDvia+1XmzhZ30aBMFUm46rneIb2sZb+Hm/SL4XgBRovQQA7RV/GOs1pmTk48bszL534QAAAABACV1NqufkRhiGl1oMyQCQMkEZAGiT+sRkXELzI+NJRlZUMwIAAACgonpzWGn5oGo5AGREUAYA2kdIgSxNNWZn9DEGAAAAoKpy03opDMN6iy2jAMiAoAwAtEF9YjIuo/mBsSQjdxuzM9cNPgAAAAAVNh6GYV6qyuQmtAPAjwnKAEB7qCZDlnzwBgAAAIAcXCcLwzBu0X/FXADkl6AMABxTfWIy/vD1rnEkI7caszO3DT4AAAAABO+GYXgp42HIS1UbAF7hdQMDAEdXn5jsC4LgmiEkIyuqyQAAAACwj0YQBL844gBdbaEyyt1jXJ+6d8S/d5D4eDK5sSwMwz7X6wDyT1AGAI4nDsn0GkMyMt2YnenUBQUAAAAACi6KouWjhkZarMyyHEVR3qodfxC3P4qiKIvrZpddLwbIP0EZADii+sRk3Gv2Q+NHRubioIzBBwAA2txi4l5GG4sA0E7XMmqBpPo4QAEIygDA0V03dmToWmN2ZtkEAABVEIZhPQiCvjY91fiu54aFQ8n8ro1P52ObfACUwOW4DVJSUScVYRjG1WTGLR6A/BOUAYAjqE9MxnfrXTR2ZOROY3ZGUAsAqJLpNr7/vhMEQTurbwAAkD9x+6OplMOfU9YBQDG8Zp4A4EiEFMiSD90AAAAAsL/UrqElFRDdWAlQEIIyAHBI9YnJKSU0ydCNxuyMVgEAAAAAsL/eMAyvpjRGbmwDKBBBGQA4hPrEZJ9e7WRoxYduAAAAAGhZx6/lhmF4IQiCK6YEoDgEZQDgcKaT/raQhenG7MyykQcAAACAloyHYXipw0OVVtUaANpEUAYAWlSfmKy7M4AMzTVmZ1QzAgAAAIDD6dg1tTAM+1SABigeQRkAaN20sSJDPnADAAAAwOFdTNojdcJlFcgBikdQBgBaUJ+YjD/wXDRWZOROY3bmpsEHAAAAgB+51cKQdKqqzEE/d65DjwvAMQjKAEBrVJMhS6rJAAAAAMDLNeIbzQ4YmytJm6S2CcMwvrly/ICfp5U6QA4JygDAAeoTk9da+MADnfJpY3amYXQBAAAA4JVaudGx3TejHfTzVqIout7mxwSgDQRlAGAf9YnJC6p5kKEVd50AAAAAwP6iKLrZQpujqXZVlQnD8FILrfpVKQfIKUEZANhfHFLoNUZk5FpjdmbZ4AMAAADAgQ664Sy+znu5TcN4tYXvEZQByClBGQB4hfrEZHxXwBXjQ0bmGrMzPkwDAAAAQGtuJhWa93Ps6s1hGF5o4brxjSiK3AAHkFOCMgDwalrekKVW7koBAAAAAL5rv7TcQhWX8TAMj1tVppVW/a4tA+SYoAwAvER9YvJqCz1moVPuNGZnbhtdAAAAADiU6y18cytBl5cKw7CvhRvc7kRRdM+0AeSXoAwAvKA+Mdkn8U/GVJMBAAAAgENKAio3DvhbF8MwrB9xbOPrdr0HfI9rywA5JygDAD8W31EwblzIyMeN2Rl3nAAAAADA0bQSVDlqVZmD/t5cFEUqRQPknKAMAOxRn5i8EATBR8aEjKy00EcZAAAAAHiFpKrMnQPG50oYhhcOM4ZhGF5t4QZL1WQACkBQBgB+SEiBLE01ZmeWzQAAAAAAHEsnqsoc1C59JYqi66YNIP8EZQAgUZ+YvBQEwQfGg4zcbczO+CANAAAAAMeUtD+aO+CnXA3DsK+VRwrDML52fPGAb3MTJkBBCMoAwPd8kCFLR+2LDAAAAAD82EFVZXpbqBLT1Mr3ub4MUBCCMgDwXTWZOKTwrrEgI7caszO3DT4AAAAAtEfSBmnlgB924M1rYRheCILgygHfdiOKIi3VAQpCUAaAyqtPTPa12LMWOkU1GQAAAABov4OqvIyHYXhQtZhWrt25vgxQIIIyAPDdh5he40BGPm7Mztwz+AAAAADQdq20Q3plUCYMw74W2i7diaLI9T2AAhGUAaDS6hOTcdnMD6s+DmRmTu9iAAAAAOiMpB3SjQN++MUwDC+94v9dbeEmS9VkAApGUAaAqrte9QEgU9caszN6FwMAAABA57QSZHlV1ZiD2i7NRVF029wBFIugDACVVZ+YjO8SuGgFkJE7jdkZQS0AAAAA6KCkLdKtAx7hShiGF/b+hzAM4/DM+AF/TzUZgAISlAGgyoQUyJIP0QAAAACQjlban794ve5VVWaaVqIoco0ZoIAEZQCopPrE5LUW7gaATrnRmJ1RkhUAAAAAUpC0R7p7wCNdDsOwL/iumkwr1chbCd8AkEOCMgBUTn1isq+F3rLQKSuqyQAAAABA6g4KtvTuuW58UDWZQFAGoLgEZQCoounkQw9kYboxO3PPyAMAAABAepI2SXMHPODVMAwvBEFw5YDvuxFF0bLpAygmQRkAKqU+MVlv4UMOdMpcY3ZGNRkAAAAAyMb1Ax41btffSst01/gACkxQBoCqUQ6TLGn5BQAAAADZmU5ao+9n/ID/fyuKIhWjAQpMUAaAyqhPTMZ9ZS+acTJypzE7c9PgAwAAAEA2knZJB1WVOYibMQEKTlAGgEqoT0z2KYdJxlSTAQAAAIDsHSfocjeKolZaMwGQY4IyAFTFVAslM6FTbjRmZxpGFwAAAACylbRNunHEg1BNBqAEBGUAKL36xOQF1TzI0Ir1BwAAAAC5cpT2S3NRFB23bRMAOSAoA0AVxCn/XjNNRq41ZmeWDT4AAAAA5EPSPunOIQ9GSAagJARlACi1+sTkpSAIPjDLZGSuMTujHCsAAAAA5M9hgi8r2i4BlIegDABl58MLWbpq9AEAAAAgf5I2SnMtHtjNKIpUjQYoCUEZAEqrPjEZhxTeNcNk5E5jdua2wQcAAACA3Gr1RstrphCgPARlACil+sRkn2oyZEw1GQAAAADIt+tJW6X93Iqi6J55BCgPQRkAyipO+PeaXTLycWN2xodnAAAAAMixpJ3SQTdcuiEToGQEZQAonfrE5IUgCD40s2RkxYdnAAAAACiM6/sc6N0oirRXByiZ100oACW03wcb6LSpxuzMslEGAAAAoMiiKLqWVO7OTBRFlzr92ElbpbCDP79jPxuAo1FRBoBSqU9MXg6C4KJZJSN3G7MzgloAAAAAAAA5paIMAGWj5Q1ZmjL6AMUWhmHcwvFC8iT2/vN+4kpijT3/v5H0uQeonDAM60EQ9AVB0Lz7+8Vz6UE3NtxNzqvBC+fX3ZYHWh8AaTjCe0LvBwEACkRQBoDSqE9MxmVAx80oGbnRmJ1x0R6gIMIwjDdw68mmR/PPtr2PCMPdytpzQRDcSzZNmn/aNOFQ9oQO9m7S9SXrdj+Nl4UNhAxol2QTuZ58XWrjefTdF/79g+TPj4Ifn19v7zm33jO51ZK8lu/LOY/9hGHY98I5rHle623HwHk/yMvsCWG9KlAaux5FkYrFANBBgjIAlEJ9YrJPNQ8ytJJ1v2YAXi25GH0p+aq/ZBO2U8aTrx9UTwjDcC7ZJIk3725HUdQwfXQgdPDSqh3Jpt1Kc6OuuWlnM5mDJGv08p5zaZY3Kfzo/JqcW2/vObcKzhRcEhTce27sO2yIITnnBXvOe8svnPu8BldIsqYupXwe2+/9YDPs5/1giSWhvuZXq+3ivS8DgA4TlAGgLKbbdccPHMF0Y3bGhXiAHEk2Qq4mF6TTCsa0qrlhslslIQzDePPuZnJB/KY7jKthT4Drcgahg95ko2ZvyCBIWt7sDRpYixUXhuHlPes079U74+O7knzFx343ObfetAFdDEfcTG5V756f2axS1Dz33dkTVrhZuYEvueQ81jyX5ek89uI5qxmcuek1uPj2rLvLrlcCQD6FURSZGgAKrT4xGV/s+J1ZJCPxxax6Y3bGRSyAjCXBg6mCbOju59aezV2vLyWyJ8BVlDV6Z89azDQUHIbh7TZunN+JoujAli1VlYQVrpZsc28uWcvTZa00E4ZhOy/yfhxFUSoVM3O6mXxrT3jVDREFVJKQwq1kDWq/kwjD8FqzDd9xRVEUtuPnvHB8zc8iV9uw7lI7DwNAVb1m5gEoAR8cydKUkAxAtsIwvBqGYXwn+NdBEHxY8JBMkNzp/tsgCJbCMLyZbPZQUPGmSRiG02EYxputXxZsjcbBlE/i3634dywMw6lkE4iSSdbptWSd/i6pcFCmO+DHk9+9eC3fjl83cnBMlRWHBsMwvB6GYfw56vMcrrcP9pz7bifvM/pycFzs44XzWB7X1WHtvh+Mf0+S35d6sQ6/OuKAaTxHez6LqCADAAUgKANAodUnJq92oCQztOpOY3ZGaW6ADMQbVslmyHISKslbe6V2iTdJPo83fZKQgo26gkg2Vm+XKMD17p6NYwGuknhhc++jEqzTVlzcs/k8LfyVnj3B1i8LFGK4mLzPuGe95FMFzmO9ye/Ll0loVdAvJ5Jw1u09AVMAoEAEZQAorPrEZJ9qMmRsygQApCu5IB1vhiwlmyFVuWNzPAkp3EsCQjbqcigJcE0ld7P/tsSB7r0BrmsCXMWTbCxXfXOvd0+VmevOq52TBGTuFTzYar3kTEXPY+8mQT8B6gwl7/ea4Sw37wFAQQnKAFBkUxW545F8utGYnWmYG4B0NCvIJBekq3zHZm8SEGpu1NkgyYE96/NeEmiqynvU8WQ9CswUxAsbyzb3vndFAKL9XgjIlOm8aL1kyHls194AtRt4UpSM9z0VZACg+ARlACik+sTkheSiPGRhRTUZgPTsuSDttf+HrggoZO+F9VmVCkcv6hWYyTcbyy0TgGiDMAzryXorW0DmRVcEV9Ozp6qg89j34tffT5IKM1oyddCeNkufVPj9HgCUiqAMAEU1bebI0LXG7MyyCQDorGRj954L0vtqBhQaNkjSZX2+1N7AjFBxDthYPrJmAELw65CS6lpfVmy9XXHe66xkXTVU8Xil8aQl0+04qJbTYyysMAwvJ+vP6ygAlIigDACFU5+YvBQEwQdmjozMNWZnBLUAOihpYzOdbOxqs9iavRskqiB0UBI8uGl97mvvHe6XcnycpWZjuS2awS9BxAMkVWQaFa7+1jzveR1uoxfWlVDqweIgx5fx+2ghv/ZIXks/t/4AoHwEZQAooutmjQy5SA7QQcmmerwh8qFxPpKLzSoIBTz23EuqBTSEtlsWB4l+py1Jumwst12vIOL+knNjXEXm3TwfZ0ouJlXeVJc5pj3Viayrw/swWYfCqseQVGTT+hUASkpQBoBCqU9MTrlzlwzdaszO3DYBAJ2hikxbfRRvlNvUbY+kisxtbZaOrNmW5HJBj78wbCx3lCDiC5IKcNeTcyPfa1aXuSkkeHiqE7VNM6yqIu4hJec2FdkAoOReN8EAFEV9YjK+wOSiJFlyVyC5VTs/F2+GNzfE4770rV6Ub4a/lnd+P94ww2QhCXPctLHbdu8272qPokhFviNKwh3XBWSOLR6/z8Mw/DR+Tx9F0XLBn0+uxBvLyTp1Hu28j5LzwuUoiu6V/cm+ShIAuW3N7euD5HU4XiveZ7cgaXM27TW3rT5MKstctQ4P5twGANUhKANAkVxzsYQMfdyYnanshXDyoXZ+ri8JwVxKQjEXkn8/zrnxj3dq1s7PxX+sJG1F4q94zTd2fj+ukhIdI4TQcc2WIfF5Y0o4oXXJRsm0u4nbLm4HcSneDLVh1x42ljNR6SBiEsy6bc21JK7qcVtodX9eczvuXevwYEIyAFAtgjIAFEJ9YrKeXFSHLKwkF+0gVbXzc81QTPPPNNrR9CatBS42/0MSoLmbXDS8vfP78ZtWAu2QtK9QVj8d8cZTPbmrXfDzAKocdZwNuzZJ2t7YWM5GJYOIQjJH0lwr9SiKVCl9gXBCaqzDg01bhwBQHYIyABSFkAJZmmrMzrgDn45LKsbEmy2XUwzGtOrd5OvDJDhzK9lEvrnz+3G/Hxyazd1MNCsgXFLJ49VsAqfGht0x2FjOlcoEEZ0fjy1ugdMXRdHVgj+PtrGmMvFhMu6XVRr8XhiGKhoBQMW8ZsIByLv6xOTlvZUNIGV3GrMz7nSmo2rn5y7Xzs/FoZOlIAg+Ty7Q5Skk8zIfxBuc8THHx147P+eCPy2JN4jCMGy4EJ2ZeCPqy6RVCy9IxuVLG3apijfsbibBD1qQbHDeE5LJlT8GEcv6BAUa2uaKc953kvab1lQ2LiaV3Sq/DoPv16Iq1gBQMSrKAJBr9YnJPtVkyNg1E0An1M7PxW094jvor5bg4nAcmvmgdn4uPl/HwbLpnd+Pa+3CjxS4AsJcsil9kHqBfp9/m9zV7n1WQiuwTH2QbNhdcnf7/pIw13QBzjVx69JG8rWcnENbfW/QDJtc2POV9wBxkMzJ78Iw/FXZWooJybRd5c95ybnstzk4lMO40+L3FuWcFb8fv1f1SoNJu003RwFABQnKAJB3UwW5wEA53WjMztw2t7RT7fzcpSSAVcZKWb3JnXhxe6b4QvK1nd+P+x1iV7LJdj3HIZm5PZu6uxu7URQdaf0mF9wvJMGZ5p95/J3/JGl7U/nqMgVrBdbcqGs1eNBcj305D6k1K3Jc1hrs5XK+sXwnCVLEX41jbv6/9NybVGtpnlMv5Xg9l6qlWM5CMs3g6nLyWn2QZugqj6/B71Y1LJPzc9ndZG3dS9b98lFfk5KAeH3P6/ClHIaqe/esw6q+9l4XAgSAahKUASC36hOTzWoLkIUV1WRop6Q10bUKhf/iDYnfCcwQ5PdO9Hiz7WZzY7edm1RRFN3bs8HyR8kmb/x1OUcbvHELiKDKYZkch2TuvhDeOm74YNeeIFdzw66eo9emcRt2L5fDikd/PIdGUXQzjQd8MbyYbELvPa/m6T3Wh0nVrkKfW5MxvpnR6/ed5NwXz/u9454T9qyXZtAqD+GZyoVlchiSuZussZtHDUi/SjKnP/qZyfvi5rnrg3Y+5hFVNiwThuFUhueClb3nuGYI0PsfAEhPGEWR4QYgl+oTk0W6s5fy+bgxOyMow7ElFWSuq461u9kxtfP7cRf+KiZnIZm7ye/jzSTMkpkkrHA5ab+Wh9DMjSqGZXIWklnZE966meam6Z71mJdNu3gsfrBhF4bh7TZuZt2JouhSC9+XCzlap801er3dG8rtsGcd56kqaWrn1jAM23mR9+OkxVea7RKb6+tmuwOsL7MnOHM5B79ft6IoupzxMXRcEkr4JAeHkqf3g33JGrycg9ffH7325k2bQ6Nnk3BKWp9RVprv8ZJznFbFAJAxQRkAcqk+MRlfsPqd2SEjc43ZmQsGn+OonZ+rJxsMZWyxdBw3ksBMpUrMV1Vy8T/NC9Avs5JshlzP64X/ZHN3KgnNZDlWlQrL5CR88MeN4bSqchxkz6bd1Yxfw36wYVfVoExO1ulc8p7melGqXiQhzalkLWcd1Ezl3NqBoMyFlNbejazPgXvOe1lWfyz1a3AOKsk03w9O5zWgkKzD5vvBrNZhrsMybQ7K3EjpHHcref3Mxfs8AOB7rxkLAHJq2sSQIS2/OLLa+bm+2vm5+Bz2pZDMS8UXI+/Vzs/5PSu55GJ/lpVk4o3dX8UbfVEUTeX57th4wyY+xmRT8uNkkyILV5JN+dLLQfhg7/q8mqfNkzgIEUXR9SRI8nYQBJ9mtCabrSDqGTx2LuRlnUZRFK/T6SK1honP+UnoIOvzalDQc+vVDq+9lWRe3s7DOXDPeS9eL79IKiGm7UoSAiidjEMyzfNYX/J+MLdVPJJ1eC1Zh79Kjj1tVXrt7fTr643kHHdZSAYA8klFGQBypz4xmbee1VTLncbsTGFK8ZMv2iwdWlz2/Kp2TOWzJySTRUuheFMh3mQobOBjzx3FUxkFjT5NgjullHH44E6yPnPXtmY/Ga/JlSTscLNKFWUyXqeFP4++TLJZP51hgLOjFUPaXFGmU1aSOch98CoMw6ze1/+qTL97GYZkSnEeS8Yvi0pH8e9qPW/BojZXlOmUG8na01oJAHJORRkAcqU+MdmnmgwZq0zLCdpnTxWZ3wnJHEocoviydn6ulHfPVtz1DEIy8QX9XyeVDwq9KdK8ozgJB3yawSF8mGxu0i0RAAAgAElEQVTMlE6ywZJF+CAOyPwiDmcULSQT/HhNpl2dozcJ3vWl+JiZCsNwKqN1ulvlowzn0ZdJnlOWFWaulPXc2qIbSRWta0WoThSfq5PKHr9Oeb1Ml6WaR/I80g7JlOb9YJCct5J1mMVr780kKEtr5pL3eleFZACgGARlAMibaznoIU91fdqYnXFBg0OpnZ+rJxt4Hxq5I/uodn7udu383IWCHj97JFUQPkh5TJqbb6UK2ybhhHjD/GdJBaY0/TYMw8vZjkB7JRvUad+F3Gz5UMiAzIteCMzcSPGh382oQlXqknX6SQYPfSupHlDq8OqeNVxPnnPaSndubcHezePCtO9qSt5b1FNsxxRfD7le9IBCEpJJ+3WvlO8Hg+/WYRbnrXczmMOiupG8hhovACgQQRkAcqM+MXnBRjMZWkmCWtCy2vm5qxm2lymbuJ1Go3Z+rmqbR6WSbPCmWQWh0JtvrYqiqBFFUT25mzhN10t0V/ulDO5q/zTZNCljZY7lpIXML5LfQ9ogo3Uavwf+yyiKLlfpDvj4ucbPOX7uGVSXKc25tQWflmHzOFkvl5LqMml4t8iVdpOQz/UUb4KqyvvBLM5b7yYheF7tV2VfewBQVoIyAOSJD99k6VpjdsaFDVpWOz93PdnMUgWrfeKx/DxpY0XBZFBevxSbb4eR3E38sxSDCaUoux+GYRzGvpniQ8bz87O4GlDZN02S3796Ri3CSiU5h6a5ToOkQkZcfSHtx82N5LlfSLFaSFCWiiEHaAawSnUeTCqVpPU6fKXA1YfSbMF5q4LvB2+mXOWo6m3jXmUlCWi5lgkABSUoA0Au1CcmLyfVBCALdxuzMzbmaUnt/Fxf3CYo5aoZVfNh0opJT/yCSDb70tqgKOXmW6vi6jIpl94fz2Dzvm2StXkzxVBjs/R+I6XHy9yeFmFZVOYohQyqL8R+nbQEq3xQPFnDaVYLCYpeMeQAd5PzYCkDWHteh9NoiVi4QFUYhtdSbMH5q6QaVhXfDzarHKUVVP1thSphtSIOy5WirSYAVJmgDAB5IaRAlqaMPq1Ighu3BftSEY9xHJZxQbYY0goi3E0uSle2+kHw/abu5RRbMV1MNr6KaDqlu9pXql56f88d7mlsHpdNmtUXmnfA+/z1gmRMfpFi4KuMVRruJK/TpW7jlbwO15NwZCf1FqnybtI+7qMUHmolqdxW+UoeSVD1Vyk9XOGrDLbJ3aqFogGgrARlAMhcfWLyWnK3MmThVmN2xl1AHCgJbNxLcSOL78ZaWCbnkgBFGuGxW8nmm4vSiaQVU1qbIx8lG2CFkWxAp1H9q3lXsQ277zbHL6WweVwaKVdfuFu1FiWHlYzNpRQDX9NJe7gyuFG1KkVxODKF890HRWjBtKeCW6fdTVrGeT+YSN5//CyFkN+4lul/DO1XvhobAJSBoAwAmapPTPap5kGGVqw/WpEENW6n3BKB7/QmYRl98XMoKcGexp3DN6paWv8gKW6OBEVqAZGszTQqZrir+AVJpYU0No8LL8XqC8GesGGpK320Q/L7nFZYplAVQ/ZxI/m9r5yUznfTBXj9TaN93B0hhZfbc97q9PvBOLhV1WsoQjIAUDKCMgBkbdrGMxmabszO2CxgX0IyuRCP/W+FZXIpjc29T6u6+daqFDdHxgvULjONDbtbNkxeLfm9TaviUeEkm95pBSSEDQ8pGau0wjIXC96CqbIhmaYUwjLjeb7BIwlOdLoyVuUqFh1WiiG/ayWqhNUqIRkAKCFBGQAyU5+YvJRSOXx4mbkCbfaRESGZ3BGWyZGkXUinW5H9Kooilb9akGJY5kreWzCltDYFD1qQVDwSlnm56ym1n618iOGoUg7LFKFiyMvcsb6+k0JYZiqPayQJTFzr8MM4j7UopbBMWSphtUpIBgBKSlAGgCx1+mIK7GeqMTvjQgevVDs/d0FIJpeEZXIgpZZLv0o22WlRimGZ3LZgSmlt2rA7BGGZHwvD8HIK1Rdiv7ZWjyfFsExvAT8fx2NyOQfHkRvJ79udDh1Pb06rynS6SrDX3ENKzltXO/x+8GJFWjAJyQBAiQnKAJCJ+sRk/KH9otEnI3caszM3DT6vUjs/F28A3xSSyS1hmex1uiLXx0IyR5NSWCbPLSA6vTZt2B2BsMz3kpBZGlUN47Ch6oltkGJY5sMCtTOJX2Ou2jx+qcsdXCsf5SmomkLoz2vuEaX0fvBaQSthtcp5DgBKTlAGgNTVJybTujj8/7N3f8dtJFffgDEvfE99CQB0BIILhWvREYiOQNoIxI1AVASmIlgqAksRmLxGoQxGYIEJvMsAWPPV6G3YWIqQ+Kd7emb6eapU6/IFCc4Me4Zzfn0O7GOMBz9z0cLYDp7nLIzGomVVVaUOuzZFEV3nniEUR1Lf6953raAbdjanvjYV7J4ohGVSjiXpi9MWRi4JG0bWUoeGUY+6yrwN9xruaOFa6cTfsi2E/q783f48O2GZVA4GPs7aOgcAAycoA0AOJ7o0kNGn9WrpZQd7jSebcyGZXmjuIxfCMu0KRZGURbwvgghxhCL5r4m/TWcK8S1cmwp2EYTf72LDMmE02LvE30bYMJFQME09auhND7rKNNeY7pw/kDiwetKRLh4nCUN/xt1EEq7FlB3d3oR729B8tM4BwPAJygDQqtl80bz0e++ok8mNIhc/Mp5smuvjjYPUG01Y5jyMyqIdqYsiQjIRhbErKUMJr6qqSrlT+TFOEwaxFeziOmlhhE1XGQ3Wc3VdX7QQQuxy0MnfUw8UAqtfEnzpg9zPSyGok+o6MO4msnAtfkz4LYbWVeamR929AIBnEJQBoG1GLpHT6Xq19MKNe40nm6bY+3dHp3ea7j92+7WghaLIsaJIEqlDCdkLCaH7QqouHQp2kYVjedzCCJtOCaGylKPBdD1qSQghpghAbHW5q4z18HFSjWDK/bt+ljCcatxNAnVdN9fMZaIv36XgdAxn1jkAKIOgDACtmc0XzR/Orx1xMtmsV0tBLe4VOpIIW/TXq/FkY9dfeilHJzZFka9d+mGHIrzoT1WoG3WkOJLy91/BLoHw+556hE3XpBxVdqPrUeuadXWT8Jt28bnm0iiSx9m5B8c2zXXvDSGuVB04jbtJK2VIdSh/i93Y4AcA5RCUAaBNKV8Ow89oQ8+PfE4YAKAd70NXIBJI3E1GUSSxEPRIWcDIVhxRsOuvMMLmQwk/a1VVbxOOrRsJybQvYQBi6zjce7tEKPkJwn0kRSePXH/fproOrkLXExJJvG4NpauMbjIAUBBBGQBaMZsvThK/HIYf+bJeLS8cIe4TOpGkHIVAez6H7kDEl6qbjKJIS8KokCG23Few67G6rk8TjwbripQBg191PcojhL0+JvrmBx0bpXUZfl6eJsW5bL0rV8Jw6o2NLe0Iwa1Uo+OGEKazwQ8ACiIoA0Bys/nihd1nZKbQxb3Gk82s6UTi6AzGgZebyaRaRxVF2pVyBFPr99rEBbvSxgLlNOh1oKqq44QbBi5DCI58ThOOYOrS74a/558hhNk+Rf6yB2F9aVOqa/JU4K9VqZ4H+95V5tIoWAAoi6AMAG04M9KEjD6sV0svO9hHqGJ4Xo8nGwXuiMLIkBT38Q+KIu0KL/9TFTtfh+BKm1L9LKcKJe0J68CQRzClCpEJdHVA4lEm03APzu1KN5koUtyzWlsDEo7hvBL4a1fidavP4VfjNgGgMIIyACQ1my9miXb6wkPchKAWfCeMXHrpyAzSmRFMUaUoimysz3mEYlSqUTetdZUJBbsUBUIFuwzCCKZUXTmyqapqlnC849tQ7CSzECJJNcqkC0Vna2IEIYAZ+zppMyx3nCg4rftrBmEEU4qRnG8yBKdjEZQBgML8yQkHIDEv1cjpZL1aKiDwHSOXBm8adu168f5MocibIlB2qsCbVfO78c8EH+Bti793qTodWTfyaY79Pwb4M6VwGYqcdEezJs0Gej5ca/E070deR/x6zfilWUsd+lKsZ590K8qqOaf/SvAB3vZwXNtGN0EAKE9V17XTDkASs/nieIAvu+mPy/Vq2ef52CQ0nmwuEu7wpjv+fHs99cLzGaqqago67yJ/2abAa33OrKqqVOvg39oo4FdV9TWE4mJqCnZ9HhnQewmvy7uSr0Oh69HXRIGuPyvocVdVVSle8n6p69qIr4gS3L8+hK5cKT/zLFGgwlqWWVVV5wm6QDehkyRdZaqqOk204cUzIAAUyOglAJKYzRcvdJMhs77tYKIl48nmREimGOelH4AIUrwwtj53Q6rzkLygGgp2sUMyI9dmJwzpHKQaU/JBYZkW6SYTX+zn0zbCxymeBz9ayzohxX13WlVV30LxOhsBQIEEZQBI5SRRAQMe4tN6tfSig++MJ5tDhdCivBpPNjqXPFFVVSmKvJda7HdDOA+XCT7Mm9BJI6UUBbtPCnb5Jbwuc0hxnd7YjEDL3LPjix2UaWPkV4r1zFrWAeHZ51OCT9K37ixtjC8DADpGUAaA6GbzxWHCefzwMzeuP37gLNHubrpLMOrpUnQGcT66JdX5SB1Q0+lo2Hp/LqqqOkzUve6sruvfE3xduM9GgDC+cEw3Eb/wQei0lkSi4LRwarekuO/2amRbXdeCMgBQIEEZAFJQiCans/VqqYDAd0JnkdeOTHF0lXm62C+4N7rJdEs4H1cJPlSy4oiC3fANpKtMit8B3WRom3t2OrFHWqXsKpPiOVo4tUPCM1Ds++5BeGbrgxTPwgBADwjKABDVbL5QiCanzXq19NKNfWK3Oac/dJl6pKqqjhKEEazP3ZSi8J6yMKJgV4a+37NTXKe6ydA2HRbSiR1COkz4WWPf0y+FUzspxfNgXzYruLcCQKEEZQCIzS5HclIM517jyaYpgk4dnWK9Hk82KQsIQ5Qi6BB79zRxfA6dKmJKOQZCwa4AdV2fJ7guW1FV1YtEGwcEfmmboEw6sYMySQIJ4V4e+28oa1kH1XX9OfJIsFGPxi9Z6wCgUIIyAEQzmy/ejkajl44omVyuV0tFWL4TAhJCVLgGHid2weWTTgjdFM5Livtn9KKdgl1x+npuUhSsjQcjB8XjRMK9N2Yo4UWijxp7PbsJQUi6Kfbz4LSqqj5sVvA3CgAUSlAGgChm88UL3WTI7K0TwB6nCUbI0D/WiAcK3RBiB18FGbutF0EZBbviCMr8lzWU1gm4Jhcz/JZqw1Ls9cxa1m0lj18CAAokKANALArR5PRxvVraZct3xpNN82LujSNDc48aTzbCMg+TIoygMNJh4fzEHnOTYvSSgl1B6rpuullc9fAnjn2dbqyhZNDH372+iTp+KQSdY3PfLUjoXBb7d78PQRndswCgUIIyADzbbL5oWqm+cyTJ5CYEteA+rg129WVOfm6xAw6KIv3Qh3b7CnbliVpITk1HLgZEN5n0Yh/jqM9vYdxhzM1QgtP9EPu+24egjPUOAAolKANADFrWk9PJerX0YoPvhO4hrxwZdrweTzYpdtsOTewX2r0qdBcsxXmKFpRJULAbKdj1Qt/+zkjRScl1CsPU9S4Wsdczz4P9EPu+O03U7QgA4NkEZQB4ltl8caQQTUZX69VSUIvvhDCEbjLcR1eZn1MYKVOKYnzM0FXs6/JL5K9HAmH8UuyxYCmlGF1nDQUeoutd3KxlPZDovpsiRAoA8GyCMgA8l5ACOZ04+uzRXBtTB4d7CMr8QBiVE7Nrx6au66/RPyjR1XXddGe7ivx1YxbtBLjK1adzFbtQ7TolF9de/8Ref6xn5Spx/BIAUCBBGQCebDZfnCpEk9GX9WrpZRvfCd1khKjYx/ilH1MUKVvsMRCCMsQgKAMMTg+6RcXsHHwTOpXQD7Gvzdj3RgCAKARlAHiS2XyhEE1ON64/fuAsckcMhseuxv1iHxtFkX6Jfb5ihluiBmUU7HqlT2GR2CNpBWWA1oUOgzG55/ZLl4PTAADRCMoA8FQK0eR0tl4tjfLgO+PJpnkJ98aR4ScEZfaL3W1HYaRfYp+vKM+KVVW9iPzceRnxa5FYX0JNCQrLAl1ALrpjFSxBt6PYIVIAgCgEZQB4tNl8MVOIJqNNCGrBfc4dFR5AUGa/2ONtFHl7JMUYiEjhAdclVz04ArELywJdQC6x77s2ufRP1PtuCD0DAHSKoAwATyGkQE6n69Xyd2eAu8aTzZHdajzQy/Fk42Xt/WIel5u6rq3X/bOJ/IljhAdi/74q2PVPH8JNOnIBQ+G+S+xzFjt8BQDwbIIyADzKbL54qxBNRpfr1VLHEPYR4uMxvKy938uIX0uRt5+6WMzSUYY+FFljX6eChkAuUTtkpehYR3KelQCAwROUAeDBZvNFs6vo1BEjoxMHn/uMJ5u3kQv8DJ/xS+kp8vZTCbu+7WzvnxILdgrLQC6xR8nRP7Gf421SAAA6R1AGgMdoQgpTR4xMPq1XS7ua2EeIj8fysvaOqqpiF0Ws2f0UO0QSI5QWNdhW17WgTP/0IXhnpB/A964ck16K/RzvHgkAdI6gDAAPMpsvDnXzIKMb1x/7jCebUyE+nsBO2e85JkBXFTd6yagSIKOYz4Q6DAIA0EmCMgA8VFOIPnC0yOR0vVp6wcZ3xpPNCyEqnsioLrjf0Ivzlx34DDySLkAArbIJAZ0hAYDBE5QB4Kdm80XT7v6NI0Umm/Vqeebgs4cQH082nmyMX0pLNwRi0e0IAKAldV3H3qgUdYwmAEAMgjIAPISQAjm9dfS5z3iyaQqn7xwcnsGsfOgHO9sBoJ8EpwEA6CRBGQB+aDZfvDWegowu16ulF2vsc+rI8Ew6ygD0x1VB58qIMAAAAEhIUAaAvWbzxQvdZMhMNxnuNZ5sjIQjBh1lAPoj9hgIAAAAoFCCMgD8yMloNDpwhMjkw3q1/Orgs4duMsSgowwAAAAAQGEEZQC412y+OByNRu8dHTK50c2IfcaTTdNp6JUDRAQ6ygDQRYKcAAAAkJCgDAD7CCmQ08l6tdRen310kwEAhkxXT2AoBNMBAOgkQRkAvjObL45Go9FrR4ZMrtar5bmDz33Gk00Tkpk6OERix35ah0P+4QAA+CnP2wAAdJKgDAD30U2GnE4cfe4znmxeuD6IzI79tARl+klBi64ydhEAWlBVVexOQBfOGwDQNYIyAPzBbL5oitAvHRUy+bReLb1AYZ8zwQaA5Lo4ImET8WsJApHKOubXrarKtQpALu5BAMDgCcoA8B+z+aIpjJw6ImRy4/pjn/Fk07yoe+MAQVJRi7w6yhDR14hfS+CSVH6P/HW7GFoDynAZ8acUuAAAoJMEZQDYdap4QEZn69UyZiGMYTESDhKr6zp2kVdQpp+Ohv4DJhgnQGKFdlexhgJD4B1TP8W+73rXAwB0jqAMAN/M5ovmRew7R4NMNoIQ7DOebI5Ho9ErBwh6R5G3n0oIkdjd3j+xr8sUBTtduQDuYZRcL/XhvgsA8CyCMgBsnTsSZHS6Xi1jdzJgOISooD0xW+1PnbdeetnBD30R+esJIPRP7HOWomAX+1lWYRnIJfZ9Vye3/hl8h0EAAEEZAJpuMke6NZDR5Xq1FNTiXuPJ5lSxHfqrqiov2XukoB3fgjL904dzFjt8IygDDIXnwf6Jet+t6zp2+AoA4NkEZQAY6SZDZqdOAPcZTzbNzsMTBwdaFXt0iEJvv3Q1jBD7ulSw65/Y5yz2NdUUAWMHZaZVVenCAOSgk1vBwr0n5maVm9KPKQDQTYIyAIWbzRcnujWQ0af1amlnEfs0I5cOHB1olY4IZetqgMRIG2IXWVON/LyK/PVcq8AQCKj2S+x7T/RwKgBADIIyAAWbzRcvdPMgoxvXH/uMJ5vm5dwbB4jE7G78no4yZetkIStBu/6Dqqrsbu+JcK5iB/tTFe1ihw0Vl4HWJbjv6pDVL7HvPbHvjQAAUQjKAJTtVLcGMjpbr5ZemLDPmSNDC+xu/F7sY/JSYaQfwnl62eEPGzvYJoDQH7EDdzd1XafqKGNMGDAUm8g/h/WsPzo/7hAAIAZBGYBCzeaL5oXzO+efTDaCEOwznmyOR6PRKwcI2heKx7EDCcdOZS90vYAlgFCuPhXsYndheCVsCGTivluu2H+LC8oAAJ0kKANQLiEFcjpdr5apdvLSf9Yn2qKr1f0URsrU9UBT7ACC67I/+hSUSfG1hQ2BHGKvZ9ayHqiqKvp5SjDKCwAgCkEZgALN5gvdGsjpcr1anjsD3Gc82TQj4aYODi0RlLlf7JfZCiP9UFpHmWlVVbFH+hBZVVWHCUaCJQvKhK5csceVWEOBHGI/D7rv9kPse87VUA8UANB/gjIAZdKtgZxOHX3uM55smtECJw4OLRKUuV/swshBit2pxBMKV10PKaYIN7xN8DWJK8XakXoEROw19LXxS0AGOmSVKfY50k0GAOgsQRmAwszmC90ayOnLerX0ooR9mhDfgaNDiwRl7pGoPbrCSLd1PqRY13Xz+3oT+cu6Lrsvdpjppq7rvgVlRq5VoG2hQ1bsbiACqh0Wgu2x/x73/gcA6CxBGYCCzOYL3RrIzfXHvcaTTdPN4I2jQ8tSF0v77DLyZ3+jI0Kn9aUIbwxEQfo2dmnH5wRfU3EZyMF9tywpngcFZQCAzhKUASiLbg3k9GG9WurewD5GwtG2ze319HdHfa8UL7UVejuoqqq3PXo+TBFAEOLtrhRrRopr6A8SdWF4pbgMZJDiedB9t4NCoD32xpWrcE8EAOgkQRmAQszmC90ayOlGEIJ9xpNNs3PtlQNEywT3fkwgoRx9CjAlGWmj21Fnpbg229rZbg0Feq+u6xRrmftuN6W4x5wP4cAAAMMlKANQDiEFcjpZr5Z2ErGP9YkctAH/gbqum/Ekm8hftmm335cRP0WoquqoT0HFuq6/JrguD3o0eqoYYa2YRv55N2Fta0OK4vKbMI4KoE1fIn+vA8G/TuplFzcAgOcQlAEowGy+eKtbAxldrVdLO4m413iyOU1QCIOHEJT5OR0Rhq+P47BSXJenCb4mz5NirWht3U8UNhy5VoEMUtx3jePskDCGM0U4VQdPAKDTBGUABm42X7zwQpXMFEW513iyeeH6IJfb66mgzM+lCDm+Cl1MyCx0pujjWM4U1+U0FInogISdjtre2Z7iWtVVBmhbirXTfbdbUrwz1DUWAOg8QRmA4TvRrYGMvqxXS8Vo9jkNrbehbZeO+M/piDB4vSxguC6LkOJc3NR1PYSgzMi1CrSpruvfE4xfapxWVfXCycwrUTeZkbFLAEAfCMoADNhsvjjUrYHMXH/cazzZNOvTO0eHTLy4fbhUXWXsIs4odOx43eMfwe72gUrYTab1MaBh5ESKYGbTVWaW4OsC7JOkm5v3BXmFoFKK8OWlsUsAQB8IygAMm24N5PRhvVp6OcI+rResYIdOVw+XrCOCXcRZ9b0jRapuODp15JfqHOR67kj1fY20AFoTOnLdJPh+J54Hs0rVgdrf+gBAL/zJaQIYptl80ezGfOP0ksmNF/jsM55sUu0Wh4fY3F5P147UwzS7Qauq+pKg+8h2F7FgQstC15Rer8HhurxM8HM0XWVO67p2XWZQVdVxomvzKozsal1d1+fNNZWgENl05jqp69rzdoeEjkgprMP4G8jpPEFH0IPw3kBHt5ZVVZWqA3Uz6lBQBgDoBUEZgOHy0pScTtarpZe57OPFGTkZu/R4Z4nG9LyvqupzrgJ2icKu7aE8I54nClU0u9vPjQxoV+JrM/c131yr7xN83dOwhrpWOyAEolKc52YDwmHvDghDdJZodO6bcN/V8bFd54k6UHsXCQD0htFLAAM0my+a3TgvnVsyuVqvloIQ3Gs82aRq7wwP5eXtI4XCxSbRl3e/aFeqokjrwm7lFGMgDlyXWaR6PrjpQEDyzLU6bFVVzRKFZBpnusnQBSGU9yXRRzk3gqk9CTu46SwMAPSKoAzAwMzmiyHtFKafUrTvZQDGk80LY1bI7Or2emrn/dOk+t19GXbhk1goiqToDJRTqmfeb2Nt+nEI+m/oIYPw/VNeq9bQ/FIGlvxtT5ekuh6n/k5sRxi5lGrN+izYBwD0iaAMwPCcDGWnML30Zb1aapnMPqfWJzJTbHqi0L0jVVeZZgTTUfIfomCJiyI5perUMQpjbYw7SSx0EEh1bXZpZ3vKa9UamlFVVWcJu7l+UnSmS0KXwctEH+mdtawVKbsLCjsBAL0iKAMwILP54jDhbkz4mRvdZNhnPNkcJpppDw/VhfEbfZfy5fdnLfeT+jzEoGLiTh3G2rTjNGHIoDMjaxJfqyNraB6hU1fK51tFZ7rI82BPhQ5kKUYujUKwT+dOAKBXBGUAhsVOeXI6W6+WXoywj2IjuZ3fXk/tyn6GxF1lDgSZ0qiq6jxhEKELUnbqMNYmocQhgy51k9lKea02a+iFAnN7wsiwlM+3HxSd6aLEXWW+rWVOfHyhW0+qjXU2TQEAvSQoAzAQs/mi+aP3tfNJJhtBLfYZTzZHCXeuwUNZo+JIGRp4FUIdRFJV1dvRaPRmyMezhU4d70Ogg4haCBl0ppvMVvg8KQuJL93r2rEzMixVp64uBr1gV8rnwZeeB+MKoyRTBtI7d88FAHgIQRmA4fAigZxO16ulFyPsY30it0+311O7siMIXWVS7SJuvAnhDp4pBBF+K+Q4puzU0TgPx5MISg4ZhDX0KuG3eKPA3IrPiTt1nSg602Whq8ynhB/R82Ak4Z6bcgSnTVMAQG8JygAMwGy+aHYmTp1LMrlcr5ZeyHOv8WRjfaILjE6JK/Xx/E1x5HlCqKOY0QUtdJX5NhrMWJtoSg8ZpF7fhGUSCsc2ZafEqxCogq47TRxS9TwYR+n3XACAvQRlAHpuNl+8UAAkM7Ooudd4srE+0QW6yUTWwi7iUSiOHLX+ww1AaK9/kXDncCfVdX0adjWn0oQ+L4RlnqeFkMFl10MGdV2vR6PRx8TfRjeGBML1m3qcnStrWtUAACAASURBVPNGL9R1/bWFTiJnOro9XUv33JQjnQAAkhKUAei/09IKIXTKx/VquXZK2MP6RG43wnzJnCTeRTwKHTwURx6hhfb6XZe6wPwyHF+eoKWQQV/W/NTBrlEIHAosR9LS9fsxBKmgF1oIqR6EkKrnwUdqYc26EewDAPpOUAagx2bzRfOy4J1zSCY3uoWwz3iyObQ+0QFnt9dTrcATCC3WU78cVxx5hBCSuUjcXr/TQrejL4k/4ytjbR4vdDdJHTL40JeQQUtraOO96/X5WgrJbPxtRU95HuyYltas09BVCACgtwRlAPotdZtb+JHT9WqpAM0+ijLktrm9nio4JRRaracOJWyLI8dZf9iO2xm3VGxIZsfbFrodvRE+eLgQkvkt8be56tvfRiHY9aGFb+V6fYaWCs6NtyFABb0S1rLU4+SEZR6opTWrGbnkfSQA0HuCMgA9NZsvjhPPGoYfuVqvll6McK/xZHNkfaIDtAJvRxuhhKY48o9QbOeOUDRaC8n8nxY7dTThg8+hkw97VFV10kJIZtTXkEEYW3LVwrdqrtcL1+vDNceq+R1vKSTzMYQNoK/aGCcnPP0TLYVkjFwCAAZDUAagv4QUyMmLEX7ErmVy+3J7PVVwakEoTLdVsPitqipdgnZUVXUUOskcdOZDdUBL3Y4ar0PRTvjgHqFg9/cWvlVvRi7tcdxC4HAUQsw6MjzATpeu1y18uysjl+i7Fp8HhafvEYJ9Fy12vzJyCQAYBEEZgB6azRfNi7Spc0cmH9arZZ+LESQ0nmxOrE9kZpdjy1ocH9J4r4vH/wlFon8Kyez1toXd7aPQyUf4YEco2LU1ruYydGXprVBwbCtwuL1e3Sf3CN0q2uzSZeQSgxACi7+29LM04Wmbx/4Y7Gujo+unEEYGABgEQRmAnpnNF80fwSfOG5k0I5fseORe48nmhR2xdMDx7fVUwalloVB92dJ3fV1yMGEnhNDGOJvearnb0TZ8cDTAQ/koIcTW1q72mxbPcVIhcNhWgfkgFJnPhQ7/K6ytTeH9Hy0GEH/peTck+IO6rs9a6ujWeFf6SLnw3NFWsO+qrmshSwBgUARlAPrn1M5hMvJihB+xPpHbRyOXsmprfMhoJ5hQVHg4hIPaCiH0Xsu725v73z9LHg8WCnZfW+zEcTSkThyhwPypxW/ZrCNrAa8/FJvftfhtm84MxpUyRG11dBuFLipfQyeoooTnjbY6CzbP98XfKwCA4RGUAeiR2XxxpDBCRkYusdd4sjlsubgAd13dXk91XMsoFKyPWgzLNIWBv4fdxIedPCgRhYLIv1oMIQxChvDB+1KuyV0tF+xGQ+3EEXbrX7X4Lach4HVWYleGnS4y/2x5dOiVLrEM1U5HtzafB/9RyjrWhKarqmruf+9b/LaDCqYCAGwJygD0i5Em5HJp5BI/YUcsOdnl2BGhcN128e9V6IowyPtUpoLI0Jy0HD7YXpODL4Rnuj4/DrwTx1HL1+sohJ2brgzFdI8Mv59fMwS9v40MU3RmyMLzYNvrybtw7x1sd5lMoWkj4gCAwRKUAeiJ2XzxNrx0h7bdhB1hcK/xZHNsfSKzo9vrqYJTR4QC9i8tf5qD0MljMO33Q6eD8wwFkbYL9Mll6HY0utPxaJbpR08mXJ85CnbNuJpBB5B2rte2fxeba/a3cM0ONnzahIGae0Xz+5lpZGjTmeFrhu8Lrarr+nOG58Fp6C4zqHtvsyaHdavt0PSvRsQBAEMmKAPQA7P5omkfe+ZckcnxerVUgOZHrE/k9Mvt9dQux44JL9XbHHeztVsg6WWhdyeA8DXDyM3BjgPJFJYZhSDpv5rQ01DGMYWuIzm6HF2F0USDl2F0ya5XYRzTYAIzYV3dBmR+a3nM0i6dGShKxufBQdx7Q0DmIsN4uFEIpvo7HwAYNEEZgH44ybTbDX5Zr5YXxR8F9hpPNqcZiw3w4fZ6apdjR4WCdo7iyOhOobcXhfWmkLMTkHmf4dnvaugjzEKBOlfHoTfbEWFN0T7TZ3iWnYJdjqDB4K/Pu0LXkRzhrq3tOrru60imMBrsPKyrOQMyoxCS8cxCcTI/Dzb33n+HwExv7iF3AjI5Ord+KiWYCgCUTVAGoONm88Vhht2a0Pi0Xi29zGWv8WTzYqidB+iFT7fX01OnqtsyF0dGobjwWxjJdNrFXcXNqKiqqprxBP/OFJAZbUMIoYvFoNV1fZFhFMTWQTjH/xuKdr0YCxE6ceQs2BVzfd4Vwl1NcXeT8WO8DOvo71VVnXX9ug2hw5Mm4BNGg73pwKaTD0IylKwDz4NvdoN/XQys7nS+Wme8346EZACAkgjKAHSfF2rk8GW9Wno5ws+c6XZFJk1IxhrVEx0ojoxCF4H3YVfxOhRRs4Vmwk7hszAG5B+j0eh1rs9SYgghFKxzhWW23oSxEJ0s2m07HO2MqslVsCs2JLMVwjKzcCxyap753oXr9mtYwzrRoWFnTV2H0OHfQ8CnC5qis2AvxevI8+DLcE9rAqufc99/QzjmOHS++t/w2XKuXUIyAEBR/uR0A3TXbL44yvhSmnI1L+G9HOGHxpPNLBT5oG1CMj3UvHSvqmrUkXXjZSii/r2qqqZLw8X2Xxh1El3owHC0868rIcOSO3Wch2vyt8wfZVu0azp2fBmNRp9TXos/EsJjx+E5sAshg+JDMlvNMQihlIuOnJtpCM28q6rqZmcdXYeuTcmE9fQwrKezjv+9rOgMOzr2PPg6/Gvuv1d3ngeT3XfCWr7916X1y3oFABRHUAag23SToW3fChLr1bL4ggQ/deYQkYGQTI91rDiyNQ2f59tnCgXfdfj3eyiYNH4PXR32CsXbF+HftpB72OEibvEhhA6FZba2RbvRnRDX+mfX31OEa3Y3xDVt6ed8iOKvz7t2wjLnmbtQ3XVw59odhVFRX3fW0AetpSGste32tV1LR+H6fNGhLjEP8UEnGfheR58HX4Z/TQBwew++bw37+pAg685atvtMOOvwGiYkAwAUSVAGoKNm88VJx15WM3xCMjzIeLI51u2KDIRkBiAUR37fFiI66CCsb9s17v32I4aizlAIIQQdDMts3Q1xjcJ5+30nyLX93w+xHZGzDXR1+T5+2XS2cX1+LxyT7ZiOLnf2m4Z/Q19L9/kljHgD7tHRsMyuktYwoT4AoFiCMgAdNJsvmpfX/lClTUIyPIZuMrRNSGZA6ro+qapq3cFgQinsGr4jhGV+D506ujIW6z7bnehDDqu6Ph8gFJkvrKOd03QlOxGSgZ8L69g6jMMkD6E+AKBo/1P6AQDoqLOOv6RnWIRkeLDxZHOq2xUt+yAkMzzhpfzfQlGR9nwQQrhfXdefQ9cV12Q+rs9HCOvoX12znXETOnUpOsMD1XXdvPv6xTrWuuZ4/9V6BQCUTlAGoGNm88Ws4220GZYvQjI81HiyabpdnThgtOiX2+upDmsDtRNMuCr9WLTgJuwa9vv0A3Vdr8NoItdku1yfT1TXddNV5jCMqyKfZs04DGsI8AghrNE8D24ct1Zsx29eFPCzAgD8kKAMQPcYaUJbPq1Xy2MhGR5Btyva0hRN/3J7PbXLceBCUfEoBDdJ40qXg4er6/pruCY/9eUz99zG9fk8dV3/Xtd1c81+6PPP0WMf67qeNeeh9AMBT7UTVBX6S+tTuOcK9QEAxRsJygB0y2y+OB6NRq+cFlrwy3q11NqeBxtPNrpd0ZamqD+7vZ56gVuIUORtnoF+Lf1YJKAg8gThmnzrmkyuCcjNXJ9xhI48f9WVoTVNqPdvdV3rtggRCP0l1axXvzbPNkJ9AAD/JSgD0BGz+eKFbjK0oHlx/pf1amnXLo9lfaIN34r6t9fTr452eeq6btaZvxh7E8V2lI2CyDPsXJOCB3FtC3bHrs+4wiiNJtz8cUg/VwddhlFLn0s/EBBbCP2598az7Szo73kAgDsEZQC6o9mJNnU+SOjbrt31amnXLo8ynmx0uyK1b0X92+vp29vrqaJpwZrOEs0IC7uJn+UydOkQio1gZxyE4EEcCnaJha4MJ6G7jOBhXNsQ4pGQF6Tj3hvNhzAazjsgAIB7CMoAdMBsvjgMQRlI4SaMWjper5Ze6PIUilmkdBW6yCjq8x87u4kvHZUH23bpaAq4ujJFdCd4YIf70ynYtajpLhOCh7+G9YHn+RS6yHhegRYI/T1Lc7z+Ep6nAQDYQ1AGoBuaP14PnAsSaLrIHBq1xFONJ5tT3a5I6MPt9XR2ez1VNOU7obvMURP2VOT9qW0BV7AxIWNtnqwJvP1ZwS6PsC4chnWCx2uu378aZQd57IT+Pnge/KltaFooFQDgAQRlADKbzRdNAeiN80BkzW7nv+oiw3OMJ5sXul2RyLddjrfXU0VTfirs3j9UILmXAm7Ldna463j0c83z6N90OcovXLdvm8CSwMyD7V6/Fz35zDBYIWwp9LffR6FpAIDHEZQByM8fscT0bQfRerVsush4octznel2RWTf1ihdZHisUOQ9Dd08FEj+G5BRwM1kp+PR34xj+s5NGLPUFOw+d+yzFa0JLAnM/FTz+/yL6xe6R+jvXp9C17YToWkAgMcRlAHIaDZfNH/gv3QOiOAm7LRvAjLCVzzbeLKZ6XZFZN92Od5eT61RPNmdIm+JHWYEZDqmKaQ3BfUwIqz0wMx/nkeNWeq2ewIzunX93/r6txCQMbYWOkzo75ttQOatrm0AAE8jKAOQyWy+eKGbDBE0BZlfQ0Dm1JglIrI+Ecu3l7i319OT2+upNYooQoFk24L/1wICCp8EZLqtKazvBGZKG8m02Q3I2NHeHzvF5u21e1XYIbjZKTYf6SAD/bKzhv2/cB8a+vPgNpAqIAMAEMGfHESAbE6NNOEZvoxGo/P1aullLtGNJ5vj0Wj0ypHlGZqXuM1u7LPb66kXuCQTCvJNsO+sqqpmDE5TLDkeyDPWJvxsnxVC+iN0ojjfuR6H3J2tCQSd677Rf2Et3V67s3Dtvh3w36tfwtrq2oUBCGtY847ttKqq47B+vR7QubVmAQAkUNV17bgCtGw2XzQ79v7tuPNIV+EF9uf1aqlgRjLjyaa5vqaOME9wtS3s6x5DLlVVvQhhmeMeFkmacMznED5Y5/gAoUges6tY0UGKnevxZCAjV7Nfo7QnFJyPwjXc52fDJsB7Ea7dz0PtelRVVcyOY0JwLXDO0un58+Bo9/2PwPR/VVW1DXLGdOKZBgDKJCgDkMFsvrjQrYEHuty+0BWOoQ3jyabZiffeweYR/lM0vb2eesFI5+wUeo86GlS43BZwvaQftqqqDkPB7m3PQjObnWtUN8NChRDd0c6/rneb2a6tF0bWQdlCaGYb+jvqaPBvN9B3IRwDAJCeoAxAy2bzRfNH+T8dd/bYvtBtCmUX69VSRwZaM55smheIX42F4wG2Ib4L4Rj6ZKdQMtv5b5tr3ibc49eKt2XrQdHucqdYZ53nOyE4s/sv50aQy/AM+219tbYCPxKCq7vPgm0/D45Cx5jdZ0L3WgCAlgnKALRsNl8YacIovBT5uvNi5Ot6tfRihKzGk03T2vmNs8Adfyjs315PFZ8YlBBY2BZJXuz898UTu3409/jfw791uN9/DcVbAVjudSfEtf3X1t8MQgZEEYrPh/esp6NnBGk24foc7ayn/1lfdV0AYvjB+vXU58FRuL+Odp8Fm/XLfRYAoBsEZQBaNJsvmj+0z+75jjl2rxDXtih213cvc41QoovGk03zUvDfTk5xLu/8wNuXtt9e4grFwB9VVXV09/9T7CCl0LVjt2C3LeTt2hdA2A0YjLZru5ABXbJzjQsTAr2wE6r5A8+EAAD9IigDAAAAAAAAAEAR/sdpBgAAAAAAAACgBIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAAAAAUQVAGAAAAAAAAAIAiCMoAAAAAAAAAAFAEQRkAAAAAAAAAAIogKAMAAAAAAAAAQBEEZQAAAAAAAAAAKIKgDAAAAAAAAAAARRCUAQAAAAAAAACgCIIyAAAAAAAA8P/ZtQMBAAAAAEH+1oNcHAEAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAItovGAAAGZ1JREFUC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAAAAC6IMAAAAAAAAAAALogwAAAAAAAAAAAuiDAAAAAAAAADUrh0QAACAMAx6/9T2cJADIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAAIEGUAQAAAAAAAAAgQZQBAAAAAAAAACBBlAEAAAAAAAAA4L9tB4UJ1ZntS/czAAAAAElFTkSuQmCC"
)
    st.markdown(f"""
    <div class="sb-brand-row">
        <img class="sb-logo sb-logo-dark"  src="data:image/png;base64,{_LOGO_DARK}"  alt="Adchor" />
        <img class="sb-logo sb-logo-light" src="data:image/png;base64,{_LOGO_LIGHT}" alt="Adchor" />
    </div>
    """, unsafe_allow_html=True)

    # ── Progress bar ──────────────────────────────────────────────────────────
    _prog_pct = min(100, int((st.session_state.step - 1) / 3 * 100))
    st.markdown(f"""
    <div class="sb-prog-wrap">
        <div class="sb-prog-label">Progress <span>{_prog_pct}%</span></div>
        <div class="sb-prog-track"><div class="sb-prog-fill" style="width:{_prog_pct}%"></div></div>
    </div>
    """, unsafe_allow_html=True)

    # ── Step indicators ───────────────────────────────────────────────────────
    _step_info = [
        ("Upload Brief",  "PDF & transcript"),
        ("Review SOW",    "AI draft content"),
        ("Pricing",       "Line items & total"),
        ("Download PDF",  "Export & send"),
    ]
    _steps_html = ""
    for i, (name, desc) in enumerate(_step_info, 1):
        if i < st.session_state.step:
            _state = "step-done"
            _icon  = "&#10003;"
        elif i == st.session_state.step:
            _state = "step-active"
            _icon  = str(i)
        else:
            _state = "step-pend"
            _icon  = str(i)
        _steps_html += f"""
        <div class="sb-step-item {_state}">
            <div class="sb-step-icon">{_icon}</div>
            <div>
                <div class="sb-step-name">{name}</div>
                <div class="sb-step-desc">{desc}</div>
            </div>
        </div>"""
    st.markdown(_steps_html, unsafe_allow_html=True)
    st.divider()

    # API Key -- loaded once from Streamlit Secrets or environment, never shown to users
    def _load_api_key():
        try:
            return st.secrets["ANTHROPIC_API_KEY"]
        except Exception:
            pass
        return os.environ.get("ANTHROPIC_API_KEY", "")

    if "api_key" not in st.session_state or not st.session_state.api_key:
        st.session_state.api_key = _load_api_key()

    if not st.session_state.api_key:
        st.markdown("""
        <div style="background:rgba(255,60,60,0.07);border:1px solid rgba(255,60,60,0.25);
            border-radius:8px;padding:10px 14px;font-size:11px;color:#ff6b6b;">
            &#9888; API key not configured.<br>
            <span style="color:#5a6278;">Contact your Adchor admin.</span>
        </div>""", unsafe_allow_html=True)
    st.divider()

    # MODIFIED: use centralised reset helper
    if st.button("&#8635; Start New SOW", use_container_width=True):
        _reset_sow_state()
        st.rerun()

    # ── Active Client chip ────────────────────────────────────────────────────
    _client_name = (st.session_state.sow_data or {}).get("client_name", "") or "No active client"
    st.markdown(f"""
    <div class="sb-client-chip">
        <div class="sb-chip-label">Active Client</div>
        <div class="sb-chip-row">
            <div class="sb-chip-dot"></div>
            <div class="sb-chip-name">{_client_name}</div>
        </div>
    </div>
    <div style='margin-top:14px;color:#141828;font-size:9px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;'>Adchor&#8482; &middot; 2026</div>
    """, unsafe_allow_html=True)

    # ── SOW Library panel ─────────────────────────────────────────────────────
    st.divider()
    _sow_lib = st.session_state.get("sow_library", {"sows": []})
    _saved   = _sow_lib.get("sows", [])
    _lib_label = f"🗂 SOW Library ({len(_saved)})" if _saved else "🗂 SOW Library"
    with st.expander(_lib_label, expanded=False):
        if not _saved:
            st.caption("No SOWs saved yet. Complete a SOW and click 'Save to Library'.")
        else:
            for _si, _entry in enumerate(_saved):
                _cn  = _entry.get("client_name",  "Unknown")
                _pn  = _entry.get("project_name", "Untitled")
                _upd = _entry.get("updated_at",   _entry.get("created_at", ""))[:10]
                st.markdown(
                    f"<div style='font-size:12px;font-weight:600;color:#e0e3f0;'>{_cn}</div>"
                    f"<div style='font-size:10px;color:#5a6278;margin-bottom:4px;'>{_pn} &middot; {_upd}</div>",
                    unsafe_allow_html=True,
                )
                _lc1, _lc2 = st.columns([2, 1])
                with _lc1:
                    if st.button("Load", key=f"lib_load_{_si}", use_container_width=True):
                        _pending = {
                            "ta_why_now":          (_entry.get("sow_data") or {}).get("why_now", ""),
                            "ta_project_overview": (_entry.get("sow_data") or {}).get("project_overview", ""),
                            "ta_core_message":     (_entry.get("sow_data") or {}).get("core_message", ""),
                            "ta_assumptions":      "\n".join((_entry.get("sow_data") or {}).get("assumptions", [])),
                            "ta_out_of_scope":     "\n".join((_entry.get("sow_data") or {}).get("out_of_scope", [])),
                            "ta_timeline_notes":   (_entry.get("sow_data") or {}).get("timeline_notes", ""),
                        }
                        for _ki in range(30):
                            st.session_state.pop(f"sd_{_ki}", None)
                            st.session_state.pop(f"st_{_ki}", None)
                        st.session_state.sow_data       = _entry.get("sow_data") or {}
                        st.session_state.pricing_items  = _entry.get("pricing_items") or []
                        st.session_state.sow_discount   = _entry.get("sow_discount") or 0
                        st.session_state.step           = 2
                        st.session_state.ai_reword_result = ""
                        st.session_state["_pending_widget_update"] = _pending
                        st.rerun()
                with _lc2:
                    if st.button("🗑", key=f"lib_del_{_si}", use_container_width=True, help="Remove from library"):
                        _saved.pop(_si)
                        _sow_lib["sows"] = _saved
                        st.session_state.sow_library = _sow_lib
                        save_sow_library(_sow_lib)
                        st.rerun()
                st.markdown("<hr style='margin:6px 0;border-color:#1e2540;'>", unsafe_allow_html=True)

        st.divider()

        # ── Download individual SOW as PDF ────────────────────────────────────
        if _saved:
            st.markdown("**Download SOW as PDF**")
            _pdf_sow_options = {
                f"{s.get('client_name','?')} — {s.get('project_name','?')} ({s.get('updated_at', s.get('created_at',''))[:10]})": i
                for i, s in enumerate(_saved)
            }
            _pdf_sow_sel = st.selectbox(
                "Select SOW",
                options=list(_pdf_sow_options.keys()),
                key="lib_pdf_sel",
                label_visibility="collapsed",
            )
            if _pdf_sow_sel:
                _pdf_idx   = _pdf_sow_options[_pdf_sow_sel]
                _pdf_entry = _saved[_pdf_idx]
                _pdf_sd    = _pdf_entry.get("sow_data") or {}
                _pdf_pi    = _pdf_entry.get("pricing_items") or []
                _pdf_disc  = float(_pdf_entry.get("sow_discount") or 0)
                _pdf_total = sum(
                    (it.get("qty", 0) * it.get("rate", 0)) for it in _pdf_pi
                ) * (1 - _pdf_disc / 100)
                try:
                    _pdf_sow_bytes = build_sow_pdf(
                        sow_data=_pdf_sd,
                        pricing_items=_pdf_pi,
                        total=_pdf_total,
                        discount=_pdf_disc,
                    )
                    _pdf_cn  = _pdf_sd.get("client_name", "SOW").replace(" ", "_")
                    _pdf_pn  = _pdf_sd.get("project_name", "").replace(" ", "_")
                    _pdf_fn  = f"SOW_{_pdf_cn}_{_pdf_pn}.pdf" if _pdf_pn else f"SOW_{_pdf_cn}.pdf"
                    st.download_button(
                        "⬇ Download PDF",
                        data=_pdf_sow_bytes,
                        file_name=_pdf_fn,
                        mime="application/pdf",
                        use_container_width=True,
                    )
                except Exception as _pdf_ex:
                    st.error(f"PDF build failed: {_pdf_ex}")

            st.markdown("**Batch Excel Report**")
            # Build Excel with one row per SOW for reporting
            try:
                import io as _io_xl
                import openpyxl as _opxl
                from openpyxl.styles import Font as _XLFont, PatternFill as _XLFill, Alignment as _XLAlign
                _wb = _opxl.Workbook()
                _ws = _wb.active
                _ws.title = "SOW Library"
                _headers = [
                    "Client", "Project", "Account Lead", "Date", "Deadline",
                    "Budget", "Pricing Items", "Discount (%)", "Total (calc)",
                    "Status", "Created", "Updated",
                ]
                _hdr_fill = _XLFill(fill_type="solid", fgColor="014BF7")
                _hdr_font = _XLFont(bold=True, color="FFFFFF")
                for _ci, _h in enumerate(_headers, 1):
                    _cell = _ws.cell(row=1, column=_ci, value=_h)
                    _cell.fill = _hdr_fill
                    _cell.font = _hdr_font
                    _cell.alignment = _XLAlign(horizontal="center", vertical="center")
                for _ri, _s in enumerate(_saved, 2):
                    _sd   = _s.get("sow_data") or {}
                    _pi   = _s.get("pricing_items") or []
                    _disc = float(_s.get("sow_discount") or 0)
                    _tot  = sum((it.get("qty", 0) * it.get("rate", 0)) for it in _pi) * (1 - _disc / 100)
                    _row  = [
                        _s.get("client_name", ""),
                        _s.get("project_name", ""),
                        _sd.get("account_lead", ""),
                        _sd.get("date", ""),
                        _sd.get("deadline", ""),
                        _sd.get("budget", ""),
                        len(_pi),
                        _disc,
                        round(_tot, 2),
                        _s.get("status", "draft"),
                        _s.get("created_at", "")[:10],
                        _s.get("updated_at", _s.get("created_at", ""))[:10],
                    ]
                    for _ci, _val in enumerate(_row, 1):
                        _ws.cell(row=_ri, column=_ci, value=_val)
                # Auto-size columns
                for _col in _ws.columns:
                    _max_w = max(len(str(_c.value or "")) for _c in _col)
                    _ws.column_dimensions[_col[0].column_letter].width = min(_max_w + 4, 40)
                _xl_buf = _io_xl.BytesIO()
                _wb.save(_xl_buf)
                _xl_buf.seek(0)
                st.download_button(
                    "⬇ Download Excel Report",
                    data=_xl_buf.getvalue(),
                    file_name="SOW_Library_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as _xl_ex:
                st.error(f"Excel export failed: {_xl_ex}")

        st.divider()
        st.download_button(
            "Download Library Backup (.json)",
            data=json.dumps(_sow_lib, indent=2),
            file_name="sow_library.json",
            mime="application/json",
            use_container_width=True,
            help="Downloads sow_library.json — commit to GitHub to persist across deploys",
        )

        # ── Import SOW from PDF ───────────────────────────────────────────────
        st.caption("Import a previous SOW PDF to pre-populate fields:")
        # Use a rotating key so the uploader resets after a successful extraction
        _pdf_uploader_key = f"sow_pdf_import_{st.session_state.get('sow_pdf_upload_count', 0)}"
        _pdf_up = st.file_uploader(
            "Upload a previous SOW PDF",
            type=["pdf"],
            key=_pdf_uploader_key,
            help="Upload a previously generated Adchor SOW PDF — AI will extract the content and pre-populate all fields for editing.",
        )
        if _pdf_up:
            if not st.session_state.get("api_key"):
                st.error("API key not configured — cannot extract SOW fields.")
            else:
                _sow_extract_result = None
                _sow_extract_error  = None
                with st.spinner("Reading PDF and extracting SOW fields…"):
                    try:
                        from pypdf import PdfReader
                        import io as _io
                        _reader = PdfReader(_io.BytesIO(_pdf_up.read()))
                        _pdf_text = "\n".join(
                            p.extract_text() or "" for p in _reader.pages
                        ).strip()

                        if not _pdf_text:
                            _sow_extract_error = "Could not extract text from this PDF. Make sure it's a text-based PDF, not a scanned image."
                        else:
                            import anthropic as _ant
                            _client = _ant.Anthropic(api_key=st.session_state.api_key)
                            _msg = _client.messages.create(
                                model="claude-sonnet-4-6",
                                max_tokens=4096,
                                messages=[{
                                    "role": "user",
                                    "content": (
                                        "You are extracting structured data from a Statement of Work (SOW) document "
                                        "for a creative agency called Adchor.\n\n"
                                        "SOW document text:\n"
                                        f"{_pdf_text[:8000]}\n\n"
                                        "Extract all available fields and return a single valid JSON object with this exact structure. "
                                        "Use empty string or empty array for fields not found:\n"
                                        '{"client_name":"","project_name":"","account_lead":"","deadline":"","budget":"","date":"",'
                                        '"project_overview":"","why_now":"","objective":"","audience":"","core_message":"",'
                                        '"scope_sections":[{"title":"","description":"","services":[],"deliverables":[]}],'
                                        '"assumptions":[],"out_of_scope":[],"timeline_notes":"","review_rounds":"2","adchor_notes":""}\n\n'
                                        "Return raw JSON only — no markdown, no code fences, no commentary."
                                    ),
                                }],
                            )
                            _extracted_text = _msg.content[0].text.strip()
                            if _extracted_text.startswith("```"):
                                _lines = _extracted_text.splitlines()
                                _extracted_text = "\n".join(
                                    _lines[1:-1] if _lines[-1].strip() == "```" else _lines[1:]
                                ).strip()
                            _sow_extract_result = json.loads(_extracted_text)
                    except json.JSONDecodeError:
                        _sow_extract_error = "AI could not parse the extracted fields. Try again."
                    except Exception as _e:
                        _sow_extract_error = f"Extraction failed: {_e}"

                if _sow_extract_result:
                    st.session_state.sow_data = _sow_extract_result
                    for _wk in ["ta_why_now", "ta_project_overview", "ta_core_message",
                                "ta_assumptions", "ta_out_of_scope", "ta_timeline_notes"]:
                        st.session_state.pop(_wk, None)
                    for _wi in range(30):
                        for _wk in [f"st_{_wi}", f"sd_{_wi}", f"ss_{_wi}", f"del_{_wi}"]:
                            st.session_state.pop(_wk, None)
                    # Rotate uploader key so it resets on next render (prevents re-extraction loop)
                    st.session_state.sow_pdf_upload_count = st.session_state.get("sow_pdf_upload_count", 0) + 1
                    st.session_state.step = 2
                    st.session_state.ai_reword_result = ""
                    st.rerun()
                elif _sow_extract_error:
                    st.error(_sow_extract_error)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 -- INPUT  (unchanged)
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.step == 1:
    st.markdown("""
    <div class="hero-bar">
        <div class="hero-eyebrow">Step 1 of 4</div>
        <div class="hero-title">Upload Brief &amp; Transcript</div>
        <div class="hero-sub">Upload your filled creative brief PDF and call transcript &mdash; our AI extracts everything and drafts the SOW</div>
    </div>
    """, unsafe_allow_html=True)

    col_brief, col_trans = st.columns(2)

    with col_brief:
        st.markdown("#### Creative Brief PDF")
        brief_file = st.file_uploader(
            "Upload filled Adchor Creative Brief (.pdf)",
            type=["pdf"],
            label_visibility="collapsed",
        )
        if brief_file:
            fields = extract_brief_fields(brief_file.read())
            if '_error' in fields:
                st.error(f"Could not read PDF: {fields['_error']}")
            elif fields:
                st.success(f"✓ {len(fields)} fields extracted from brief")
                with st.expander("Preview extracted data"):
                    for k, v in fields.items():
                        if v:
                            st.text(f"{k.replace('_',' ').title()}: {v}")
                st.session_state.brief_fields = fields
            else:
                st.warning("No filled fields found -- brief may not be completed yet.")
                st.session_state.brief_fields = {}

    with col_trans:
        st.markdown("#### Call Transcript")
        source = st.radio("Source", ["Paste text", "Upload .txt"], horizontal=True, label_visibility="collapsed")
        if source == "Paste text":
            transcript = st.text_area(
                "Paste Zoom / Teams transcript",
                height=220,
                placeholder="Paste the full call transcript here. The AI will extract decisions, requirements, and context...",
                label_visibility="collapsed",
            )
        else:
            tf = st.file_uploader("Upload .txt transcript", type=["txt"], label_visibility="collapsed")
            transcript = tf.read().decode("utf-8") if tf else ""
            if transcript:
                st.success(f"✓ {len(transcript.split())} words loaded")
        st.session_state.transcript = transcript

    st.divider()
    col_skip, col_gen = st.columns([1, 1])
    with col_skip:
        if st.button("Skip AI -- Enter Manually", use_container_width=True):
            st.session_state.sow_data = get_empty_sow()
            st.session_state.step = 2
            st.rerun()
    with col_gen:
        has_input = bool(st.session_state.brief_fields or st.session_state.transcript)
        has_key   = bool(st.session_state.get("api_key"))
        disabled  = not has_input
        if st.button("Generate SOW", use_container_width=True, type="primary", disabled=disabled):
            if not has_key:
                st.error("AI generation is not available -- API key not configured. Contact your Adchor admin.")
            else:
                with st.spinner("Reading your brief and transcript — drafting the SOW..."):
                    try:
                        sow = generate_sow_content(
                            st.session_state.brief_fields,
                            st.session_state.transcript,
                            st.session_state.api_key,
                        )
                        st.session_state.sow_data = sow
                        st.session_state.step = 2
                        st.rerun()
                    except Exception as e:
                        import traceback as _tb
                        st.error(f"Generation failed: {e}")
                        with st.expander("Full error details"):
                            st.code(_tb.format_exc())
        if disabled:
            st.caption("Upload a brief or paste a transcript to continue.")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 -- REVIEW & EDIT CONTENT
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.step == 2:
    st.markdown("""
    <div class="hero-bar">
        <div class="hero-eyebrow">Step 2 of 4</div>
        <div class="hero-title">Review &amp; Edit SOW Content</div>
        <div class="hero-sub">All fields are editable &mdash; refine the AI draft before moving to pricing</div>
    </div>
    """, unsafe_allow_html=True)

    sow = st.session_state.sow_data or {}

    # ════════════════════════════════════════════════════════════════════════════
    # AI INSTRUCTION BOX — whole-SOW editor via free-form Claude instruction
    # ════════════════════════════════════════════════════════════════════════════
    st.markdown(
        '<div class="ai-instr-panel">'
        '<div class="ai-instr-label">AI SOW Editor</div>'
        '<div class="ai-instr-sub">'
        'Tell the AI what to fix and it will update all relevant fields automatically. '
        'You can still edit anything manually after the update.'
        '</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    _ai_sow_instr = st.text_area(
        "AI instruction",
        key="ai_sow_instruction",
        height=85,
        label_visibility="collapsed",
        placeholder=(
            'e.g. "The client name is wrong — update it to Acme Corp."  ·  '
            '"Remove brand audit from services and add social media management."  ·  '
            '"Rewrite the project overview to sound more strategic and concise."  ·  '
            '"The assumptions are incomplete — add: client provides all assets by Day 5."'
        ),
    )

    _upd_col, _ = st.columns([2, 3])
    with _upd_col:
        _ai_sow_clicked = st.button(
            "Update SOW",
            type="primary",
            use_container_width=True,
            key="btn_ai_sow_update",
        )

    if _ai_sow_clicked:
        if not _ai_sow_instr.strip():
            st.warning("Please type an instruction before submitting.")
        elif not st.session_state.get("api_key"):
            st.error("API key not configured. Contact your Adchor admin.")
        else:
            with st.spinner("Updating your SOW…"):
                try:
                    _updated_sow = _ai_update_sow(
                        sow,
                        _ai_sow_instr.strip(),
                        st.session_state.api_key,
                    )
                    # Write updated SOW back as the source of truth
                    st.session_state.sow_data = _updated_sow
                    # Clear all field widget keys so they reinitialise from new sow_data
                    for _wk in _SOW_FIELD_WIDGET_KEYS:
                        st.session_state.pop(_wk, None)
                    for _wi in range(30):
                        for _wk in [f"st_{_wi}", f"sd_{_wi}", f"ss_{_wi}", f"del_{_wi}"]:
                            st.session_state.pop(_wk, None)
                    # Clear the instruction box
                    st.session_state.pop("ai_sow_instruction", None)
                    st.session_state["ai_sow_update_status"] = (
                        "✓ SOW updated! Review the fields below — you can still edit anything manually."
                    )
                    st.rerun()
                except Exception as _exc:
                    st.error(f"Update failed: {_exc}")

    if st.session_state.get("ai_sow_update_status"):
        st.success(st.session_state["ai_sow_update_status"])
        st.session_state["ai_sow_update_status"] = ""

    st.divider()

    # ── Client Details ────────────────────────────────────────────────────────
    with st.expander("Client & Project Details", expanded=True):
        c1, c2, c3 = st.columns(3)
        sow["client_name"]    = c1.text_input("Client Name",        value=sow.get("client_name", ""))
        sow["project_name"]   = c2.text_input("Project / Campaign", value=sow.get("project_name", ""))
        sow["version"]        = c3.text_input("Version",            value=sow.get("version", "v1.0"))
        c4, c5, c6 = st.columns(3)
        sow["account_lead"]   = c4.text_input("Account Lead",       value=sow.get("account_lead", ""))
        sow["business_owner"] = c5.text_input("Business Owner",     value=sow.get("business_owner", ""))
        sow["date"]           = c6.text_input("Date",               value=sow.get("date", datetime.today().strftime("%B %d, %Y")))
        c7, c8 = st.columns(2)
        sow["final_deadline"] = c7.text_input("Final Deadline",     value=sow.get("final_deadline", ""))
        sow["budget_range"]   = c8.text_input("Budget Range",       value=sow.get("budget_range", ""))

    # ── Strategic Summary ─────────────────────────────────────────────────────
    # MODIFIED: explicit keys added to why_now, project_overview, core_message
    # so the AI Reword Apply button can force-update these widgets.
    with st.expander("Strategic Summary", expanded=True):
        sow["why_now"] = st.text_area(
            "Why This, Why Now",
            value=sow.get("why_now", ""),
            height=70,
            key="ta_why_now",
            help="The trigger -- what changed or what window exists.",
        )
        sow["project_overview"] = st.text_area(
            "Project Overview",
            value=sow.get("project_overview", ""),
            height=100,
            key="ta_project_overview",
            help="Executive summary of the engagement.",
        )
        sow["core_message"] = st.text_input(
            "Core Message",
            value=sow.get("core_message", ""),
            key="ta_core_message",
            help="One sentence that drives everything.",
        )

    # ── Scope Sections ────────────────────────────────────────────────────────
    st.markdown('<div class="sec-bar">SCOPE OF SERVICES</div>', unsafe_allow_html=True)
    st.caption("Each section becomes a collapsible block in the SOW. Add as many as needed.")

    sections = sow.get("scope_sections", [{"title": "", "description": "", "services": [], "deliverables": []}])
    for i, sec in enumerate(sections):
        label = sec.get("title") or f"Section {i + 1}"
        with st.expander(f"{label}", expanded=(i == 0)):
            sec["title"]       = st.text_input("Section Title", value=sec.get("title", ""),       key=f"st_{i}")
            # key sd_{i} already existed; kept unchanged — Apply targets this key
            sec["description"] = st.text_area("Description",   value=sec.get("description", ""), height=80, key=f"sd_{i}")
            c_svc, c_del = st.columns(2)
            with c_svc:
                svc_raw = st.text_area(
                    "Services Included (one per line)",
                    value="\n".join(sec.get("services", [])),
                    height=120, key=f"ss_{i}",
                )
                sec["services"] = [l.strip() for l in svc_raw.splitlines() if l.strip()]
            with c_del:
                del_raw = st.text_area(
                    "Primary Deliverables (one per line)",
                    value="\n".join(sec.get("deliverables", [])),
                    height=120, key=f"del_{i}",
                )
                sec["deliverables"] = [l.strip() for l in del_raw.splitlines() if l.strip()]
            if st.button(f"Remove Section {i + 1}", key=f"rm_sec_{i}"):
                sections.pop(i)
                st.rerun()
        sections[i] = sec

    if st.button("+ Add Scope Section"):
        sections.append({"title": "", "description": "", "services": [], "deliverables": []})
        st.rerun()
    sow["scope_sections"] = sections

    # ── Assumptions & Out of Scope ────────────────────────────────────────────
    # MODIFIED: explicit keys added so Apply can force-update these widgets
    with st.expander("Assumptions & Out of Scope"):
        c_left, c_right = st.columns(2)
        with c_left:
            assume_raw = st.text_area(
                "Assumptions (one per line)",
                value="\n".join(sow.get("assumptions", [])),
                height=150,
                key="ta_assumptions",
            )
            sow["assumptions"] = [l.strip() for l in assume_raw.splitlines() if l.strip()]
        with c_right:
            oos_raw = st.text_area(
                "Out of Scope (one per line)",
                value="\n".join(sow.get("out_of_scope", [])),
                height=150,
                key="ta_out_of_scope",
            )
            sow["out_of_scope"] = [l.strip() for l in oos_raw.splitlines() if l.strip()]
        sow["timeline_notes"] = st.text_input(
            "Timeline Notes",
            value=sow.get("timeline_notes", ""),
            key="ta_timeline_notes",
        )
        sow["review_rounds"] = st.text_input(
            "Creative Review Rounds",
            value=sow.get("review_rounds", "2"),
        )

    # ════════════════════════════════════════════════════════════════════════════
    # NEW: AI Reword Assistant
    # ════════════════════════════════════════════════════════════════════════════
    with st.expander("AI Reword Assistant", expanded=False):
        st.caption(
            "Select a SOW field, load its content, ask the AI to improve it, "
            "then apply the result directly back into the field."
        )

        # Build field map: label → (sow_field_key, widget_key, field_type)
        # field_type is "text" for strings or "list" for newline-separated lists
        field_map = {
            "Why This, Why Now":   ("why_now",         "ta_why_now",          "text"),
            "Project Overview":    ("project_overview", "ta_project_overview", "text"),
            "Core Message":        ("core_message",     "ta_core_message",     "text"),
            "Assumptions":         ("assumptions",      "ta_assumptions",      "list"),
            "Out of Scope":        ("out_of_scope",     "ta_out_of_scope",     "list"),
            "Timeline Notes":      ("timeline_notes",   "ta_timeline_notes",   "text"),
        }
        for _i, _sec in enumerate(sow.get("scope_sections", [])):
            _title = _sec.get("title") or f"Section {_i + 1}"
            field_map[f"Scope: {_title} — Description"] = (
                f"scope_{_i}_description", f"sd_{_i}", "text"
            )

        ai_r1c1, ai_r1c2 = st.columns([1, 1])
        with ai_r1c1:
            selected_label = st.selectbox(
                "Field to improve",
                list(field_map.keys()),
                key="ai_field_select",
            )
        fkey, wkey, ftype = field_map[selected_label]

        # Resolve current value from the live sow dict (reflects latest edits
        # made in this render pass — Step 2→3 sync source of truth)
        if fkey.startswith("scope_") and fkey.endswith("_description"):
            _idx = int(fkey.split("_")[1])
            _secs = sow.get("scope_sections", [])
            current_val = _secs[_idx].get("description", "") if _idx < len(_secs) else ""
        elif ftype == "list":
            current_val = "\n".join(sow.get(fkey, []))
        else:
            current_val = sow.get(fkey, "")

        with ai_r1c2:
            _presets = [
                "Make this more strategic and client-facing",
                "Make this more concise",
                "Improve clarity and readability",
                "Make this more persuasive",
                "Make this more professional",
                "Simplify for a non-technical audience",
                "Custom instruction…",
            ]
            preset_choice = st.selectbox("Improvement style", _presets, key="ai_preset")

        if preset_choice == "Custom instruction…":
            instruction = st.text_input(
                "Custom instruction",
                key="ai_custom_instr",
                placeholder="e.g. Make this sound more urgent and results-driven",
            )
        else:
            instruction = preset_choice

        # ── Current value display ─────────────────────────────────────────────
        if current_val:
            _preview = current_val[:400] + ("…" if len(current_val) > 400 else "")
            st.markdown(
                f'<div class="ai-box"><strong style="color:#14a4fe;font-size:10px;'
                f'letter-spacing:1px;text-transform:uppercase;">Current field value</strong>'
                f'<br><br>{_preview}</div>',
                unsafe_allow_html=True,
            )

        # ── Load & edit input ─────────────────────────────────────────────────
        _load_col, _ = st.columns([1, 3])
        with _load_col:
            if st.button("Load field into editor", key="ai_load_btn", use_container_width=True):
                st.session_state["ai_input_text"] = current_val
                st.rerun()

        input_text = st.text_area(
            "Text to improve (editable)",
            key="ai_input_text",
            height=110,
            placeholder="Click 'Load field into editor' above, or paste any text you want to improve.",
        )

        # ── Reword button ─────────────────────────────────────────────────────
        _rw_col, _ = st.columns([1, 2])
        with _rw_col:
            if st.button("Reword", type="primary", use_container_width=True, key="btn_reword"):
                if not input_text.strip():
                    st.warning("Please load a field or paste some text first.")
                elif not instruction.strip():
                    st.warning("Please choose or enter an instruction.")
                elif not st.session_state.get("api_key"):
                    st.error("API key not configured. Contact your Adchor admin.")
                else:
                    with st.spinner("Refining your copy…"):
                        try:
                            _result = _ai_reword(
                                input_text,
                                instruction,
                                st.session_state.api_key,
                            )
                            st.session_state.ai_reword_result     = _result
                            st.session_state.ai_reword_target     = fkey
                            st.session_state.ai_reword_widget_key = wkey
                            st.session_state.ai_reword_field_type = ftype
                            # Force-update result display widget on next render
                            st.session_state["ai_result_area"]    = _result
                            st.rerun()
                        except Exception as _exc:
                            st.error(f"Reword failed: {_exc}")

        # ── Result + Apply ────────────────────────────────────────────────────
        if st.session_state.get("ai_reword_result"):
            st.markdown(
                "<span style='font-size:11px;font-weight:700;color:#00ff79;"
                "letter-spacing:1px;text-transform:uppercase;'>✓ AI suggestion</span>"
                " <span style='color:#5a6278;font-size:11px;'>(editable before applying)</span>",
                unsafe_allow_html=True,
            )
            edited_result = st.text_area(
                "AI suggestion",
                height=130,
                key="ai_result_area",
                label_visibility="collapsed",
            )

            _ap1, _ap2 = st.columns([1, 1])
            _t_fkey  = st.session_state.get("ai_reword_target",     fkey)
            _t_wkey  = st.session_state.get("ai_reword_widget_key", wkey)
            _t_ftype = st.session_state.get("ai_reword_field_type", ftype)

            with _ap1:
                if st.button("← Apply to field", type="primary", use_container_width=True, key="btn_apply"):
                    # 1. Update sow_data (persistent source of truth)
                    if _t_fkey.startswith("scope_") and _t_fkey.endswith("_description"):
                        _si = int(_t_fkey.split("_")[1])
                        _sc = sow.get("scope_sections", [])
                        if _si < len(_sc):
                            _sc[_si]["description"] = edited_result
                            sow["scope_sections"] = _sc
                    elif _t_ftype == "list":
                        sow[_t_fkey] = [l.strip() for l in edited_result.splitlines() if l.strip()]
                    else:
                        sow[_t_fkey] = edited_result

                    st.session_state.sow_data         = sow
                    st.session_state.ai_reword_result = ""
                    # 2. Defer widget-key update to top of next rerun (avoids
                    #    StreamlitAPIException: can't set already-rendered widget)
                    if _t_wkey:
                        st.session_state["_pending_widget_update"] = {_t_wkey: edited_result}
                    st.success(f"✓ Applied to '{selected_label}'! Scroll up to review.")
                    st.rerun()

            with _ap2:
                if st.button("✕ Discard", use_container_width=True, key="btn_discard"):
                    st.session_state.ai_reword_result = ""
                    st.rerun()

    # ── Save & navigate ───────────────────────────────────────────────────────
    # sow_data is the single source of truth for Step 3 pricing sync.
    # Saving here ensures any manual edit made in this render is persisted
    # before Step 3 reads it.
    st.session_state.sow_data = sow

    st.divider()
    col_back, col_save, col_next = st.columns([1, 1, 1])
    with col_back:
        if st.button("← Back", use_container_width=True):
            st.session_state.step = 1
            st.rerun()
    with col_save:
        if st.button("Save SOW", use_container_width=True):
            upsert_sow_to_library(
                sow,
                st.session_state.get("pricing_items", []),
                st.session_state.get("sow_discount", 0),
            )
            st.success("✓ Saved to SOW Library!")
    with col_next:
        if st.button("Pricing →", use_container_width=True, type="primary"):
            # Auto-save when advancing to pricing
            upsert_sow_to_library(
                sow,
                st.session_state.get("pricing_items", []),
                st.session_state.get("sow_discount", 0),
            )
            st.session_state.step = 3
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 -- PRICING BUILDER  (unchanged except pdf_downloaded reset on enter)
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.step == 3:
    st.markdown("""
    <div class="hero-bar">
        <div class="hero-eyebrow">Step 3 of 4</div>
        <div class="hero-title">Investment &amp; Pricing</div>
        <div class="hero-sub">Price the services from your SOW &mdash; click any service to add it, then set price and quantity</div>
    </div>
    """, unsafe_allow_html=True)

    library   = st.session_state.pricing_library
    lib_items = library.get("items", [])
    items     = st.session_state.pricing_items
    # Step 3 always reads from st.session_state.sow_data — the single source of
    # truth written at the end of every Step 2 render — so any edit made in
    # Step 2 is automatically reflected here without extra wiring.
    sow       = st.session_state.sow_data

    # ── Section A: Services from your SOW ─────────────────────────────────────
    scope_sections = sow.get("scope_sections", [])
    if scope_sections:
        st.markdown("##### Services from your SOW")
        st.caption("Click a service to add it as a line item below.")
        sow_cols = st.columns(3)
        for i, sec in enumerate(scope_sections):
            title = sec.get("title", "").strip()
            desc  = sec.get("description", "")[:80] + "…" if len(sec.get("description","")) > 80 else sec.get("description","")
            if not title:
                continue
            with sow_cols[i % 3]:
                st.markdown(f"""
                <div style="background:#1a1d24;border:1px solid #2a2d3a;border-left:4px solid #014bf7;
                    border-radius:8px;padding:12px 14px;margin-bottom:8px;">
                    <div style="color:white;font-weight:700;font-size:13px;">{title}</div>
                    <div style="color:#9aa0b0;font-size:11px;margin-top:4px;">{desc}</div>
                </div>""", unsafe_allow_html=True)
                if st.button(f"+ Add to pricing", key=f"sow_add_{i}", use_container_width=True):
                    items.append({
                        "name":        title,
                        "description": sec.get("description","")[:100],
                        "category":    "SOW Service",
                        "unit_price":  0,
                        "qty":         1,
                        "total":       0,
                    })
                    st.session_state.pricing_items = items
                    st.rerun()
        st.divider()

    # ── Section B: Pricing Table ───────────────────────────────────────────────
    st.markdown("##### Line Items")
    if not items:
        st.info("No items yet -- click a service above or use '+ Add Row' below.")
    else:
        h1, h2, h3, h4, h5, h6 = st.columns([3, 3, 1, 1.5, 1.5, 0.4])
        for col, hdr in zip([h1,h2,h3,h4,h5,h6], ["Service","Description","Qty","Unit Price ($)","Total",""]):
            col.markdown(f"<span style='font-size:11px;font-weight:700;color:#9aa0b0;text-transform:uppercase;'>{hdr}</span>",
                         unsafe_allow_html=True)
        st.markdown("<hr style='margin:4px 0 8px;border-color:#2a2d3a;'>", unsafe_allow_html=True)

        to_remove = []
        for i, item in enumerate(items):
            c1, c2, c3, c4, c5, c6 = st.columns([3, 3, 1, 1.5, 1.5, 0.4])
            item["name"]        = c1.text_input("", value=item.get("name",""),        key=f"n_{i}", label_visibility="collapsed", placeholder="Service name")
            item["description"] = c2.text_input("", value=item.get("description",""), key=f"d_{i}", label_visibility="collapsed", placeholder="Brief description")
            qty_raw = c3.text_input("", value=str(item.get("qty", 1)), key=f"q_{i}", label_visibility="collapsed")
            try:
                item["qty"] = max(1, int(qty_raw))
            except ValueError:
                item["qty"] = 1
            price_raw = c4.text_input("", value=str(item.get("unit_price", "")), key=f"p_{i}",
                                      label_visibility="collapsed", placeholder="0")
            try:
                item["unit_price"] = float(str(price_raw).replace(",","").replace("$","").strip()) if price_raw else 0.0
            except ValueError:
                item["unit_price"] = 0.0
            item["total"] = item["qty"] * item["unit_price"]
            c5.markdown(f"<div style='padding-top:8px;font-weight:700;color:white;'>${item['total']:,.0f}</div>",
                        unsafe_allow_html=True)
            if c6.button("✕", key=f"del_{i}"):
                to_remove.append(i)

        for idx in sorted(to_remove, reverse=True):
            items.pop(idx)
        if to_remove:
            st.session_state.pricing_items = items
            st.rerun()
        st.session_state.pricing_items = items

    if st.button("＋ Add Row", use_container_width=False):
        items.append({"name":"","description":"","category":"","unit_price":0,"qty":1,"total":0})
        st.session_state.pricing_items = items
        st.rerun()

    st.divider()

    # ── Section C: Totals ──────────────────────────────────────────────────────
    subtotal = sum(i.get("total", 0) for i in items)
    tc1, tc2 = st.columns([3, 1])
    with tc2:
        disc_raw = st.text_input("Discount (%)", value=str(int(st.session_state.sow_discount)) if st.session_state.sow_discount else "", placeholder="0", key="disc_input")
        try:
            discount_pct = float(str(disc_raw).replace(",","").replace("%","").strip()) if disc_raw else 0.0
            discount_pct = max(0.0, min(100.0, discount_pct))
        except ValueError:
            discount_pct = 0.0
        discount_amt = subtotal * discount_pct / 100
        st.session_state.sow_discount = discount_pct
        final_total = subtotal - discount_amt
        st.session_state.sow_total = final_total
        disc_line = f'<div class="sub">Discount ({discount_pct:.0f}%): -${discount_amt:,.0f}</div>' if discount_pct else ""
        st.markdown(f"""
        <div class="pricing-total">
            <div class="sub">Subtotal: ${subtotal:,.0f}</div>
            {disc_line}
            Total Investment: ${final_total:,.0f}
        </div>""", unsafe_allow_html=True)

    # ── Section D: Library (collapsed) ────────────────────────────────────────
    with st.expander("Add from Service Library or save new item"):
        lib_col1, lib_col2 = st.columns(2)
        with lib_col1:
            st.markdown("**Quick-add from library**")
            if lib_items:
                lib_labels = [it["name"] for it in lib_items]
                sel = st.selectbox("Pick a service", ["-- Select --"] + lib_labels, key="lib_sel", label_visibility="collapsed")
                if sel != "-- Select --" and st.button("Add to pricing", use_container_width=True):
                    idx = lib_labels.index(sel)
                    new = lib_items[idx].copy()
                    new["qty"] = 1
                    new["total"] = new["unit_price"]
                    items.append(new)
                    st.session_state.pricing_items = items
                    st.rerun()
            else:
                st.info("Library is empty.")
        with lib_col2:
            st.markdown("**Save a new item to library**")
            save_name  = st.text_input("Service Name", key="save_name")
            save_desc  = st.text_input("Description",  key="save_desc")
            save_cat   = st.text_input("Category",     key="save_cat")
            save_price_raw = st.text_input("Default Price ($)", key="save_price", placeholder="0")
            try:
                save_price = float(save_price_raw.replace(",","").replace("$","").strip()) if save_price_raw else 0.0
            except ValueError:
                save_price = 0.0
            if st.button("Save to Library", use_container_width=True):
                if save_name.strip():
                    entry = {"name": save_name.strip(), "description": save_desc.strip(),
                             "category": save_cat.strip(), "unit_price": save_price}
                    if not any(it["name"] == entry["name"] for it in lib_items):
                        lib_items.append(entry)
                        st.session_state.pricing_library["items"] = lib_items
                        save_library(st.session_state.pricing_library)
                        st.success(f"✓ '{save_name}' saved to pricing library.")
                    else:
                        st.warning(f"'{save_name}' already exists in the library.")
        st.divider()
        st.caption("Download pricing library:")
        _dl1, _dl2, _dl3 = st.columns(3)

        # ── JSON download ─────────────────────────────────────────────────────
        with _dl1:
            st.download_button(
                "JSON",
                data=json.dumps(st.session_state.pricing_library, indent=2),
                file_name="pricing_library.json",
                mime="application/json",
                use_container_width=True,
            )

        # ── Excel download ────────────────────────────────────────────────────
        with _dl2:
            try:
                import openpyxl, io as _io2
                _wb = openpyxl.Workbook()
                _ws = _wb.active
                _ws.title = "Pricing Library"
                _ws.append(["Name", "Description", "Category", "Unit Price"])
                for _it in lib_items:
                    _ws.append([
                        _it.get("name", ""),
                        _it.get("description", ""),
                        _it.get("category", ""),
                        float(_it.get("unit_price", 0)),
                    ])
                _xl_buf = _io2.BytesIO()
                _wb.save(_xl_buf)
                _xl_buf.seek(0)
                st.download_button(
                    "Excel",
                    data=_xl_buf.getvalue(),
                    file_name="pricing_library.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception:
                st.caption("Excel unavailable")

        # ── PDF download ──────────────────────────────────────────────────────
        with _dl3:
            try:
                import io as _io3
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib import colors as _rl_colors
                from reportlab.lib.units import inch
                _pdf_buf = _io3.BytesIO()
                _doc = SimpleDocTemplate(_pdf_buf, pagesize=(8.5*inch, 11*inch),
                                         leftMargin=0.5*inch, rightMargin=0.5*inch,
                                         topMargin=0.6*inch, bottomMargin=0.5*inch)
                _styles = getSampleStyleSheet()
                _table_data = [["Name", "Category", "Unit Price"]]
                for _it in lib_items:
                    _table_data.append([
                        _it.get("name", ""),
                        _it.get("category", ""),
                        f"${float(_it.get('unit_price', 0)):,.0f}",
                    ])
                _tbl = Table(_table_data, colWidths=[3.5*inch, 1.5*inch, 1.2*inch])
                _tbl.setStyle(TableStyle([
                    ("BACKGROUND",  (0,0), (-1,0),  colors.HexColor("#014bf7")),
                    ("TEXTCOLOR",   (0,0), (-1,0),  _rl_colors.white),
                    ("FONTNAME",    (0,0), (-1,0),  "Helvetica-Bold"),
                    ("FONTSIZE",    (0,0), (-1,-1), 9),
                    ("ROWBACKGROUNDS", (0,1), (-1,-1), [_rl_colors.white, _rl_colors.HexColor("#f4f6fc")]),
                    ("GRID",        (0,0), (-1,-1), 0.5, _rl_colors.HexColor("#e4e8f4")),
                    ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
                    ("TOPPADDING",  (0,0), (-1,-1), 5),
                    ("BOTTOMPADDING",(0,0),(-1,-1), 5),
                ]))
                _doc.build([
                    Paragraph("Adchor Pricing Library", _styles["Title"]),
                    Spacer(1, 0.2*inch),
                    _tbl,
                ])
                _pdf_buf.seek(0)
                st.download_button(
                    "PDF",
                    data=_pdf_buf.getvalue(),
                    file_name="pricing_library.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception:
                st.caption("PDF unavailable")

        # ── Upload / restore ──────────────────────────────────────────────────
        st.caption("Restore or import pricing items (JSON, Excel, or PDF rate card):")
        _pl_up = st.file_uploader(
            "Upload pricing library or rate card",
            type=["json", "xlsx", "pdf"],
            key="pricing_lib_upload",
        )
        if _pl_up:
            _ext = _pl_up.name.rsplit(".", 1)[-1].lower()
            if _ext == "json":
                try:
                    _pl_restored = json.load(_pl_up)
                    if "items" in _pl_restored:
                        st.session_state.pricing_library = _pl_restored
                        save_library(_pl_restored)
                        st.success(f"✓ Restored {len(_pl_restored['items'])} pricing items.")
                        st.rerun()
                    else:
                        st.error("Invalid pricing library JSON file.")
                except Exception:
                    st.error("Could not read JSON file.")

            elif _ext == "xlsx":
                try:
                    import openpyxl, io as _io4
                    _wb2 = openpyxl.load_workbook(_io4.BytesIO(_pl_up.read()))
                    _ws2 = _wb2.active
                    _new_items = []
                    for _row in _ws2.iter_rows(min_row=2, values_only=True):
                        _nm = str(_row[0] or "").strip()
                        if not _nm:
                            continue
                        _new_items.append({
                            "name":        _nm,
                            "description": str(_row[1] or "").strip(),
                            "category":    str(_row[2] or "").strip(),
                            "unit_price":  float(_row[3] or 0),
                        })
                    if _new_items:
                        _pl_new = {"items": _new_items}
                        st.session_state.pricing_library = _pl_new
                        save_library(_pl_new)
                        st.success(f"✓ Imported {len(_new_items)} items from Excel.")
                        st.rerun()
                    else:
                        st.error("No items found in Excel file. Ensure headers are: Name, Description, Category, Unit Price.")
                except Exception as _xe:
                    st.error(f"Could not read Excel file: {_xe}")

            elif _ext == "pdf":
                if not st.session_state.get("api_key"):
                    st.error("API key not configured — cannot extract pricing from PDF.")
                else:
                    with st.spinner("Extracting pricing items from PDF…"):
                        try:
                            from pypdf import PdfReader
                            import io as _io5
                            _pr = PdfReader(_io5.BytesIO(_pl_up.read()))
                            _ptxt = "\n".join(p.extract_text() or "" for p in _pr.pages).strip()
                            if not _ptxt:
                                st.error("Could not extract text from this PDF.")
                            else:
                                import anthropic as _ant2
                                _cl2 = _ant2.Anthropic(api_key=st.session_state.api_key)
                                _m2  = _cl2.messages.create(
                                    model="claude-sonnet-4-6",
                                    max_tokens=2048,
                                    messages=[{"role": "user", "content": (
                                        "Extract all pricing / service items from this rate card or pricing document.\n\n"
                                        f"Document text:\n{_ptxt[:6000]}\n\n"
                                        "Return a JSON array of objects with keys: "
                                        "name, description, category, unit_price (number).\n"
                                        "Return raw JSON array only — no markdown, no code fences."
                                    )}],
                                )
                                _raw = _m2.content[0].text.strip()
                                if _raw.startswith("```"):
                                    _lns = _raw.splitlines()
                                    _raw = "\n".join(_lns[1:-1] if _lns[-1].strip() == "```" else _lns[1:]).strip()
                                _pit = json.loads(_raw)
                                if isinstance(_pit, list) and _pit:
                                    _pl_pdf = {"items": _pit}
                                    st.session_state.pricing_library = _pl_pdf
                                    save_library(_pl_pdf)
                                    st.success(f"✓ Extracted {len(_pit)} pricing items from PDF.")
                                    st.rerun()
                                else:
                                    st.error("No pricing items found in this PDF.")
                        except Exception as _pe:
                            st.error(f"PDF extraction failed: {_pe}")

    st.divider()
    col_back, col_save3, col_next = st.columns([1, 1, 1])
    with col_back:
        if st.button("← Back", use_container_width=True):
            st.session_state.step = 2
            st.rerun()
    with col_save3:
        if st.button("Save SOW", use_container_width=True, key="save_sow_s3"):
            upsert_sow_to_library(
                st.session_state.sow_data or {},
                st.session_state.get("pricing_items", []),
                st.session_state.get("sow_discount", 0),
            )
            st.success("✓ Saved to SOW Library!")
    with col_next:
        # Auto-save when advancing to PDF
        if st.button("Generate PDF →", use_container_width=True, type="primary"):
            _pi = st.session_state.get("pricing_items", [])
            upsert_sow_to_library(
                st.session_state.sow_data or {},
                _pi,
                st.session_state.get("sow_discount", 0),
            )
            # Auto-add all quoted line items to the pricing library
            autosave_pricing_items_to_library(_pi)
            st.session_state.pdf_downloaded = False
            st.session_state.step = 4
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 -- DOWNLOAD
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.step == 4:
    st.markdown("""
    <div class="hero-bar">
        <div class="hero-eyebrow">Step 4 of 4</div>
        <div class="hero-title">Download &amp; Send</div>
        <div class="hero-sub">Generate your branded PDF and deliver via Adobe Sign</div>
    </div>
    """, unsafe_allow_html=True)

    sow     = st.session_state.sow_data or {}
    client  = sow.get("client_name", "Client")
    project = sow.get("project_name", "Project")
    total        = st.session_state.sow_total
    discount_pct = st.session_state.sow_discount
    subtotal_s4  = sum(i.get("total", 0) for i in st.session_state.pricing_items)
    discount_amt = subtotal_s4 * discount_pct / 100

    st.success(f"✓  SOW ready: **{client} -- {project}**  ·  Investment: **${total:,.0f}**")

    col_pdf, col_steps = st.columns([1, 1])

    with col_pdf:
        st.markdown("#### Download PDF")
        st.markdown("Branded PDF ready for Adobe Sign.")
        with st.spinner("Building PDF..."):
            try:
                pdf_bytes = build_sow_pdf(
                    sow_data=sow,
                    pricing_items=st.session_state.pricing_items,
                    total=total,
                    discount=discount_amt,
                )
                filename = f"{client}_{project}_SOW.pdf".replace(" ", "_").replace("/", "-")

                # MODIFIED: capture the return value of download_button.
                # It returns True on the rerun triggered by the click,
                # which is the reliable signal that the file was sent to the browser.
                st.markdown("""
                <style>
                div[data-testid="stDownloadButton"][data-key="main_pdf_dl"] button {
                    background: linear-gradient(135deg,#014bf7,#021de0) !important;
                    color: #ffffff !important;
                    border: none !important;
                    font-weight: 700 !important;
                }
                div[data-testid="stDownloadButton"][data-key="main_pdf_dl"] button:hover {
                    background: linear-gradient(135deg,#0255ff,#0330f5) !important;
                    box-shadow: 0 4px 18px rgba(1,75,247,0.55) !important;
                }
                </style>
                """, unsafe_allow_html=True)
                _downloaded = st.download_button(
                    label=f"Download {filename}",
                    data=pdf_bytes,
                    file_name=filename,
                    mime="application/pdf",
                    use_container_width=True,
                    key="main_pdf_dl",
                )
                if _downloaded:
                    st.session_state.pdf_downloaded = True
                    # Auto-save final SOW to library with "final" status
                    _final_entry = upsert_sow_to_library(
                        sow,
                        st.session_state.pricing_items,
                        st.session_state.sow_discount,
                    )
                    # Mark as final
                    _lib = st.session_state.get("sow_library", {"sows": []})
                    _client_n  = (sow or {}).get("client_name", "").strip()
                    _project_n = (sow or {}).get("project_name", "").strip()
                    for _s in _lib.get("sows", []):
                        if _s.get("client_name") == _client_n and _s.get("project_name") == _project_n:
                            _s["status"] = "final"
                    save_sow_library(_lib)

            except Exception as e:
                st.error(f"PDF generation error: {e}")

        # Download status indicator
        if st.session_state.get("pdf_downloaded"):
            st.markdown(
                "<div style='margin-top:8px;padding:8px 14px;background:rgba(0,255,121,0.07);"
                "border:1px solid rgba(0,255,121,0.25);border-radius:8px;font-size:12px;"
                "color:#00ff79;font-weight:600;'>✓ PDF downloaded</div>",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                "<div style='margin-top:8px;padding:8px 14px;background:rgba(255,180,0,0.07);"
                "border:1px solid rgba(255,180,0,0.25);border-radius:8px;font-size:12px;"
                "color:#ffb400;'>Click the button above to download the PDF before finishing.</div>",
                unsafe_allow_html=True,
            )

    with col_steps:
        st.markdown("#### Next Steps")
        st.markdown("""
        <div class="dl-note">
        <strong>After downloading:</strong><br><br>
        1. Open <strong>Adobe Sign</strong><br>
        2. Upload the PDF and add signature fields for the client and your account lead<br>
        3. Send for e-signature<br>
        4. Once fully signed, attach to the <strong>Monday.com</strong> project card and move to Active<br><br>
        <em>Tip: Keep the PDF file name as-is -- it follows the Client_Project_SOW convention.</em>
        </div>
        """, unsafe_allow_html=True)

    st.divider()

    # MODIFIED: two-button row — Back on the left, Done on the right.
    col_back, _, col_done = st.columns([1, 3, 1])

    with col_back:
        if st.button("← Back to Pricing", use_container_width=True):
            # Going back resets the download flag — new pricing requires a new download.
            st.session_state.pdf_downloaded = False
            st.session_state.step = 3
            st.rerun()

    with col_done:
        # NEW: Done button — only completes the workflow after the PDF is downloaded.
        if st.button("✓ Done", use_container_width=True, type="primary"):
            if st.session_state.get("pdf_downloaded", False):
                _reset_sow_state()
                st.rerun()
            else:
                st.warning("Please make sure to download the PDF before finishing the SOW.")
